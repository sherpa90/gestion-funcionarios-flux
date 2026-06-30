from django.views.generic import TemplateView, View
from django.shortcuts import redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from permisos.models import SolicitudPermiso
from licencias.models import LicenciaMedica
from asistencia.models import RegistroAsistencia, HorarioFuncionario, DiaHorario, AlegacionAsistencia
from users.models import CustomUser
from core.services import BusinessDayCalculator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, time, date
from django.utils.timezone import now
from .calc_utils import calcular_inasistencias_reales

# Mapeo día de semana (Python weekday 0=Lunes) → nombre en español
DIA_SEMANA_MAP = {
    0: 'Lunes', 1: 'Martes', 2: 'Miércoles',
    3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'
}

class ReportesView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista unificada y minimalista de reportes"""
    template_name = 'reportes/reportes.html'

    def test_func(self):
        # Acceso para Director, Secretaria, Admin y Directivos
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener parámetros de filtro
        search = self.request.GET.get('search', '')
        year = self.request.GET.get('year', '')
        mes = self.request.GET.get('mes', '')
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')
        sort_by = self.request.GET.get('sort', 'name')
        
        # Base queryset: incluir todos los funcionarios del sistema
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'])
        
        # Filtro de búsqueda por nombre o RUN
        if search:
            funcionarios = funcionarios.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(run__icontains=search)
            )
        
        # Preparar datos de cada funcionario
        empleados_data = []
        for functorio in funcionarios:
            # Obtener permisos aprobados (sin filtros para cálculo cronológico completo)
            todos_permisos = SolicitudPermiso.objects.filter(
                usuario=functorio,
                estado='APROBADO'
            ).order_by('fecha_inicio', 'created_at')
            
            # Permisos con filtro para mostrar en el reporte
            permisos = todos_permisos
            if year:
                permisos = permisos.filter(fecha_inicio__year=year)
            if mes:
                permisos = permisos.filter(fecha_inicio__month=mes)
            if fecha_inicio:
                permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
            
            # Calcular días disponibles cronológicamente por cada permiso
            BASE_DIAS_ADMINISTRATIVOS = 6.0
            permisos_con_dias = []
            dias_acumulados = 0.0
            
            for p in todos_permisos:
                dias_acumulados += float(p.dias_solicitados)
                dias_restantes = max(BASE_DIAS_ADMINISTRATIVOS - dias_acumulados, 0)
                # Solo incluir si pasa los filtros
                include = True
                if year and p.fecha_inicio.year != int(year):
                    include = False
                if mes and p.fecha_inicio.month != int(mes):
                    include = False
                if fecha_inicio and p.fecha_inicio < datetime.strptime(fecha_inicio, '%Y-%m-%d').date():
                    include = False
                if fecha_fin and p.fecha_inicio > datetime.strptime(fecha_fin, '%Y-%m-%d').date():
                    include = False
                
                if include:
                    permisos_con_dias.append({
                        'permiso': p,
                        'dias_disponibles': dias_restantes,
                    })
            
            dias_usados = sum(p['permiso'].dias_solicitados for p in permisos_con_dias)
            dias_disponibles_calculados = permisos_con_dias[-1]['dias_disponibles'] if permisos_con_dias else BASE_DIAS_ADMINISTRATIVOS
            
            # Obtener licencias médicas
            licencias = LicenciaMedica.objects.filter(usuario=functorio)
            
            if year:
                licencias = licencias.filter(fecha_inicio__year=year)
            if mes:
                licencias = licencias.filter(fecha_inicio__month=mes)
            if fecha_inicio:
                licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                licencias = licencias.filter(fecha_inicio__lte=fecha_fin)
            
            total_licencias = licencias.count()
            dias_licencias = licencias.aggregate(Sum('dias'))['dias__sum'] or 0

            # Obtener registros de asistencia para contar atrasos e inasistencias
            registros_asistencia = RegistroAsistencia.objects.filter(funcionario=functorio)
            if year:
                registros_asistencia = registros_asistencia.filter(fecha__year=year)
            if mes:
                registros_asistencia = registros_asistencia.filter(fecha__month=mes)
            if fecha_inicio:
                registros_asistencia = registros_asistencia.filter(fecha__gte=fecha_inicio)
            if fecha_fin:
                registros_asistencia = registros_asistencia.filter(fecha__lte=fecha_fin)

            es_sereno_func = (
                getattr(functorio, 'funcion', None) == 'SERENO' or
                getattr(functorio, 'tipo_funcionario', None) == 'SERENO'
            )
            if es_sereno_func:
                total_atrasos = 0
                total_inasistencias = 0
                total_minutos_retraso = 0
            else:
                # Determinar rango de fechas para el cálculo de inasistencias reales
                f_inicio_calc = None
                f_fin_calc = None
                
                # Convertir fechas string a date si es necesario
                def parse_date(d_val):
                    if isinstance(d_val, str) and d_val:
                        return datetime.strptime(d_val, '%Y-%m-%d').date()
                    elif isinstance(d_val, datetime):
                        return d_val.date()
                    return d_val

                if fecha_inicio:
                    f_inicio_calc = parse_date(fecha_inicio)
                elif year:
                    y_int = int(year)
                    if mes:
                        m_int = int(mes)
                        f_inicio_calc = date(y_int, m_int, 1)
                    else:
                        f_inicio_calc = date(y_int, 1, 1)

                if fecha_fin:
                    f_fin_calc = parse_date(fecha_fin)
                elif year:
                    y_int = int(year)
                    if mes:
                        m_int = int(mes)
                        import calendar
                        _, last_day = calendar.monthrange(y_int, m_int)
                        f_fin_calc = date(y_int, m_int, last_day)
                    else:
                        f_fin_calc = date(y_int, 12, 31)

                total_atrasos = registros_asistencia.filter(estado='RETRASO').count()
                total_inasistencias = calcular_inasistencias_reales(functorio, f_inicio_calc, f_fin_calc)
                total_minutos_retraso = registros_asistencia.filter(estado='RETRASO').aggregate(
                    total=Sum('minutos_retraso'))['total'] or 0
            
            empleados_data.append({
                'funcionario': functorio,
                'cargo': functorio.get_funcion_display() or functorio.get_tipo_funcionario_display() or functorio.get_role_display(),
                # Días disponibles calculados cronológicamente hasta la fecha del filtro
                'dias_disponibles': dias_disponibles_calculados,
                'dias_usados': dias_usados,
                'total_licencias': total_licencias,
                'dias_licencias': dias_licencias,
                'permisos': [p['permiso'] for p in permisos_con_dias],
                'permisos_con_dias': permisos_con_dias,
                'licencias': licencias.order_by('fecha_inicio'),
                'total_atrasos': total_atrasos,
                'total_inasistencias': total_inasistencias,
                'total_minutos_retraso': total_minutos_retraso,
            })
        
        # Aplicar ordenamiento
        if sort_by == 'name':
            empleados_data.sort(key=lambda x: (x['funcionario'].first_name, x['funcionario'].last_name))
        elif sort_by == 'name_desc':
            empleados_data.sort(key=lambda x: (x['funcionario'].first_name, x['funcionario'].last_name), reverse=True)
        elif sort_by == 'dias':
            empleados_data.sort(key=lambda x: x['dias_disponibles'], reverse=True)
        elif sort_by == 'dias_asc':
            empleados_data.sort(key=lambda x: x['dias_disponibles'])
        elif sort_by == 'dias_usados':
            empleados_data.sort(key=lambda x: x['dias_usados'], reverse=True)
        elif sort_by == 'dias_usados_asc':
            empleados_data.sort(key=lambda x: x['dias_usados'])
        elif sort_by == 'licencias':
            empleados_data.sort(key=lambda x: x['total_licencias'], reverse=True)
        elif sort_by == 'licencias_asc':
            empleados_data.sort(key=lambda x: x['total_licencias'])
        elif sort_by == 'dias_licencias':
            empleados_data.sort(key=lambda x: x['dias_licencias'], reverse=True)
        elif sort_by == 'dias_licencias_asc':
            empleados_data.sort(key=lambda x: x['dias_licencias'])
        elif sort_by == 'inasistencias':
            empleados_data.sort(key=lambda x: x['total_inasistencias'], reverse=True)
        elif sort_by == 'inasistencias_asc':
            empleados_data.sort(key=lambda x: x['total_inasistencias'])
        elif sort_by == 'atrasos':
            empleados_data.sort(key=lambda x: x['total_atrasos'], reverse=True)
        elif sort_by == 'atrasos_asc':
            empleados_data.sort(key=lambda x: x['total_atrasos'])
        
        context['empleados_data'] = empleados_data
        context['filtros'] = {
            'search': search,
            'year': year,
            'mes': mes,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
        
        # Años disponibles para filtro
        from datetime import datetime
        permisos_years = set(SolicitudPermiso.objects.dates('fecha_inicio', 'year').values_list('fecha_inicio', flat=True))
        licencias_years = set(LicenciaMedica.objects.dates('fecha_inicio', 'year').values_list('fecha_inicio', flat=True))
        all_years = sorted(set([d.year for d in permisos_years] + [d.year for d in licencias_years]), reverse=True)
        context['years'] = all_years if all_years else [datetime.now().year]
        context['current_sort'] = sort_by
        
        # --- Estadísticas para gráficos ---
        stats_year = int(year) if year and year.isdigit() else now().year
        
        # Permisos por mes
        permisos_mes = SolicitudPermiso.objects.filter(
            estado='APROBADO',
            fecha_inicio__year=stats_year
        ).values('fecha_inicio__month').annotate(total=Sum('dias_solicitados')).order_by('fecha_inicio__month')
        
        permisos_data = [0] * 12
        for p in permisos_mes:
            if p['fecha_inicio__month']:
                permisos_data[p['fecha_inicio__month'] - 1] = float(p['total'] or 0)
        
        # Licencias por mes
        licencias_mes = LicenciaMedica.objects.filter(
            fecha_inicio__year=stats_year
        ).values('fecha_inicio__month').annotate(total=Sum('dias')).order_by('fecha_inicio__month')
        
        licencias_data = [0] * 12
        for l in licencias_mes:
            if l['fecha_inicio__month']:
                licencias_data[l['fecha_inicio__month'] - 1] = int(l['total'] or 0)
        
# Atrasos por mes (minutos acumulados) - excluir serenos
        atrasos_mes = RegistroAsistencia.objects.filter(
            estado='RETRASO',
            fecha__year=stats_year
        ).exclude(
            Q(funcionario__funcion='SERENO') | Q(funcionario__tipo_funcionario='SERENO')
        ).values('fecha__month').annotate(total=Sum('minutos_retraso')).order_by('fecha__month')

        atrasos_data = [0] * 12
        for a in atrasos_mes:
            if a['fecha__month']:
                atrasos_data[a['fecha__month'] - 1] = int(a['total'] or 0)

        context['stats'] = {
            'year': stats_year,
            'permisos_mensuales': permisos_data,
            'licencias_mensuales': licencias_data,
            'atrasos_mensuales': atrasos_data,
            'total_permisos_anual': sum(permisos_data),
            'total_licencias_anual': sum(licencias_data),
            'total_atrasos_anual': sum(atrasos_data),
        }
        
        return context


class PDFIndividualView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generar Pdf de un solo empleado"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request, usuario_id):
        try:
            functorio = CustomUser.objects.get(pk=usuario_id)
        except CustomUser.DoesNotExist:
            return HttpResponse("Funcionario no encontrado", status=404)
        
        # Obtener parámetros de filtro
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        
        # Obtener todos los permisos aprobados del usuario (para cálculo cronológico)
        todos_permisos = SolicitudPermiso.objects.filter(
            usuario=functorio,
            estado='APROBADO'
        ).order_by('fecha_inicio', 'created_at')
        
        # Permisos con filtro para mostrar en el reporte
        permisos = todos_permisos
        if year:
            permisos = permisos.filter(fecha_inicio__year=year)
        if mes:
            permisos = permisos.filter(fecha_inicio__month=mes)
        if fecha_inicio:
            permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
        if fecha_fin:
            permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
        
        # Calcular días disponibles cronológicamente por cada permiso
        BASE_DIAS_ADMINISTRATIVOS = 6.0
        permisos_con_dias = []
        dias_acumulados = 0.0
        
        for p in todos_permisos:
            dias_acumulados += float(p.dias_solicitados)
            dias_restantes = max(BASE_DIAS_ADMINISTRATIVOS - dias_acumulados, 0)
            # Solo incluir si pasa los filtros
            include = True
            if year and p.fecha_inicio.year != int(year):
                include = False
            if mes and p.fecha_inicio.month != int(mes):
                include = False
            if fecha_inicio and p.fecha_inicio < datetime.strptime(fecha_inicio, '%Y-%m-%d').date():
                include = False
            if fecha_fin and p.fecha_inicio > datetime.strptime(fecha_fin, '%Y-%m-%d').date():
                include = False
            
            if include:
                permisos_con_dias.append({
                    'permiso': p,
                    'dias_disponibles': dias_restantes,
                })
        
        dias_usados = sum(p['permiso'].dias_solicitados for p in permisos_con_dias)
        
        # Obtener licencias
        licencias = LicenciaMedica.objects.filter(usuario=functorio).order_by('fecha_inicio')
        
        if year:
            licencias = licencias.filter(fecha_inicio__year=year)
        if mes:
            licencias = licencias.filter(fecha_inicio__month=mes)
        if fecha_inicio:
            licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
        if fecha_fin:
            licencias = licencias.filter(fecha_inicio__lte=fecha_fin)

        # Obtener registros de asistencia para contar atrasos e inasistencias
        registros_asistencia = RegistroAsistencia.objects.filter(funcionario=functorio)
        if year:
            registros_asistencia = registros_asistencia.filter(fecha__year=year)
        if mes:
            registros_asistencia = registros_asistencia.filter(fecha__month=mes)
        if fecha_inicio:
            registros_asistencia = registros_asistencia.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            registros_asistencia = registros_asistencia.filter(fecha__lte=fecha_fin)

        es_sereno_func = (
            getattr(functorio, 'funcion', None) == 'SERENO' or
            getattr(functorio, 'tipo_funcionario', None) == 'SERENO'
        )
        if es_sereno_func:
            total_inasistencias = 0
            total_atrasos = 0
            total_minutos_retraso = 0
        else:
            total_inasistencias = registros_asistencia.filter(estado='AUSENTE').count()
            total_atrasos = registros_asistencia.filter(estado='RETRASO').count()
            total_minutos_retraso = registros_asistencia.filter(estado='RETRASO').aggregate(
                total=Sum('minutos_retraso'))['total'] or 0
        
        html_string = render_to_string('reportes/pdf_individual.html', {
            'functorio': functorio,
            'cargo': functorio.get_funcion_display() or functorio.get_tipo_funcionario_display() or functorio.get_role_display(),
            'permisos': [p['permiso'] for p in permisos_con_dias],
            'permisos_con_dias': permisos_con_dias,
            'licencias': licencias,
            'dias_usados': dias_usados,
            'total_dias_licencias': licencias.aggregate(Sum('dias'))['dias__sum'] or 0,
            'total_inasistencias': total_inasistencias,
            'total_atrasos': total_atrasos,
            'total_minutos_retraso': total_minutos_retraso,
            'year': year,
            'mes': mes,
            'mes_nombre': {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}.get(int(mes) if mes else 0, ''),
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'director': CustomUser.objects.filter(role='DIRECTOR').first(),
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=reporte_{functorio.run}.pdf'
        response.write(result)
        return response



class MiReportePDFView(LoginRequiredMixin, View):
    """Genera el PDF individual del usuario actualmente autenticado."""

    def get(self, request):
        functorio = request.user

        year = request.GET.get('year', '')
        mes  = request.GET.get('mes', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin    = request.GET.get('fecha_fin', '')

        # Obtener todos los permisos aprobados del usuario (para cálculo cronológico)
        todos_permisos = SolicitudPermiso.objects.filter(
            usuario=functorio,
            estado='APROBADO'
        ).order_by('fecha_inicio', 'created_at')
        
        # Permisos con filtro para mostrar en el reporte
        permisos = todos_permisos
        if year:        permisos = permisos.filter(fecha_inicio__year=year)
        if mes:         permisos = permisos.filter(fecha_inicio__month=mes)
        if fecha_inicio:permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
        if fecha_fin:   permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
        
        # Calcular días disponibles cronológicamente por cada permiso
        BASE_DIAS_ADMINISTRATIVOS = 6.0
        permisos_con_dias = []
        dias_acumulados = 0.0
        
        for p in todos_permisos:
            dias_acumulados += float(p.dias_solicitados)
            dias_restantes = max(BASE_DIAS_ADMINISTRATIVOS - dias_acumulados, 0)
            # Solo incluir si pasa los filtros
            include = True
            if year and p.fecha_inicio.year != int(year):
                include = False
            if mes and p.fecha_inicio.month != int(mes):
                include = False
            if fecha_inicio and p.fecha_inicio < datetime.strptime(fecha_inicio, '%Y-%m-%d').date():
                include = False
            if fecha_fin and p.fecha_inicio > datetime.strptime(fecha_fin, '%Y-%m-%d').date():
                include = False
            
            if include:
                permisos_con_dias.append({
                    'permiso': p,
                    'dias_disponibles': dias_restantes,
                })
        
        dias_usados = sum(p['permiso'].dias_solicitados for p in permisos_con_dias)

        licencias = LicenciaMedica.objects.filter(usuario=functorio).order_by('fecha_inicio')
        if year:        licencias = licencias.filter(fecha_inicio__year=year)
        if mes:         licencias = licencias.filter(fecha_inicio__month=mes)
        if fecha_inicio:licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
        if fecha_fin:   licencias = licencias.filter(fecha_inicio__lte=fecha_fin)

        registros_asistencia = RegistroAsistencia.objects.filter(funcionario=functorio)
        if year:        registros_asistencia = registros_asistencia.filter(fecha__year=year)
        if mes:         registros_asistencia = registros_asistencia.filter(fecha__month=mes)
        if fecha_inicio:registros_asistencia = registros_asistencia.filter(fecha__gte=fecha_inicio)
        if fecha_fin:   registros_asistencia = registros_asistencia.filter(fecha__lte=fecha_fin)

        es_sereno_func = (
            getattr(functorio, 'funcion', None) == 'SERENO' or
            getattr(functorio, 'tipo_funcionario', None) == 'SERENO'
        )
        if es_sereno_func:
            total_inasistencias = 0
            total_atrasos = 0
            total_minutos_retraso = 0
        else:
            total_inasistencias = registros_asistencia.filter(estado='AUSENTE').count()
            total_atrasos       = registros_asistencia.filter(estado='RETRASO').count()
            total_minutos_retraso = registros_asistencia.filter(estado='RETRASO').aggregate(
                total=Sum('minutos_retraso'))['total'] or 0

        html_string = render_to_string('reportes/pdf_individual.html', {
            'functorio': functorio,
            'cargo': functorio.get_funcion_display() or functorio.get_tipo_funcionario_display() or functorio.get_role_display(),
            'permisos': [p['permiso'] for p in permisos_con_dias],
            'permisos_con_dias': permisos_con_dias,
            'licencias': licencias,
            'dias_usados': dias_usados,
            'total_dias_licencias': licencias.aggregate(Sum('dias'))['dias__sum'] or 0,
            'total_inasistencias': total_inasistencias,
            'total_atrasos': total_atrasos,
            'total_minutos_retraso': total_minutos_retraso,
            'year': year,
            'mes': mes,
            'mes_nombre': {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                           7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}.get(int(mes) if mes else 0, ''),
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'director': CustomUser.objects.filter(role='DIRECTOR').first(),
        })

        html   = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=mi_reporte_{functorio.run}.pdf'
        response.write(result)
        return response


class PDFColectivoView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generar Pdf de todos los empleados filtrados"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        # Obtener parámetros de filtro
        search = request.GET.get('search', '')
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        
        # Filtrar funcionarios - EXCLUIR ADMIN
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA'])
        if search:
            funcionarios = funcionarios.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(run__icontains=search)
            )
        
        # Preparar datos
        empleados_data = []
        total_dias_disponibles = 0
        total_licencias = 0
        
        for functorio in funcionarios.order_by('first_name', 'last_name'):
            permisos = SolicitudPermiso.objects.filter(usuario=functorio, estado='APROBADO')
            if year:
                permisos = permisos.filter(fecha_inicio__year=year)
            if mes:
                permisos = permisos.filter(fecha_inicio__month=mes)
            if fecha_inicio:
                permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
            
            licencias = LicenciaMedica.objects.filter(usuario=functorio)
            if year:
                licencias = licencias.filter(fecha_inicio__year=year)
            if mes:
                licencias = licencias.filter(fecha_inicio__month=mes)
            if fecha_inicio:
                licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                licencias = licencias.filter(fecha_inicio__lte=fecha_fin)
            
            dias_usados = permisos.aggregate(Sum('dias_solicitados'))['dias_solicitados__sum'] or 0
            dias_lic = licencias.aggregate(Sum('dias'))['dias__sum'] or 0
            total_lic = licencias.count()

            # Obtener registros de asistencia para contar atrasos e inasistencias
            registros_asistencia = RegistroAsistencia.objects.filter(funcionario=functorio)
            if year:
                registros_asistencia = registros_asistencia.filter(fecha__year=year)
            if mes:
                registros_asistencia = registros_asistencia.filter(fecha__month=mes)
            if fecha_inicio:
                registros_asistencia = registros_asistencia.filter(fecha__gte=fecha_inicio)
            if fecha_fin:
                registros_asistencia = registros_asistencia.filter(fecha__lte=fecha_fin)

            es_sereno_func = (
                getattr(functorio, 'funcion', None) == 'SERENO' or
                getattr(functorio, 'tipo_funcionario', None) == 'SERENO'
            )
            if es_sereno_func:
                total_atrasos = 0
                total_inasistencias = 0
                total_minutos_retraso = 0
            else:
                total_atrasos = registros_asistencia.filter(estado='RETRASO').count()
                total_inasistencias = registros_asistencia.filter(estado='AUSENTE').count()
                total_minutos_retraso = registros_asistencia.filter(estado='RETRASO').aggregate(
                    total=Sum('minutos_retraso'))['total'] or 0

            total_dias_disponibles += float(functorio.dias_disponibles)
            total_licencias += total_lic

            empleados_data.append({
                'funcionario': functorio,
                'dias_disponibles': functorio.dias_disponibles,
                'dias_usados': dias_usados,
                'total_licencias': total_lic,
                'dias_licencias': dias_lic,
                'total_atrasos': total_atrasos,
                'total_inasistencias': total_inasistencias,
                'total_minutos_retraso': total_minutos_retraso,
            })
        
        html_string = render_to_string('reportes/pdf_colectivo.html', {
            'empleados_data': empleados_data,
            'year': year,
            'mes': mes,
            'mes_nombre': {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}.get(mes, ''),
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'total_funcionarios': len(empleados_data),
            'total_dias_disponibles': total_dias_disponibles,
            'total_licencias': total_licencias,
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'director': CustomUser.objects.filter(role='DIRECTOR').first(),
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename=reporte_colectivo.pdf'
        response.write(result)
        return response


class ExportarExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar reporte detallado a Excel"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        # Obtener parámetros de filtro
        search = request.GET.get('search', '')
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        
        # Filtrar funcionarios
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'])
        if search:
            funcionarios = funcionarios.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(run__icontains=search)
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Detallado"
        
        # Encabezados
        ws.append(['Nombre', 'RUN', 'Cargo', 'Inasistencias sin Justificar', 'Atrasos', 'Min. Retraso Total', 'Días Disponibles', 'Días Usados', 'Días Licencia', 'Total Licencias'])
        
        # Preparar datos
        for functorio in funcionarios.order_by('first_name', 'last_name'):
            permisos = SolicitudPermiso.objects.filter(usuario=functorio, estado='APROBADO')
            if year:
                permisos = permisos.filter(fecha_inicio__year=year)
            if mes:
                permisos = permisos.filter(fecha_inicio__month=mes)
            if fecha_inicio:
                permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
            
            licencias = LicenciaMedica.objects.filter(usuario=functorio)
            if year:
                licencias = licencias.filter(fecha_inicio__year=year)
            if mes:
                licencias = licencias.filter(fecha_inicio__month=mes)
            if fecha_inicio:
                licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                licencias = licencias.filter(fecha_inicio__lte=fecha_fin)
            
            dias_usados = permisos.aggregate(Sum('dias_solicitados'))['dias_solicitados__sum'] or 0
            dias_licencia = licencias.aggregate(Sum('dias'))['dias__sum'] or 0

            # Contar inasistencias y atrasos
            registros = RegistroAsistencia.objects.filter(funcionario=functorio)
            if year:
                registros = registros.filter(fecha__year=year)
            if mes:
                registros = registros.filter(fecha__month=mes)
            if fecha_inicio:
                registros = registros.filter(fecha__gte=fecha_inicio)
            if fecha_fin:
                registros = registros.filter(fecha__lte=fecha_fin)

            es_sereno_func = (
                getattr(functorio, 'funcion', None) == 'SERENO' or
                getattr(functorio, 'tipo_funcionario', None) == 'SERENO'
            )
            total_inasistencias = registros.filter(estado='AUSENTE').count() if not es_sereno_func else 0
            total_atrasos = registros.filter(estado='RETRASO').count() if not es_sereno_func else 0
            total_minutos_retraso = (
                registros.filter(estado='RETRASO').aggregate(
                    total=Sum('minutos_retraso'))['total'] or 0
            ) if not es_sereno_func else 0
            
            ws.append([
                functorio.get_full_name(),
                functorio.run,
                functorio.get_funcion_display() or functorio.get_tipo_funcionario_display() or functorio.get_role_display(),
                total_inasistencias,
                total_atrasos,
                total_minutos_retraso,
                functorio.dias_disponibles,
                dias_usados,
                dias_licencia,
                licencias.count()
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_detallado.xlsx'
        wb.save(response)
        return response

class ReporteMensualDiasAdministrativosView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generar Pdf mensual de resumen de días administrativos"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        # Obtener parámetros
        year = request.GET.get('year', str(datetime.now().year))
        mes = request.GET.get('mes', str(datetime.now().month))
        
        try:
            year = int(year)
            mes = int(mes)
        except ValueError:
            year = datetime.now().year
            mes = datetime.now().month
        
        # Obtener todos los permisos aprobados del mes
        permisos = SolicitudPermiso.objects.filter(
            estado='APROBADO',
            fecha_inicio__year=year,
            fecha_inicio__month=mes
        ).select_related('usuario').order_by('fecha_inicio', 'created_at')
        
        # Calcular días disponibles cronológicamente para todo el año hasta cada permiso
        BASE_DIAS_ADMINISTRATIVOS = 6.0
        
        # Obtener TODOS los permisos del año para cada funcionario para cálculo cronológico
        usuarios_con_permisos = {}
        for p in permisos:
            if p.usuario_id not in usuarios_con_permisos:
                usuarios_con_permisos[p.usuario_id] = list(SolicitudPermiso.objects.filter(
                    usuario=p.usuario,
                    estado='APROBADO',
                    fecha_inicio__year=year
                ).order_by('fecha_inicio', 'created_at'))
        
        # Calcular días disponibles cronológicamente para cada funcionario
        usuarios_dias_calculados = {}
        for user_id, user_permisos in usuarios_con_permisos.items():
            dias_acumulados = 0.0
            user_dias = {}
            for p in user_permisos:
                dias_acumulados += float(p.dias_solicitados)
                user_dias[p.id] = max(BASE_DIAS_ADMINISTRATIVOS - dias_acumulados, 0)
            usuarios_dias_calculados[user_id] = user_dias
        
        # Preparar datos para el reporte - cada permiso es una fila
        empleados_data = []
        
        for permiso in permisos:
            dias_disp = usuarios_dias_calculados.get(permiso.usuario_id, {}).get(permiso.id, BASE_DIAS_ADMINISTRATIVOS)
            empleados_data.append({
                'funcionario': permiso.usuario,
                'cargo': permiso.usuario.get_funcion_display() or permiso.usuario.get_tipo_funcionario_display() or permiso.usuario.get_role_display(),
                'run': permiso.usuario.run,
                'nombre_completo': permiso.usuario.get_full_name() or permiso.usuario.username,
                'dias_solicitados': permiso.dias_solicitados,
                'dias_disponibles': dias_disp,
                'fecha_desde': permiso.fecha_inicio,
                'fecha_hasta': permiso.fecha_termino,
                'fecha_solicitud': permiso.created_at,
            })
        
        # Generar Pdf
        html_string = render_to_string('reportes/reporte_mensual_dias_administrativos.html', {
            'empleados_data': empleados_data,
            'year': year,
            'mes': mes,
            'mes_nombre': {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}.get(mes, ''),
            'total_funcionarios': len(empleados_data),
            'total_dias': sum(e['dias_solicitados'] for e in empleados_data),
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'director': CustomUser.objects.filter(role='DIRECTOR').first(),
            'establecimiento': 'Dirección de Educación Municipal Los Lagos',
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=reporte_dias_administrativos_{year}_{mes:02d}.pdf'
        response.write(result)
        return response


class ReporteMensualDiasAdministrativosExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generar Excel mensual de resumen de días administrativos"""

    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        # Obtener parámetros
        year = request.GET.get('year', str(datetime.now().year))
        mes = request.GET.get('mes', str(datetime.now().month))

        try:
            year = int(year)
            mes = int(mes)
        except ValueError:
            year = datetime.now().year
            mes = datetime.now().month

        # Obtener todos los permisos aprobados del mes
        permisos = SolicitudPermiso.objects.filter(
            estado='APROBADO',
            fecha_inicio__year=year,
            fecha_inicio__month=mes
        ).select_related('usuario').order_by('fecha_inicio', 'created_at')
        
        # Calcular días disponibles cronológicamente para todo el año hasta cada permiso
        BASE_DIAS_ADMINISTRATIVOS = 6.0
        
        # Obtener TODOS los permisos del año para cada funcionario para cálculo cronológico
        usuarios_con_permisos = {}
        for p in permisos:
            if p.usuario_id not in usuarios_con_permisos:
                usuarios_con_permisos[p.usuario_id] = list(SolicitudPermiso.objects.filter(
                    usuario=p.usuario,
                    estado='APROBADO',
                    fecha_inicio__year=year
                ).order_by('fecha_inicio', 'created_at'))
        
        # Calcular días disponibles cronológicamente para cada funcionario
        usuarios_dias_calculados = {}
        for user_id, user_permisos in usuarios_con_permisos.items():
            dias_acumulados = 0.0
            user_dias = {}
            for p in user_permisos:
                dias_acumulados += float(p.dias_solicitados)
                user_dias[p.id] = max(BASE_DIAS_ADMINISTRATIVOS - dias_acumulados, 0)
            usuarios_dias_calculados[user_id] = user_dias

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Días Administrativos {mes:02d}-{year}"

        # Encabezados
        headers = ['N°', 'Nombre Completo', 'RUN', 'Cargo', 'Días Solicitados', 'Días Disponibles', 'Fecha Desde', 'Fecha Hasta', 'Fecha Solicitud']
        ws.append(headers)

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15

        for i, permiso in enumerate(permisos, 1):
            dias_disp = usuarios_dias_calculados.get(permiso.usuario_id, {}).get(permiso.id, BASE_DIAS_ADMINISTRATIVOS)
            ws.append([
                i,
                permiso.usuario.get_full_name() or permiso.usuario.username,
                permiso.usuario.run,
                permiso.usuario.get_funcion_display() or permiso.usuario.get_tipo_funcionario_display() or permiso.usuario.get_role_display(),
                permiso.dias_solicitados,
                dias_disp,
                permiso.fecha_inicio.strftime("%d/%m/%Y") if permiso.fecha_inicio else "",
                permiso.fecha_termino.strftime("%d/%m/%Y") if permiso.fecha_termino else "",
                permiso.created_at.strftime("%d/%m/%Y") if permiso.created_at else ""
            ])

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'reporte_dias_administrativos_{year}_{mes:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class ExportarDAEMExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar reporte DAEM/DAEM3 a Excel"""

    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')
        tipo = request.GET.get('tipo', '').lower()  # Para DAEM3, convertir a minúsculas

        # Si es DAEM3, mostrar formato de asistencia laboral
        if tipo in ['docentes', 'docente', 'asistentes', 'asistente']:
            return self._exportar_daem3_excel(year, mes, tipo)

        # Formato regular DAEM (permisos administrativos)
        wb = openpyxl.Workbook()

        # Pestaña 1: Nómina Funcionarios
        ws_nomina = wb.active
        ws_nomina.title = "Nomina Funcionarios"

        # Agregar encabezado
        ws_nomina.append(['N°', 'Nombre Completo', 'RUN', 'Cargo'])
        for col in ['A']:
            ws_nomina.column_dimensions[col].width = 10
        for col in ['B', 'C', 'D']:
            ws_nomina.column_dimensions[col].width = 30

        # Pestaña 2: Permisos Administrativos
        ws_permisos = wb.create_sheet(title="Permisos Administrativos")
        ws_permisos.append(['N°', 'Nombre Completo', 'RUN', 'Cargo', 'Tipo de Permiso Administrativo', 'Fecha de Inicio', 'Fecha de Término', 'Cantidad de Días', 'Observaciones'])
        for col in ['A']:
            ws_permisos.column_dimensions[col].width = 10
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws_permisos.column_dimensions[col].width = 25

        # Obtener funcionarios
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']).order_by('first_name', 'last_name')

        # Procesar nómina (empezar desde fila 2, después de los encabezados)
        for i, f in enumerate(funcionarios, 2):
            ws_nomina.cell(row=i, column=1).value = i - 1  # N°
            ws_nomina.cell(row=i, column=2).value = f.get_full_name() or f.username  # Nombre Completo
            ws_nomina.cell(row=i, column=3).value = f.run  # RUN
            ws_nomina.cell(row=i, column=4).value = f.get_funcion_display() or ""  # Cargo

        # Obtener todas las justificaciones del mes/año para estos funcionarios
        from asistencia.models import RegistroAsistencia
        registros_justificados = RegistroAsistencia.objects.filter(
            funcionario__in=funcionarios,
            estado='JUSTIFICADO'
        ).order_by('fecha')

        if year:
            registros_justificados = registros_justificados.filter(fecha__year=year)
        if mes:
            registros_justificados = registros_justificados.filter(fecha__month=mes)

        # Agrupar por funcionario
        justificaciones_por_func = {}
        for r in registros_justificados:
            if r.justificacion_manual:
                if r.hora_entrada_real and r.hora_salida_real:
                    hora_str = f"{r.hora_entrada_real.strftime('%H:%M')} a {r.hora_salida_real.strftime('%H:%M')}"
                elif r.hora_entrada_real:
                    hora_str = r.hora_entrada_real.strftime('%H:%M')
                elif r.hora_salida_real:
                    hora_str = r.hora_salida_real.strftime('%H:%M')
                else:
                    hora_str = "--:--"
                
                just_text = f"{r.fecha.strftime('%d-%m-%Y')} ({hora_str}): {r.justificacion_manual}"
                if r.funcionario_id not in justificaciones_por_func:
                    justificaciones_por_func[r.funcionario_id] = []
                justificaciones_por_func[r.funcionario_id].append(just_text)

        permisos = SolicitudPermiso.objects.filter(estado='APROBADO', usuario__in=funcionarios).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name', 'fecha_inicio')
        if year:
            permisos = permisos.filter(fecha_inicio__year=year)
        if mes:
            permisos = permisos.filter(fecha_inicio__month=mes)

        for i, p in enumerate(permisos, 1):
            justs = justificaciones_por_func.get(p.usuario.id, [])
            justs_str = "; ".join(justs) if justs else ""

            ws_permisos.append([
                i,
                p.usuario.get_full_name() or p.usuario.username,
                p.usuario.run,
                p.usuario.get_funcion_display() or "",
                "",
                p.fecha_inicio.strftime("%d-%m-%Y") if p.fecha_inicio else "",
                p.fecha_termino.strftime("%d-%m-%Y") if p.fecha_termino else "",
                float(p.dias_solicitados),
                ""
            ])

        # Firma
        firma_row = len(permisos) + 10
        ws_permisos.cell(row=firma_row, column=2).value = "Director Colegio Los Alerces"
        ws_permisos.cell(row=firma_row + 2, column=2).value = "Puerto Montt, " + datetime.now().strftime("%d de %B de %Y")

        # Ajustar ancho de columnas
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws_permisos.column_dimensions[col].width = 20

        # Pestaña 3: Licencias Médicas
        ws_licencias = wb.create_sheet(title="Licencias Medicas")
        ws_licencias.append(['N°', 'Nombre Completo', 'Cargo', 'Fecha de Inicio', 'Fecha de Término', 'Cantidad de Días', 'Observaciones'])
        for col in ['A']:
            ws_licencias.column_dimensions[col].width = 10
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws_licencias.column_dimensions[col].width = 25

        # Procesar licencias médicas
        licencias = LicenciaMedica.objects.filter(usuario__in=funcionarios).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name', 'fecha_inicio')
        if year:
            licencias = licencias.filter(fecha_inicio__year=year)
        if mes:
            licencias = licencias.filter(fecha_inicio__month=mes)

        for i, l in enumerate(licencias, 2):  # Start from row 2 (after headers)
            ws_licencias.append([
                i - 1,  # N°
                l.usuario.get_full_name() or l.usuario.username,
                l.usuario.get_funcion_display() or "",
                l.fecha_inicio.strftime("%d-%m-%Y") if l.fecha_inicio else "",
                l.fecha_termino.strftime("%d-%m-%Y") if l.fecha_termino else "",
                l.dias,
                ""  # Observaciones vacío
            ])

        # Firma en la pestaña de licencias
        firma_row_lic = max(len(licencias) + 5, 5)
        ws_licencias.cell(row=firma_row_lic, column=2).value = "Director Colegio Los Alerces"
        ws_licencias.cell(row=firma_row_lic + 2, column=2).value = "Puerto Montt, " + datetime.now().strftime("%d de %B de %Y")

        # Pestaña 4: Horarios Funcionarios
        from asistencia.models import HorarioFuncionario, DiaHorario
        from django.db.models import Q as QModel
        
        ws_horarios = wb.create_sheet(title="Horarios Funcionarios")
        
        # Verificar si hay serenos
        tiene_serenos = CustomUser.objects.filter(
            role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']
        ).filter(
            QModel(funcion='SERENO') | QModel(tipo_funcionario='SERENO')
        ).exists()
        
        # Encabezados dinámicos
        if tiene_serenos:
            headers_horarios = ['N°', 'Funcionario', 'RUN', 'Cargo',
                                'Lunes S1', 'Martes S1', 'Miércoles S1', 'Jueves S1', 'Viernes S1', 'Sábado S1', 'Domingo S1',
                                'Lunes S2', 'Martes S2', 'Miércoles S2', 'Jueves S2', 'Viernes S2', 'Sábado S2', 'Domingo S2']
        else:
            headers_horarios = ['N°', 'Funcionario', 'RUN', 'Cargo',
                                'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        ws_horarios.append(headers_horarios)
        
        # Set column widths
        ws_horarios.column_dimensions['A'].width = 8   # N°
        ws_horarios.column_dimensions['B'].width = 30  # Funcionario
        ws_horarios.column_dimensions['C'].width = 15  # RUN
        ws_horarios.column_dimensions['D'].width = 25  # Cargo
        if tiene_serenos:
            for col_letter in ['E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
                ws_horarios.column_dimensions[col_letter].width = 14
        else:
            for col_letter in ['E','F','G','H','I','J','K']:
                ws_horarios.column_dimensions[col_letter].width = 18

        # Obtener horarios con prefetch
        horarios_con_dias = HorarioFuncionario.objects.filter(activo=True).prefetch_related('dias')
        
        # Construir diccionarios separados para cada funcionario
        horarios_dict_s1 = {}
        horarios_dict_s2 = {}
        horarios_dict_universal = {}
        
        for horario in horarios_con_dias:
            d_s1, d_s2, d_uni = {}, {}, {}
            
            for dia in horario.dias.all():
                dia_nombre = DIA_SEMANA_MAP[dia.dia_semana]
                if dia.activo and dia.hora_entrada:
                    hora_str = f"{dia.hora_entrada.strftime('%H:%M')}"
                    if dia.hora_salida:
                        hora_str += f" - {dia.hora_salida.strftime('%H:%M')}"
                else:
                    hora_str = "Libre"
                
                if dia.semana_tipo == 1:
                    d_s1[dia_nombre] = hora_str
                elif dia.semana_tipo == 2:
                    d_s2[dia_nombre] = hora_str
                else:
                    d_uni[dia_nombre] = hora_str
            
            horarios_dict_s1[horario.funcionario_id] = d_s1
            horarios_dict_s2[horario.funcionario_id] = d_s2
            horarios_dict_universal[horario.funcionario_id] = d_uni

        # Procesar horarios semanales
        row_idx = 2
        for func in funcionarios:
            es_sereno = (
                getattr(func, 'funcion', None) == 'SERENO'
                or getattr(func, 'tipo_funcionario', None) == 'SERENO'
            )
            
            if tiene_serenos and es_sereno:
                # Formato extendido para serenos con semana 1 y semana 2
                d_s1 = horarios_dict_s1.get(func.id, {})
                d_s2 = horarios_dict_s2.get(func.id, {})
                d_uni = horarios_dict_universal.get(func.id, {})
                
                # Merge universal into both weeks
                for dia, hora in d_uni.items():
                    if dia not in d_s1: d_s1[dia] = hora
                    if dia not in d_s2: d_s2[dia] = hora
                
                dias_nombres = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
                ws_horarios.cell(row=row_idx, column=1).value = row_idx - 1
                ws_horarios.cell(row=row_idx, column=2).value = func.get_full_name() or func.username
                ws_horarios.cell(row=row_idx, column=3).value = func.run
                ws_horarios.cell(row=row_idx, column=4).value = func.get_funcion_display() or ""
                
                for col_offset, dia_nombre in enumerate(dias_nombres, 5):
                    ws_horarios.cell(row=row_idx, column=col_offset).value = d_s1.get(dia_nombre, 'Libre')
                for col_offset, dia_nombre in enumerate(dias_nombres, 12):
                    ws_horarios.cell(row=row_idx, column=col_offset).value = d_s2.get(dia_nombre, 'Libre')
            else:
                # Formato simple para no serenos
                dias_dict = {}
                horario = getattr(func, 'horario', None)
                if horario:
                    for d in horario.dias.all():
                        if d.activo and d.hora_entrada:
                            hora_str = f"{d.hora_entrada.strftime('%H:%M')}"
                            if d.hora_salida:
                                hora_str += f" - {d.hora_salida.strftime('%H:%M')}"
                            dias_dict[d.dia_semana] = hora_str
                        else:
                            dias_dict[d.dia_semana] = "Libre"
                
                ws_horarios.cell(row=row_idx, column=1).value = row_idx - 1
                ws_horarios.cell(row=row_idx, column=2).value = func.get_full_name() or func.username
                ws_horarios.cell(row=row_idx, column=3).value = func.run
                ws_horarios.cell(row=row_idx, column=4).value = func.get_funcion_display() or ""
                
                for dia_num in range(7):
                    ws_horarios.cell(row=row_idx, column=5 + dia_num).value = dias_dict.get(dia_num, 'Libre')
            
            row_idx += 1

        # Pestaña 5: Justificaciones (Todas las justificaciones de asistencia del mes)
        ws_justificaciones = wb.create_sheet(title="Justificaciones")
        ws_justificaciones.append(['N°', 'Nombre Completo', 'RUN', 'Cargo', 'Justificaciones Asistencia'])
        
        ws_justificaciones.column_dimensions['A'].width = 8
        ws_justificaciones.column_dimensions['B'].width = 30
        ws_justificaciones.column_dimensions['C'].width = 15
        ws_justificaciones.column_dimensions['D'].width = 25
        ws_justificaciones.column_dimensions['E'].width = 60

        just_idx = 1
        for f in funcionarios:
            justs = justificaciones_por_func.get(f.id, [])
            if justs:
                justs_str = "; ".join(justs)
                ws_justificaciones.append([
                    just_idx,
                    f.get_full_name() or f.username,
                    f.run,
                    f.get_funcion_display() or "",
                    justs_str
                ])
                just_idx += 1

        # Respuesta HTTP
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'DAEM_{year}_{int(mes):02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _exportar_daem3_excel(self, year, mes, tipo):
        """Exportar DAEM3 - Informe de Asistencia Laboral"""
        from asistencia.models import RegistroAsistencia

        # Determinar el tipo de personal
        if tipo in ['docentes', 'docente']:
            tipo_display = 'Docentes'
            # Filtrar por docentes usando el campo tipo_funcionario
            funcionarios = CustomUser.objects.filter(
                tipo_funcionario='DOCENTE'
            ).order_by('first_name', 'last_name')
        else:  # asistentes
            tipo_display = 'Asistente de la Educación'
            # Filtrar por asistentes usando el campo tipo_funcionario
            funcionarios = CustomUser.objects.filter(
                tipo_funcionario='ASISTENTE'
            ).order_by('first_name', 'last_name')

        # Obtener registros del mes
        registros_mes = RegistroAsistencia.objects.filter(
            fecha__year=int(year),
            fecha__month=int(mes),
            funcionario__in=funcionarios
        ).select_related('funcionario')

        # Pre-cargar datos para conteo de ausencias virtuales
        from asistencia.models import DiaFestivo, AnoEscolar, HorarioFuncionario
        from permisos.models import SolicitudPermiso
        from licencias.models import LicenciaMedica
        from datetime import date as date_cls, timedelta
        import calendar as cal_module

        today = date_cls.today()
        primer_dia_mes = date_cls(int(year), int(mes), 1)
        ultimo_dia_mes = date_cls(int(year), int(mes), cal_module.monthrange(int(year), int(mes))[1])
        ultimo_dia = min(ultimo_dia_mes, today)

        festivos = set(DiaFestivo.objects.filter(
            fecha__year=int(year), fecha__month=int(mes)
        ).values_list('fecha', flat=True))

        ano_escolar = AnoEscolar.objects.filter(ano=int(year)).first()

        horarios_dict = {}
        for horario in HorarioFuncionario.objects.filter(
            funcionario__in=funcionarios, activo=True
        ).prefetch_related('dias'):
            horarios_dict[horario.funcionario_id] = set(
                horario.dias.filter(activo=True).values_list('dia_semana', flat=True)
            )

        # Agrupar por funcionario y calcular atrasos e inasistencias
        funcionarios_data = []
        for func in funcionarios:
            func_registros = registros_mes.filter(funcionario=func)

            # Calcular atrasos (minutos acumulado mensual)
            total_atrasos = sum(r.minutos_retraso or 0 for r in func_registros if r.estado == 'RETRASO')

            # Inasistencias en BD (estado guardado como AUSENTE)
            ausencias_db = func_registros.filter(estado='AUSENTE').count()

            # Pre-cargar licencias y permisos del funcionario para el mes
            licencias_func = set()
            for lic in LicenciaMedica.objects.filter(
                usuario=func, fecha_inicio__lte=ultimo_dia_mes
            ):
                fin_lic = lic.fecha_inicio + timedelta(days=lic.dias - 1)
                inicio = max(lic.fecha_inicio, primer_dia_mes)
                fin = min(fin_lic, ultimo_dia_mes)
                d_lic = inicio
                while d_lic <= fin:
                    licencias_func.add(d_lic)
                    d_lic += timedelta(days=1)

            permisos_func = set()
            for perm in SolicitudPermiso.objects.filter(
                usuario=func, estado='APROBADO',
                fecha_inicio__lte=ultimo_dia_mes
            ).filter(
                Q(fecha_termino__gte=primer_dia_mes) | Q(fecha_termino__isnull=True)
            ):
                inicio = max(perm.fecha_inicio, primer_dia_mes)
                fin = perm.fecha_termino or ultimo_dia_mes
                fin = min(fin, ultimo_dia_mes)
                d_perm = inicio
                while d_perm <= fin:
                    permisos_func.add(d_perm)
                    d_perm += timedelta(days=1)

            # Inasistencias virtuales: días laborales pasados sin ningún registro
            # y sin cobertura de licencia/permiso (igual a lo que muestra la vista)
            fechas_con_registro = set(func_registros.values_list('fecha', flat=True))
            es_sereno = (func.funcion == 'SERENO') or (func.tipo_funcionario == 'SERENO')
            
            # Si no tiene horario, no debe tener atrasos ni inasistencias (aplica a todos)
            tiene_horario = func.id in horarios_dict
            if not tiene_horario:
                total_atrasos = 0
                total_inasistencias = 0
            else:
                dias_laborales = horarios_dict.get(func.id, set())

                ausencias_virtuales = 0
                d = primer_dia_mes
                while d <= ultimo_dia:
                    if d not in fechas_con_registro and d >= func.date_joined.date():
                        dia_semana = d.weekday()
                        if dia_semana >= 5 and not es_sereno:
                            d += timedelta(days=1)
                            continue
                        if d in festivos or d in licencias_func or d in permisos_func:
                            d += timedelta(days=1)
                            continue
                        en_ano_escolar = True
                        if ano_escolar:
                            en_ano_escolar = (
                                ano_escolar.sem1_inicio <= d <= ano_escolar.sem1_fin or
                                ano_escolar.sem2_inicio <= d <= ano_escolar.sem2_fin
                            )
                        if not en_ano_escolar:
                            d += timedelta(days=1)
                            continue
                        if dia_semana in dias_laborales or es_sereno:
                            ausencias_virtuales += 1
                    d += timedelta(days=1)

                total_inasistencias = ausencias_db + ausencias_virtuales

            # Solo incluir si tiene inasistencias injustificadas o atrasos acumulados >= 60 minutos
            if total_inasistencias > 0 or total_atrasos >= 60:
                funcionarios_data.append({
                    'funcionario': func,
                    'atrasos': total_atrasos,
                    'inasistencias': total_inasistencias
                })

        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Informe Asistencia Laboral"

        # Título
        ws['A1'] = "Informe Asistencia Laboral"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Establecimiento
        ws['A3'] = "Establecimiento: Colegio Los Alerces"

        # Mes de Informe
        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        ws['A5'] = f"Mes de Informe: {meses[int(mes)-1]}"

        # Tipo
        ws['A7'] = tipo_display

        # Encabezados de tabla
        headers = ['Nombre y Apellidos', 'RUN', 'Atrasos', 'Inasistencias']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Datos
        for i, data in enumerate(funcionarios_data, 10):
            ws.cell(row=i, column=1).value = data['funcionario'].get_full_name() or data['funcionario'].username
            ws.cell(row=i, column=2).value = data['funcionario'].run
            # Convert minutes to hours and minutes string
            total_minutes = data['atrasos']
            hrs = total_minutes // 60
            mins = total_minutes % 60
            ws.cell(row=i, column=3).value = f"{hrs}h {mins}m"
            ws.cell(row=i, column=4).value = data['inasistencias']

            # Bordes para las celdas de datos
            for col in range(1, 5):
                cell = ws.cell(row=i, column=col)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        # Firma
        firma_row = len(funcionarios_data) + 12
        ws.cell(row=firma_row, column=1).value = "Firma y Timbre Director"

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'Informe_Asistencia_Laboral_{tipo_display}_{year}_{int(mes):02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportarDAEMPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar reporte DAEM/DAEM3 a PDF"""

    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')
        tipo = request.GET.get('tipo', '').lower()  # Para DAEM3, convertir a minúsculas

        # Si es DAEM3, mostrar formato de asistencia laboral
        if tipo in ['docentes', 'docente', 'asistentes', 'asistente']:
            return self._exportar_daem3_pdf(year, mes, tipo)

        # Formato regular DAEM (permisos administrativos)
        # Obtener funcionarios
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']).order_by('first_name', 'last_name')

        # Preparar datos
        nomina_data = []
        for i, f in enumerate(funcionarios, 1):
            nomina_data.append({
                'numero': i,
                'nombre': f.get_full_name() or f.username,
                'run': f.run,
                'cargo': f.get_funcion_display() or ""
            })

        permisos_data = []
        permisos = SolicitudPermiso.objects.filter(estado='APROBADO', usuario__in=funcionarios).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name', 'fecha_inicio')
        if year:
            permisos = permisos.filter(fecha_inicio__year=year)
        if mes:
            permisos = permisos.filter(fecha_inicio__month=mes)

        for i, p in enumerate(permisos, 1):
            permisos_data.append({
                'numero': i,
                'nombre': p.usuario.get_full_name() or p.usuario.username,
                'run': p.usuario.run,
                'cargo': p.usuario.get_funcion_display() or "",
                'tipo_permiso': "",
                'fecha_inicio': p.fecha_inicio.strftime("%d-%m-%Y") if p.fecha_inicio else "",
                'fecha_termino': p.fecha_termino.strftime("%d-%m-%Y") if p.fecha_termino else "",
                'cantidad_dias': float(p.dias_solicitados),
                'observaciones': ""
            })

        # Renderizar template HTML para PDF
        html_content = render_to_string('reportes/daem_pdf.html', {
            'nomina_data': nomina_data,
            'permisos_data': permisos_data,
            'year': year,
            'mes': mes,
            'fecha_actual': datetime.now(),
        })

        # Generar PDF
        pdf_file = HTML(string=html_content).write_pdf()

        # Crear respuesta HTTP
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'DAEM_{year}_{int(mes):02d}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _exportar_daem3_pdf(self, year, mes, tipo):
        """Exportar DAEM3 - Informe de Asistencia Laboral"""
        from asistencia.models import RegistroAsistencia

        # Determinar el tipo de personal
        if tipo in ['docentes', 'docente']:
            tipo_display = 'Docentes'
            # Filtrar por docentes usando el campo tipo_funcionario
            funcionarios = CustomUser.objects.filter(
                tipo_funcionario='DOCENTE'
            ).order_by('first_name', 'last_name')
        else:  # asistentes
            tipo_display = 'Asistente de la Educación'
            # Filtrar por asistentes usando el campo tipo_funcionario
            funcionarios = CustomUser.objects.filter(
                tipo_funcionario='ASISTENTE'
            ).order_by('first_name', 'last_name')

        # Obtener registros del mes
        registros_mes = RegistroAsistencia.objects.filter(
            fecha__year=int(year),
            fecha__month=int(mes),
            funcionario__in=funcionarios
        ).select_related('funcionario')

        # Agrupar por funcionario y calcular atrasos e inasistencias
        # Pre-cargar datos para conteo de ausencias virtuales
        from asistencia.models import DiaFestivo, AnoEscolar, HorarioFuncionario
        from datetime import date as date_cls, timedelta
        import calendar as cal_module

        today = date_cls.today()
        primer_dia_mes = date_cls(int(year), int(mes), 1)
        ultimo_dia_mes = date_cls(int(year), int(mes), cal_module.monthrange(int(year), int(mes))[1])
        ultimo_dia = min(ultimo_dia_mes, today)

        festivos = set(DiaFestivo.objects.filter(
            fecha__year=int(year), fecha__month=int(mes)
        ).values_list('fecha', flat=True))

        ano_escolar = AnoEscolar.objects.filter(ano=int(year)).first()

        horarios_dict = {}
        for horario in HorarioFuncionario.objects.filter(
            funcionario__in=funcionarios, activo=True
        ).prefetch_related('dias'):
            horarios_dict[horario.funcionario_id] = set(
                horario.dias.filter(activo=True).values_list('dia_semana', flat=True)
            )

        # Agrupar por funcionario y calcular atrasos e inasistencias
        funcionarios_data = []
        for func in funcionarios:
            func_registros = registros_mes.filter(funcionario=func)

            # Calcular atrasos (minutos acumulado mensual)
            total_atrasos = sum(r.minutos_retraso or 0 for r in func_registros if r.estado == 'RETRASO')

            # Inasistencias en BD (estado guardado como AUSENTE)
            ausencias_db = func_registros.filter(estado='AUSENTE').count()

            # Pre-cargar licencias y permisos del funcionario para excluirlos
            from permisos.models import SolicitudPermiso
            from licencias.models import LicenciaMedica

            licencias_func = set()
            for lic in LicenciaMedica.objects.filter(
                usuario=func, fecha_inicio__lte=ultimo_dia_mes
            ):
                fin_lic = lic.fecha_inicio + timedelta(days=lic.dias - 1)
                inicio = max(lic.fecha_inicio, primer_dia_mes)
                fin = min(fin_lic, ultimo_dia_mes)
                d_lic = inicio
                while d_lic <= fin:
                    licencias_func.add(d_lic)
                    d_lic += timedelta(days=1)

            permisos_func = set()
            for perm in SolicitudPermiso.objects.filter(
                usuario=func, estado='APROBADO',
                fecha_inicio__lte=ultimo_dia_mes
            ).filter(
                Q(fecha_termino__gte=primer_dia_mes) | Q(fecha_termino__isnull=True)
            ):
                inicio = max(perm.fecha_inicio, primer_dia_mes)
                fin = perm.fecha_termino or ultimo_dia_mes
                fin = min(fin, ultimo_dia_mes)
                d_perm = inicio
                while d_perm <= fin:
                    permisos_func.add(d_perm)
                    d_perm += timedelta(days=1)

            # Inasistencias virtuales: días laborales pasados sin ningún registro
            # y sin cobertura de licencia/permiso (igual a lo que muestra la vista)
            fechas_con_registro = set(func_registros.values_list('fecha', flat=True))
            es_sereno = (func.funcion == 'SERENO') or (func.tipo_funcionario == 'SERENO')
            
            # Si no tiene horario, no debe tener atrasos ni inasistencias (aplica a todos)
            tiene_horario = func.id in horarios_dict
            if not tiene_horario:
                total_atrasos = 0
                total_inasistencias = 0
            else:
                dias_laborales = horarios_dict.get(func.id, set())

                ausencias_virtuales = 0
                d = primer_dia_mes
                while d <= ultimo_dia:
                    if d not in fechas_con_registro and d >= func.date_joined.date():
                        dia_semana = d.weekday()
                        if dia_semana >= 5 and not es_sereno:
                            d += timedelta(days=1)
                            continue
                        if d in festivos or d in licencias_func or d in permisos_func:
                            d += timedelta(days=1)
                            continue
                        en_ano_escolar = True
                        if ano_escolar:
                            en_ano_escolar = (
                                ano_escolar.sem1_inicio <= d <= ano_escolar.sem1_fin or
                                ano_escolar.sem2_inicio <= d <= ano_escolar.sem2_fin
                            )
                        if not en_ano_escolar:
                            d += timedelta(days=1)
                            continue
                        if dia_semana in dias_laborales or es_sereno:
                            ausencias_virtuales += 1
                    d += timedelta(days=1)

                total_inasistencias = ausencias_db + ausencias_virtuales

            # Solo incluir si tiene inasistencias injustificadas o atrasos acumulados >= 60 minutos
            if total_inasistencias > 0 or total_atrasos >= 60:
                funcionarios_data.append({
                    'funcionario': func,
                    'atrasos': f"{total_atrasos // 60}h {total_atrasos % 60}m",
                    'inasistencias': total_inasistencias
                })

        # Mes en texto
        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        mes_texto = meses[int(mes)-1]

        # Renderizar template HTML para PDF
        html_content = render_to_string('reportes/daem3_pdf.html', {
            'titulo': 'Informe Asistencia Laboral',
            'establecimiento': 'Colegio Los Alerces',
            'mes_informe': mes_texto,
            'tipo_personal': tipo_display,
            'funcionarios': funcionarios_data,
            'fecha_actual': datetime.now(),
        })

        # Generar PDF
        pdf_file = HTML(string=html_content).write_pdf()

        # Crear respuesta HTTP
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'Informe_Asistencia_Laboral_{tipo_display}_{year}_{int(mes):02d}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class ExportarHorariosExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar los horarios semanales de todos los funcionarios a Excel"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        from asistencia.models import HorarioFuncionario, DiaHorario

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horarios del Personal"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        # Verificar si hay serenos
        tiene_serenos = CustomUser.objects.filter(
            is_active=True,
            role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']
        ).filter(
            Q(funcion='SERENO') | Q(tipo_funcionario='SERENO')
        ).exists()
        
        # Encabezados siempre normales (solo serenos se expanden en filas Semana 1 / Semana 2)
        headers = ['N°', 'Funcionario', 'RUN', 'Cargo',
                    'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo',
                    'Horas Reales (Ciclo)', 'Prom. Semanal Ajustado (44h + carry)']
        
        ws.append(headers)
        
        # Aplicar estilos a encabezados
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # Ajustar anchos de columna (formato normal)
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        for col_letter in ['E','F','G','H','I','J','K']:
            ws.column_dimensions[col_letter].width = 14
        ws.column_dimensions['L'].width = 18   # Horas Reales
        ws.column_dimensions['M'].width = 22   # Promedio ajustado

        # Obtener todos los funcionarios activos
        funcionarios = CustomUser.objects.filter(
            is_active=True,
            role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']
        ).order_by('first_name', 'last_name')

        for i, f in enumerate(funcionarios, 1):
            horario = getattr(f, 'horario', None)
            es_sereno = (
                getattr(f, 'funcion', None) == 'SERENO'
                or getattr(f, 'tipo_funcionario', None) == 'SERENO'
            )
            total_minutos = 0

            if es_sereno:
                # Solo serenos se expanden en Semana 1 / Semana 2
                d_s1 = {}
                d_s2 = {}
                raw_s1 = 0
                raw_s2 = 0

                if horario:
                    for day in range(7):
                        nombre = DIA_SEMANA_MAP[day]

                        d1 = horario.dias.filter(dia_semana=day, semana_tipo=1).first()
                        if not d1:
                            d1 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()

                        d2 = horario.dias.filter(dia_semana=day, semana_tipo=2).first()
                        if not d2:
                            d2 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()

                        def calc_minutes(dh):
                            if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                                h1, m1 = dh.hora_entrada.hour, dh.hora_entrada.minute
                                h2, m2 = dh.hora_salida.hour, dh.hora_salida.minute
                                min1 = h1 * 60 + m1
                                min2 = h2 * 60 + m2
                                if min2 < min1:
                                    min2 += 24 * 60
                                return min2 - min1
                            return 0

                        def get_str(dh):
                            if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                                return f"{dh.hora_entrada.strftime('%H:%M')} - {dh.hora_salida.strftime('%H:%M')}"
                            return "Libre"

                        mins1 = calc_minutes(d1)
                        mins2 = calc_minutes(d2)

                        raw_s1 += mins1
                        raw_s2 += mins2

                        d_s1[nombre] = get_str(d1)
                        d_s2[nombre] = get_str(d2)
                else:
                    for nombre in DIA_SEMANA_MAP.values():
                        d_s1[nombre] = "Libre"
                        d_s2[nombre] = "Libre"

                # Suma real del ciclo
                total_real_minutos = raw_s1 + raw_s2
                h_real = total_real_minutos // 60
                m_real = total_real_minutos % 60
                horas_reales_str = f"{h_real}h {m_real}m" if m_real > 0 else f"{h_real}h"
                if total_real_minutos == 0:
                    horas_reales_str = "No configurado"

                # Promedio ajustado (44h + carry)
                MAX = 44 * 60
                adj1 = raw_s1
                adj2 = raw_s2
                if adj1 > MAX:
                    excess = adj1 - MAX
                    adj1 = MAX
                    adj2 += excess
                prom_min = (adj1 + adj2) // 2
                h_aj = prom_min // 60
                m_aj = prom_min % 60
                prom_ajustado_str = f"{h_aj}h {m_aj}m" if m_aj > 0 else f"{h_aj}h"
                if prom_min == 0:
                    prom_ajustado_str = "N/C"

                # Fila Semana 1
                ws.append([
                    i, f"{f.get_full_name()} (Semana 1)", f.run, f.get_funcion_display() or f.get_role_display(),
                    d_s1.get('Lunes','Libre'), d_s1.get('Martes','Libre'), d_s1.get('Miércoles','Libre'),
                    d_s1.get('Jueves','Libre'), d_s1.get('Viernes','Libre'), d_s1.get('Sábado','Libre'), d_s1.get('Domingo','Libre'),
                    horas_reales_str, prom_ajustado_str
                ])
                # Fila Semana 2
                ws.append([
                    i, f"{f.get_full_name()} (Semana 2)", f.run, f.get_funcion_display() or f.get_role_display(),
                    d_s2.get('Lunes','Libre'), d_s2.get('Martes','Libre'), d_s2.get('Miércoles','Libre'),
                    d_s2.get('Jueves','Libre'), d_s2.get('Viernes','Libre'), d_s2.get('Sábado','Libre'), d_s2.get('Domingo','Libre'),
                    "", ""
                ])

            else:
                # Formato normal para no serenos (siempre 1 fila)
                dias_dict = {}
                if horario:
                    for d in horario.dias.all():
                        if d.activo and d.hora_entrada and d.hora_salida:
                            dias_dict[d.dia_semana] = f"{d.hora_entrada.strftime('%H:%M')} - {d.hora_salida.strftime('%H:%M')}"
                            h1, m1 = d.hora_entrada.hour, d.hora_entrada.minute
                            h2, m2 = d.hora_salida.hour, d.hora_salida.minute
                            min1 = h1 * 60 + m1
                            min2 = h2 * 60 + m2
                            if min2 < min1: min2 += 24 * 60
                            total_minutos += (min2 - min1)
                        else:
                            dias_dict[d.dia_semana] = "Libre"
                
                h_total = total_minutos // 60
                m_total = total_minutos % 60
                horas_str = f"{h_total}h {m_total}m" if m_total > 0 else f"{h_total}h"
                if total_minutos == 0: horas_str = "No configurado"

                ws.append([
                    i, f.get_full_name(), f.run, f.get_funcion_display() or f.get_role_display(),
                    dias_dict.get(0,'Libre'), dias_dict.get(1,'Libre'), dias_dict.get(2,'Libre'),
                    dias_dict.get(3,'Libre'), dias_dict.get(4,'Libre'), dias_dict.get(5,'Libre'), dias_dict.get(6,'Libre'),
                    horas_str, "N/A"
                ])
            for cell in ws[ws.max_row][4:]:
                cell.alignment = alignment

        # Generar archivo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=horarios_personal_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class ExportarHorarioIndividualPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar el horario semanal de un funcionario individual a PDF"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request, usuario_id):
        from asistencia.models import HorarioFuncionario, DiaHorario
        
        try:
            funcionario = CustomUser.objects.get(pk=usuario_id)
        except CustomUser.DoesNotExist:
            return HttpResponse("Funcionario no encontrado", status=404)
        
        es_sereno = (
            getattr(funcionario, 'funcion', None) == 'SERENO' or
            getattr(funcionario, 'tipo_funcionario', None) == 'SERENO'
        )
        
        horario = getattr(funcionario, 'horario', None)
        total_minutos = 0
        
        if es_sereno:
            total_minutos_s1 = 0
            total_minutos_s2 = 0
            dias_s1 = {}
            dias_s2 = {}
            
            if horario:
                for day in range(7):
                    dia_nombre = list(DIA_SEMANA_MAP.values())[day]

                    d1 = horario.dias.filter(dia_semana=day, semana_tipo=1).first()
                    if not d1:
                        d1 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()

                    d2 = horario.dias.filter(dia_semana=day, semana_tipo=2).first()
                    if not d2:
                        d2 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()

                    def calc_mins(dh):
                        if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                            h1, m1 = dh.hora_entrada.hour, dh.hora_entrada.minute
                            h2, m2 = dh.hora_salida.hour, dh.hora_salida.minute
                            min1 = h1 * 60 + m1
                            min2 = h2 * 60 + m2
                            if min2 < min1:
                                min2 += 24 * 60
                            return min2 - min1
                        return 0

                    def get_hora_str(dh):
                        if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                            return f"{dh.hora_entrada.strftime('%H:%M')} - {dh.hora_salida.strftime('%H:%M')}"
                        return "Libre"

                    total_minutos_s1 += calc_mins(d1)
                    total_minutos_s2 += calc_mins(d2)
                    dias_s1[dia_nombre] = get_hora_str(d1)
                    dias_s2[dia_nombre] = get_hora_str(d2)
            
            # Suma real del ciclo
            total_real = total_minutos_s1 + total_minutos_s2
            h_real = total_real // 60
            m_real = total_real % 60
            horas_reales_str = f"{h_real}h {m_real}m" if m_real > 0 else f"{h_real}h"
            if total_real == 0:
                horas_reales_str = "N/C"

            MAX = 44 * 60
            adj1 = total_minutos_s1
            adj2 = total_minutos_s2
            if adj1 > MAX:
                excess = adj1 - MAX
                adj1 = MAX
                adj2 += excess
            prom_min = (adj1 + adj2) // 2
            h_prom = prom_min // 60
            m_prom = prom_min % 60
            prom_ajustado_str = f"{h_prom}h {m_prom}m" if m_prom > 0 else f"{h_prom}h"
            
            horas_str = horas_reales_str
        else:
            dias_s1 = {}
            dias_s2 = {}
            dias_data = {i: "Libre" for i in range(7)}
            if horario:
                for d in horario.dias.all():
                    if d.activo and d.hora_entrada and d.hora_salida:
                        dias_data[d.dia_semana] = f"{d.hora_entrada.strftime('%H:%M')} - {d.hora_salida.strftime('%H:%M')}"
                        h1, m1 = d.hora_entrada.hour, d.hora_entrada.minute
                        h2, m2 = d.hora_salida.hour, d.hora_salida.minute
                        min1 = h1 * 60 + m1
                        min2 = h2 * 60 + m2
                        if min2 < min1:
                            min2 += 24 * 60
                        total_minutos += (min2 - min1)

            h_total = total_minutos // 60
            m_total = total_minutos % 60
            horas_str = f"{h_total}h {m_total}m" if m_total > 0 else f"{h_total}h"
            if total_minutos == 0:
                horas_str = "N/C"

        empleado_data = {
            'nombre': funcionario.get_full_name(),
            'run': funcionario.run,
            'cargo': funcionario.get_funcion_display() or funcionario.get_role_display(),
            'total_horas': horas_str
        }
        
        if es_sereno:
            dias_names = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
            empleado_data['es_sereno'] = True
            empleado_data['dias_s1'] = [dias_s1.get(dia, 'Libre') for dia in dias_names]
            empleado_data['dias_s2'] = [dias_s2.get(dia, 'Libre') for dia in dias_names]
        else:
            dias_data = {i: "Libre" for i in range(7)}
            if horario:
                for d in horario.dias.all():
                    if d.activo and d.hora_entrada and d.hora_salida:
                        dias_data[d.dia_semana] = f"{d.hora_entrada.strftime('%H:%M')} - {d.hora_salida.strftime('%H:%M')}"
            empleado_data['dias'] = [dias_data[i] for i in range(7)]

        html_string = render_to_string('reportes/pdf_horarios.html', {
            'empleados_data': [empleado_data],
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'director': CustomUser.objects.filter(role='DIRECTOR').first(),
            'es_individual': True,
            'es_landscape_force': es_sereno
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=horario_{funcionario.run}.pdf'
        response.write(result)
        return response


class ExportarHorariosPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar los horarios semanales de todos los funcionarios a PDF"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        from asistencia.models import HorarioFuncionario, DiaHorario
        
        # Obtener todos los funcionarios activos
        funcionarios = CustomUser.objects.filter(
            is_active=True,
            role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']
        ).order_by('first_name', 'last_name')

        empleados_data = []
        
        # Check if there are serenos
        tiene_serenos = funcionarios.filter(
            Q(funcion='SERENO') | Q(tipo_funcionario='SERENO')
        ).exists()
        
        DIA_NAMES = {0: 'Lun', 1: 'Mar', 2: 'Mié', 3: 'Jue', 4: 'Vie', 5: 'Sáb', 6: 'Dom'}

        for f in funcionarios:
            es_sereno = (
                getattr(f, 'funcion', None) == 'SERENO' or
                getattr(f, 'tipo_funcionario', None) == 'SERENO'
            )
            
            horario = getattr(f, 'horario', None)
            total_minutos = 0
            
            if tiene_serenos and es_sereno:
                # Build separate week schedules for serenos
                dias_s1 = {i: "Libre" for i in range(7)}
                dias_s2 = {i: "Libre" for i in range(7)}
                
                if horario:
                    # Robust lookup + raw per week + adjusted promedio
                    raw_s1 = 0
                    raw_s2 = 0
                    for day in range(7):
                        d1 = horario.dias.filter(dia_semana=day, semana_tipo=1).first()
                        if not d1:
                            d1 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()
                        
                        d2 = horario.dias.filter(dia_semana=day, semana_tipo=2).first()
                        if not d2:
                            d2 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()

                        def calc_mins(dh):
                            if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                                h1, m1 = dh.hora_entrada.hour, dh.hora_entrada.minute
                                h2, m2 = dh.hora_salida.hour, dh.hora_salida.minute
                                min1 = h1 * 60 + m1
                                min2 = h2 * 60 + m2
                                if min2 < min1: min2 += 24 * 60
                                return min2 - min1
                            return 0

                        def get_str(dh):
                            if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                                return f"{dh.hora_entrada.strftime('%H:%M')} - {dh.hora_salida.strftime('%H:%M')}"
                            return "Libre"

                        raw_s1 += calc_mins(d1)
                        raw_s2 += calc_mins(d2)

                        dias_s1[day] = get_str(d1)
                        dias_s2[day] = get_str(d2)
                    
                # Suma real del ciclo
                total_real = raw_s1 + raw_s2
                h_real = total_real // 60
                m_real = total_real % 60
                horas_reales_str = f"{h_real}h {m_real}m" if m_real > 0 else f"{h_real}h"
                if total_real == 0: horas_reales_str = "N/C"

                # Ajuste 44h + carry → promedio semanal
                MAX = 44 * 60
                adj1 = raw_s1
                adj2 = raw_s2
                if adj1 > MAX:
                    excess = adj1 - MAX
                    adj1 = MAX
                    adj2 += excess
                prom_min = (adj1 + adj2) // 2
                h_prom = prom_min // 60
                m_prom = prom_min % 60
                prom_ajustado_str = f"{h_prom}h {m_prom}m" if m_prom > 0 else f"{h_prom}h"

                empleados_data.append({
                    'nombre': f.get_full_name(),
                    'run': f.run,
                    'cargo': f.get_funcion_display() or f.get_role_display(),
                    'es_sereno': True,
                    'dias_s1': [dias_s1[i] for i in range(7)],
                    'dias_s2': [dias_s2[i] for i in range(7)],
                    'total_horas': horas_reales_str,           # suma real
                    'promedio_ajustado': prom_ajustado_str     # 44h + carry
                })
            else:
                # Normal format for non-serenos
                dias_list = []
                dias_data = {i: "Libre" for i in range(7)}
                total_minutos = 0
                
                if horario:
                    for d in horario.dias.all():
                        if d.activo and d.hora_entrada and d.hora_salida:
                            dias_data[d.dia_semana] = f"{d.hora_entrada.strftime('%H:%M')} - {d.hora_salida.strftime('%H:%M')}"
                            h1, m1 = d.hora_entrada.hour, d.hora_entrada.minute
                            h2, m2 = d.hora_salida.hour, d.hora_salida.minute
                            min1 = h1 * 60 + m1
                            min2 = h2 * 60 + m2
                            if min2 < min1: min2 += 24 * 60
                            total_minutos += (min2 - min1)

                h_total = total_minutos // 60
                m_total = total_minutos % 60
                horas_str = f"{h_total}h {m_total}m" if m_total > 0 else f"{h_total}h"
                if total_minutos == 0: horas_str = "N/C"

                empleados_data.append({
                    'nombre': f.get_full_name(),
                    'run': f.run,
                    'cargo': f.get_funcion_display() or f.get_role_display(),
                    'dias': [dias_data[i] for i in range(7)],
                    'total_horas': horas_str
                })

        html_string = render_to_string('reportes/pdf_horarios.html', {
            'empleados_data': empleados_data,
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'director': CustomUser.objects.filter(role='DIRECTOR').first(),
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=horarios_personal_{datetime.now().strftime("%Y%m%d")}.pdf'
        response.write(result)
        return response

class MiHorarioPDFView(LoginRequiredMixin, View):
    """Genera el PDF del horario semanal del usuario actualmente autenticado."""

    def get(self, request):
        f = request.user
        horario = getattr(f, 'horario', None)
        total_minutos = 0

        es_sereno = (
            getattr(f, 'funcion', None) == 'SERENO' or
            getattr(f, 'tipo_funcionario', None) == 'SERENO'
        )

        if es_sereno:
            dias_s1 = {}
            dias_s2 = {}
            raw_s1 = 0
            raw_s2 = 0

            if horario:
                for day in range(7):
                    dia_nombre = list(DIA_SEMANA_MAP.values())[day]

                    d1 = horario.dias.filter(dia_semana=day, semana_tipo=1).first()
                    if not d1:
                        d1 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()

                    d2 = horario.dias.filter(dia_semana=day, semana_tipo=2).first()
                    if not d2:
                        d2 = horario.dias.filter(dia_semana=day, semana_tipo=None).first()

                    def calc_mins(dh):
                        if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                            h1, m1 = dh.hora_entrada.hour, dh.hora_entrada.minute
                            h2, m2 = dh.hora_salida.hour, dh.hora_salida.minute
                            min1 = h1 * 60 + m1
                            min2 = h2 * 60 + m2
                            if min2 < min1: min2 += 24 * 60
                            return min2 - min1
                        return 0

                    def get_str(dh):
                        if dh and dh.activo and dh.hora_entrada and dh.hora_salida:
                            return f"{dh.hora_entrada.strftime('%H:%M')} - {dh.hora_salida.strftime('%H:%M')}"
                        return "Libre"

                    raw_s1 += calc_mins(d1)
                    raw_s2 += calc_mins(d2)

                    dias_s1[dia_nombre] = get_str(d1)
                    dias_s2[dia_nombre] = get_str(d2)

            # Suma real del ciclo
            total_real = raw_s1 + raw_s2
            h_real = total_real // 60
            m_real = total_real % 60
            horas_reales_str = f"{h_real}h {m_real}m" if m_real > 0 else f"{h_real}h"
            if total_real == 0:
                horas_reales_str = "N/C"

            # Promedio ajustado (44h + carry)
            MAX = 44 * 60
            adj1 = raw_s1
            adj2 = raw_s2
            if adj1 > MAX:
                excess = adj1 - MAX
                adj1 = MAX
                adj2 += excess
            prom_min = (adj1 + adj2) // 2
            h_prom = prom_min // 60
            m_prom = prom_min % 60
            prom_ajustado_str = f"{h_prom}h {m_prom}m" if m_prom > 0 else f"{h_prom}h"

            empleado_data = {
                'nombre': f.get_full_name(),
                'run': f.run,
                'cargo': f.get_funcion_display() or f.get_role_display(),
                'es_sereno': True,
                'dias_s1': [dias_s1.get(list(DIA_SEMANA_MAP.values())[day], 'Libre') for day in range(7)],
                'dias_s2': [dias_s2.get(list(DIA_SEMANA_MAP.values())[day], 'Libre') for day in range(7)],
                'total_horas': horas_reales_str,
                'promedio_ajustado': prom_ajustado_str
            }
        else:
            dias_data = {i: "Libre" for i in range(7)}
            if horario:
                for d in horario.dias.all():
                    if d.activo and d.hora_entrada and d.hora_salida:
                        dias_data[d.dia_semana] = f"{d.hora_entrada.strftime('%H:%M')} - {d.hora_salida.strftime('%H:%M')}"
                        h1, m1 = d.hora_entrada.hour, d.hora_entrada.minute
                        h2, m2 = d.hora_salida.hour, d.hora_salida.minute
                        min1 = h1 * 60 + m1
                        min2 = h2 * 60 + m2
                        if min2 < min1: min2 += 24 * 60
                        total_minutos += (min2 - min1)

            h_total = total_minutos // 60
            m_total = total_minutos % 60
            horas_str = f"{h_total}h {m_total}m" if m_total > 0 else f"{h_total}h"
            if total_minutos == 0: horas_str = "N/C"

            empleado_data = {
                'nombre': f.get_full_name(),
                'run': f.run,
                'cargo': f.get_funcion_display() or f.get_role_display(),
                'dias': [dias_data[i] for i in range(7)],
                'total_horas': horas_str
            }

        html_string = render_to_string('reportes/pdf_horarios.html', {
            'empleados_data': [empleado_data],
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'director': CustomUser.objects.filter(role='DIRECTOR').first(),
            'es_individual': True,
            'es_landscape_force': es_sereno
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=mi_horario_{f.run}.pdf'
        response.write(result)
        return response


class ExportarJustificacionesExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar Excel con justificaciones aprobadas"""

    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Justificaciones Aprobadas"

        # Headers
        headers = ['N°', 'Nombre Completo', 'RUN', 'Cargo', 'Fecha', 'Tipo', 'Hora de Llegada', 'Estado Alegación', 'Revisado Por', 'Justificación Manual']
        ws.append(headers)

        # Column widths
        for col, width in [('A', 10), ('B', 25), ('C', 15), ('D', 20), ('E', 15), ('F', 15), ('G', 30), ('H', 20), ('I', 25)]:
            ws.column_dimensions[col].width = width

        # Query alegaciones aprobadas
        alegaciones = AlegacionAsistencia.objects.filter(
            estado='APROBADA'
        ).select_related(
            'registro_asistencia__funcionario',
            'revisado_por'
        )

        # Query justificaciones manuales (registros justificados por admin)
        justificaciones_manuales = RegistroAsistencia.objects.filter(
            justificado_por__isnull=False
        ).select_related('funcionario', 'justificado_por')

        if year:
            alegaciones = alegaciones.filter(registro_asistencia__fecha__year=year)
            justificaciones_manuales = justificaciones_manuales.filter(fecha__year=year)
        if mes:
            alegaciones = alegaciones.filter(registro_asistencia__fecha__month=mes)
            justificaciones_manuales = justificaciones_manuales.filter(fecha__month=mes)

        # Combinar y ordenar todas las justificaciones
        todas_justificaciones = []

        # Agregar alegaciones aprobadas
        for alegacion in alegaciones.order_by('registro_asistencia__fecha', 'registro_asistencia__funcionario__first_name'):
            reg = alegacion.registro_asistencia
            tipo = "Atraso" if reg.estado == 'RETRASO' else "Inasistencia" if reg.estado == 'AUSENTE' else reg.get_estado_display()

            todas_justificaciones.append({
                'tipo_justificacion': 'alegacion',
                'registro': reg,
                'alegacion': alegacion,
                'tipo': tipo,
                'hora_llegada': reg.hora_entrada_real.strftime('%H:%M') if reg.hora_entrada_real else '-',
                'estado': alegacion.get_estado_display(),
                'revisado_por': alegacion.revisado_por.get_full_name() if alegacion.revisado_por else "",
                'justificacion_manual': reg.justificacion_manual or ""
            })

        # Agregar justificaciones manuales
        for reg in justificaciones_manuales.order_by('fecha', 'funcionario__first_name'):
            tipo = "Atraso" if reg.estado == 'RETRASO' else "Inasistencia" if reg.estado == 'AUSENTE' else reg.get_estado_display()

            todas_justificaciones.append({
                'tipo_justificacion': 'manual',
                'registro': reg,
                'alegacion': None,
                'tipo': tipo,
                'hora_llegada': reg.hora_entrada_real.strftime('%H:%M') if reg.hora_entrada_real else '-',
                'estado': "Justificado Manualmente",
                'revisado_por': reg.justificado_por.get_full_name() if reg.justificado_por else "",
                'justificacion_manual': reg.justificacion_manual or ""
            })

        # Ordenar todas las justificaciones por fecha y nombre
        todas_justificaciones.sort(key=lambda x: (x['registro'].fecha, x['registro'].funcionario.first_name))

        for i, just in enumerate(todas_justificaciones, 1):
            reg = just['registro']

            ws.append([
                i,
                reg.funcionario.get_full_name() or reg.funcionario.username,
                reg.funcionario.run,
                reg.funcionario.get_funcion_display() or "",
                reg.fecha.strftime("%d-%m-%Y") if reg.fecha else "",
                just['tipo'],
                just['hora_llegada'],
                just['estado'],
                just['revisado_por'],
                just['justificacion_manual'][:100] + "..." if just['justificacion_manual'] and len(just['justificacion_manual']) > 100 else just['justificacion_manual'] or ""
            ])

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"justificaciones_aprobadas"
        if mes and year:
            filename += f"_{mes}_{year}"
        elif year:
            filename += f"_{year}"
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

        wb.save(response)
        return response


class ExportarJustificacionesPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar PDF con justificaciones aprobadas"""

    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')

        # Query alegaciones aprobadas
        alegaciones = AlegacionAsistencia.objects.filter(
            estado='APROBADA'
        ).select_related(
            'registro_asistencia__funcionario',
            'revisado_por'
        )

        # Query justificaciones manuales (registros justificados por admin)
        justificaciones_manuales = RegistroAsistencia.objects.filter(
            justificado_por__isnull=False
        ).select_related('funcionario', 'justificado_por')

        if year:
            alegaciones = alegaciones.filter(registro_asistencia__fecha__year=year)
            justificaciones_manuales = justificaciones_manuales.filter(fecha__year=year)
        if mes:
            alegaciones = alegaciones.filter(registro_asistencia__fecha__month=mes)
            justificaciones_manuales = justificaciones_manuales.filter(fecha__month=mes)

        # Preparar datos para el template
        justificaciones = []

        # Agregar alegaciones aprobadas
        for alegacion in alegaciones.order_by('registro_asistencia__fecha', 'registro_asistencia__funcionario__first_name'):
            reg = alegacion.registro_asistencia
            tipo = "Atraso" if reg.estado == 'RETRASO' else "Inasistencia" if reg.estado == 'AUSENTE' else reg.get_estado_display()

            justificaciones.append({
                'nombre_completo': reg.funcionario.get_full_name() or reg.funcionario.username,
                'run': reg.funcionario.run,
                'cargo': reg.funcionario.get_funcion_display() or "",
                'fecha': reg.fecha.strftime("%d-%m-%Y") if reg.fecha else "",
                'tipo': tipo,
                'hora_llegada': reg.hora_entrada_real.strftime('%H:%M') if reg.hora_entrada_real else '-',
                'estado': alegacion.get_estado_display(),
                'revisado_por': alegacion.revisado_por.get_full_name() if alegacion.revisado_por else "",
                'justificacion_manual': reg.justificacion_manual[:200] + "..." if reg.justificacion_manual and len(reg.justificacion_manual) > 200 else reg.justificacion_manual or ""
            })

        # Agregar justificaciones manuales
        for reg in justificaciones_manuales.order_by('fecha', 'funcionario__first_name'):
            tipo = "Atraso" if reg.estado == 'RETRASO' else "Inasistencia" if reg.estado == 'AUSENTE' else reg.get_estado_display()

            justificaciones.append({
                'nombre_completo': reg.funcionario.get_full_name() or reg.funcionario.username,
                'run': reg.funcionario.run,
                'cargo': reg.funcionario.get_funcion_display() or "",
                'fecha': reg.fecha.strftime("%d-%m-%Y") if reg.fecha else "",
                'tipo': tipo,
                'hora_llegada': reg.hora_entrada_real.strftime('%H:%M') if reg.hora_entrada_real else '-',
                'estado': "Justificado Manualmente",
                'revisado_por': reg.justificado_por.get_full_name() if reg.justificado_por else "",
                'justificacion_manual': reg.justificacion_manual or ""
            })

        # Ordenar todas las justificaciones por fecha y nombre
        justificaciones.sort(key=lambda x: (x['fecha'], x['nombre_completo']))

        html_string = render_to_string('reportes/pdf_justificaciones.html', {
            'justificaciones': justificaciones,
            'fecha_exportacion': now().strftime('%d/%m/%Y %H:%M'),
            'year': year,
            'mes': mes
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        filename = f"justificaciones_aprobadas"
        if mes and year:
            filename += f"_{mes}_{year}"
        elif year:
            filename += f"_{year}"
        response['Content-Disposition'] = f'inline; filename={filename}.pdf'
        response.write(result)
        return response



