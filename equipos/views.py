from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings
from weasyprint import HTML
from django.db.models import Prefetch, Q
from .models import Equipo, PrestamoEquipo, FallaEquipo, LugarEquipo, HitoMantenimiento, TicketBitacora
from users.models import CustomUser
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from admin_dashboard.utils import registrar_log, get_client_ip

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}


def _sanitizar_texto(texto):
    if not texto:
        return ""
    return str(texto).replace("{{", "").replace("}}", "").replace("{%", "").replace("%}", "").strip()


# ==========================================
# 1. INVENTARIO DE EQUIPOS
# ==========================================

@login_required
def inventario_equipos(request):
    """Módulo exclusivo para la gestión integral del inventario de equipos"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    equipos_base = Equipo.objects.all().prefetch_related(
        Prefetch(
            "prestamos",
            queryset=PrestamoEquipo.objects.filter(activo=True).select_related("funcionario"),
            to_attr="prestamo_activo_list"
        )
    ).select_related("lugar")

    stats = {
        "total": equipos_base.count(),
        "disponibles": equipos_base.filter(estado="DISPONIBLE").count(),
        "asignados": equipos_base.filter(estado="ASIGNADO").count(),
        "reparacion": equipos_base.filter(estado="EN_REPARACION").count(),
        "baja": equipos_base.filter(estado="BAJA").count(),
    }

    tipo = request.GET.get("tipo")
    estado = request.GET.get("estado")
    lugar_id = request.GET.get("lugar")
    search = request.GET.get("search")
    orden = request.GET.get("orden", "marca")

    equipos = equipos_base
    if tipo:
        equipos = equipos.filter(tipo=tipo)
    if estado:
        equipos = equipos.filter(estado=estado)
    if lugar_id:
        equipos = equipos.filter(lugar_id=lugar_id)
    if search:
        equipos = equipos.filter(
            Q(marca__icontains=search) |
            Q(modelo__icontains=search) |
            Q(numero_serie__icontains=search) |
            Q(numero_inventario__icontains=search) |
            Q(lugar__nombre__icontains=search)
        )

    orden_mapping = {
        "marca": "marca",
        "-marca": "-marca",
        "tipo": "tipo",
        "-tipo": "-tipo",
        "estado": "estado",
        "-estado": "-estado",
        "fecha_origen": "fecha_origen",
        "-fecha_origen": "-fecha_origen",
        "lugar": "lugar__nombre",
        "-lugar": "-lugar__nombre",
    }
    orden_query = orden_mapping.get(orden, "marca")
    equipos = equipos.order_by(orden_query)

    for eq in equipos:
        eq.marca = _sanitizar_texto(eq.marca)
        eq.modelo = _sanitizar_texto(eq.modelo)
        eq.numero_serie = _sanitizar_texto(eq.numero_serie)
        eq.numero_inventario = _sanitizar_texto(eq.numero_inventario)
        if eq.lugar:
            eq.lugar.nombre = _sanitizar_texto(eq.lugar.nombre)
        eq.prestamo_activo = eq.prestamo_activo_list[0] if eq.prestamo_activo_list else None

    context = {
        "equipos": equipos,
        "stats": stats,
        "tipos": Equipo.TIPO_CHOICES,
        "estados": Equipo.ESTADO_CHOICES,
        "lugares": LugarEquipo.objects.filter(activo=True).order_by("nombre"),
        "filtros_activos": bool(tipo or estado or lugar_id or search),
        "selected_tipo": tipo,
        "selected_estado": estado,
        "selected_lugar": lugar_id,
        "search": search,
        "orden": orden,
        "ordenes": [
            ("marca", "Marca A-Z"),
            ("-marca", "Marca Z-A"),
            ("tipo", "Tipo A-Z"),
            ("-tipo", "Tipo Z-A"),
            ("estado", "Estado A-Z"),
            ("-estado", "Estado Z-A"),
            ("fecha_origen", "Fecha Origen más antigua"),
            ("-fecha_origen", "Fecha Origen más reciente"),
            ("lugar", "Lugar A-Z"),
            ("-lugar", "Lugar Z-A"),
        ],
    }
    return render(request, "equipos/inventario_equipos.html", context)


# ==========================================
# 2. GESTIÓN DE EQUIPOS Y ASIGNACIONES
# ==========================================

@login_required
def lista_equipos(request):
    """Gestión de Equipos por Funcionario y Préstamos Rápidos (solo administradores)"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    funcionario_id = request.GET.get("funcionario_id")
    funcionarios = CustomUser.objects.filter(is_active=True).order_by("first_name", "last_name")

    selected_funcionario = None
    equipos_funcionario = []

    if funcionario_id:
        selected_funcionario = get_object_or_404(CustomUser, id=funcionario_id)
        prestamos_activos = PrestamoEquipo.objects.filter(
            funcionario=selected_funcionario, 
            activo=True
        ).select_related("equipo", "equipo__lugar").order_by("-fecha_asignacion")
        
        for p in prestamos_activos:
            eq = p.equipo
            eq.prestamo_activo = p
            eq.marca = _sanitizar_texto(eq.marca)
            eq.modelo = _sanitizar_texto(eq.modelo)
            eq.numero_serie = _sanitizar_texto(eq.numero_serie)
            eq.numero_inventario = _sanitizar_texto(eq.numero_inventario)
            equipos_funcionario.append(eq)

    prestamos_diarios_activos = PrestamoEquipo.objects.filter(
        activo=True, 
        es_prestamo_diario=True
    ).select_related("equipo", "funcionario", "equipo__lugar").order_by("-fecha_asignacion")

    equipos_disponibles = Equipo.objects.filter(estado="DISPONIBLE").order_by("tipo", "marca", "modelo")

    context = {
        "funcionarios": funcionarios,
        "selected_funcionario": selected_funcionario,
        "equipos_funcionario": equipos_funcionario,
        "prestamos_diarios_activos": prestamos_diarios_activos,
        "equipos_disponibles": equipos_disponibles,
        "total_equipos_asignados": len(equipos_funcionario),
        "total_prestamos_diarios": prestamos_diarios_activos.count(),
    }
    return render(request, "equipos/lista_equipos.html", context)


# ==========================================
# 3. SISTEMA DE PRÉSTAMOS DIARIOS
# ==========================================

@login_required
def prestamos_diarios(request):
    """Panel de control simple y ligero de Préstamos Diarios con filtro por mes"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    hoy = timezone.now().date()
    mes_filtro = request.GET.get("mes")
    
    if not mes_filtro:
        mes_filtro = hoy.strftime("%Y-%m")

    try:
        anio_sel, mes_sel = map(int, mes_filtro.split("-"))
    except ValueError:
        anio_sel, mes_sel = hoy.year, hoy.month
        mes_filtro = f"{anio_sel:04d}-{mes_sel:02d}"

    prestamos_activos_hoy = PrestamoEquipo.objects.filter(
        activo=True,
        es_prestamo_diario=True
    ).select_related("equipo", "funcionario", "asignado_por").order_by("-fecha_asignacion")

    prestamos_mes = PrestamoEquipo.objects.filter(
        es_prestamo_diario=True,
        fecha_asignacion__year=anio_sel,
        fecha_asignacion__month=mes_sel
    ).select_related("equipo", "funcionario", "asignado_por").order_by("-fecha_asignacion")

    total_mes = prestamos_mes.count()
    devueltos_mes = prestamos_mes.filter(activo=False).count()
    activos_mes = prestamos_mes.filter(activo=True).count()

    fechas_prestamos = PrestamoEquipo.objects.filter(
        es_prestamo_diario=True
    ).dates("fecha_asignacion", "month", order="DESC")
    
    meses_disponibles = []
    meses_set = {mes_filtro, hoy.strftime("%Y-%m")}
    for f in fechas_prestamos:
        meses_set.add(f.strftime("%Y-%m"))

    for m_str in sorted(list(meses_set), reverse=True):
        y, m = map(int, m_str.split("-"))
        meses_disponibles.append({
            "codigo": m_str,
            "nombre": f"{MESES_ES.get(m, '')} {y}",
            "seleccionado": m_str == mes_filtro
        })

    funcionarios = CustomUser.objects.filter(is_active=True).order_by("first_name", "last_name")
    equipos_disponibles = Equipo.objects.filter(estado="DISPONIBLE").order_by("tipo", "marca", "modelo")

    context = {
        "prestamos_activos_hoy": prestamos_activos_hoy,
        "prestamos_mes": prestamos_mes,
        "mes_seleccionado": mes_filtro,
        "nombre_mes_seleccionado": f"{MESES_ES.get(mes_sel, '')} {anio_sel}",
        "meses_disponibles": meses_disponibles,
        "total_mes": total_mes,
        "devueltos_mes": devueltos_mes,
        "activos_mes": activos_mes,
        "funcionarios": funcionarios,
        "equipos_disponibles": equipos_disponibles,
    }
    return render(request, "equipos/prestamos_diarios.html", context)


@login_required
def crear_prestamo_diario(request):
    """Crea un préstamo diario de implementos a un funcionario"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect("dashboard")

    if request.method == "POST":
        funcionario_id = request.POST.get("funcionario_id")
        equipo_id = request.POST.get("equipo_id")
        observaciones = request.POST.get("observaciones", "").strip()

        if not funcionario_id or not equipo_id:
            messages.error(request, "Debes seleccionar tanto un funcionario como un equipo.")
            return redirect(request.META.get("HTTP_REFERER", "prestamos_diarios"))

        funcionario = get_object_or_404(CustomUser, id=funcionario_id)
        equipo = get_object_or_404(Equipo, id=equipo_id)

        if equipo.estado != "DISPONIBLE":
            messages.warning(request, f"El equipo {equipo} no se encuentra disponible.")
            return redirect(request.META.get("HTTP_REFERER", "prestamos_diarios"))

        prestamo = PrestamoEquipo.objects.create(
            equipo=equipo,
            funcionario=funcionario,
            observaciones=observaciones,
            asignado_por=request.user,
            es_prestamo_diario=True,
            activo=True
        )

        registrar_log(
            usuario=request.user,
            tipo="CREATE",
            accion="Préstamo Diario de Equipo",
            descripcion=f"Préstamo diario: {equipo} entregado a {funcionario.get_full_name()}",
            ip_address=get_client_ip(request)
        )
        messages.success(request, f"Préstamo diario registrado: {equipo} entregado a {funcionario.get_full_name()}.")

    return redirect(request.META.get("HTTP_REFERER", "prestamos_diarios"))


@login_required
def terminar_prestamo_diario(request, prestamo_id):
    """Finaliza un préstamo diario y devuelve el equipo a DISPONIBLE"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect("dashboard")

    prestamo = get_object_or_404(PrestamoEquipo, id=prestamo_id)
    prestamo.activo = False
    prestamo.fecha_devolucion = timezone.now().date()
    prestamo.fecha_devolucion_real = timezone.now()
    prestamo.save()

    registrar_log(
        usuario=request.user,
        tipo="UPDATE",
        accion="Devolución Préstamo Diario",
        descripcion=f"Devolución de préstamo diario: {prestamo.equipo} por {prestamo.funcionario.get_full_name()}",
        ip_address=get_client_ip(request)
    )
    messages.success(request, f"Préstamo de {prestamo.equipo} terminado exitosamente. Equipo nuevamente disponible.")
    return redirect(request.META.get("HTTP_REFERER", "prestamos_diarios"))


@login_required
def reporte_prestamos_diarios_pdf(request):
    """Genera informe PDF de préstamos diarios ordenado y filtrado por mes"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a este reporte.")
        return redirect("dashboard")

    hoy = timezone.now().date()
    mes_filtro = request.GET.get("mes") or hoy.strftime("%Y-%m")

    try:
        anio_sel, mes_sel = map(int, mes_filtro.split("-"))
    except ValueError:
        anio_sel, mes_sel = hoy.year, hoy.month

    prestamos = PrestamoEquipo.objects.filter(
        es_prestamo_diario=True,
        fecha_asignacion__year=anio_sel,
        fecha_asignacion__month=mes_sel
    ).select_related("equipo", "funcionario", "asignado_por").order_by("fecha_asignacion")

    nombre_mes = f"{MESES_ES.get(mes_sel, '')} {anio_sel}"
    titulo = f"Reporte de Préstamos Diarios de Equipamiento - {nombre_mes}"

    html_string = render(request, "equipos/prestamos_diarios_pdf.html", {
        "prestamos": prestamos,
        "nombre_mes": nombre_mes,
        "titulo": titulo,
        "total": prestamos.count(),
        "devueltos": prestamos.filter(activo=False).count(),
        "activos": prestamos.filter(activo=True).count(),
        "fecha_emision": timezone.now(),
        "generado_por": request.user,
    }).content.decode("utf-8")

    pdf = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    filename = f"prestamos_diarios_{anio_sel}_{mes_sel:02d}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ==========================================
# 4. SISTEMA DE TICKETS Y BITÁCORA GENERAL TI
# ==========================================

@login_required
def gestion_tickets_bitacora(request):
    """Panel general de Bitácora TI, Tickets de soporte y averías reportadas"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    hoy = timezone.now().date()
    mes_filtro = request.GET.get("mes")
    
    if not mes_filtro:
        mes_filtro = hoy.strftime("%Y-%m")

    try:
        anio_sel, mes_sel = map(int, mes_filtro.split("-"))
    except ValueError:
        anio_sel, mes_sel = hoy.year, hoy.month
        mes_filtro = f"{anio_sel:04d}-{mes_sel:02d}"

    categoria_filtro = request.GET.get("categoria")
    estado_filtro = request.GET.get("estado")
    search = request.GET.get("search")

    tickets_base = TicketBitacora.objects.filter(
        fecha_actividad__year=anio_sel,
        fecha_actividad__month=mes_sel
    ).select_related("funcionario", "lugar", "equipo", "creado_por")

    if categoria_filtro:
        tickets_base = tickets_base.filter(categoria=categoria_filtro)
    if estado_filtro:
        tickets_base = tickets_base.filter(estado=estado_filtro)
    if search:
        tickets_base = tickets_base.filter(
            Q(titulo__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(resolucion__icontains=search) |
            Q(funcionario__first_name__icontains=search) |
            Q(funcionario__last_name__icontains=search) |
            Q(lugar__nombre__icontains=search) |
            Q(lugar_personalizado__icontains=search)
        )

    tickets = tickets_base.order_by("-fecha_actividad", "-fecha_creacion")

    stats_tickets = {
        "total": TicketBitacora.objects.filter(fecha_actividad__year=anio_sel, fecha_actividad__month=mes_sel).count(),
        "resueltos": TicketBitacora.objects.filter(fecha_actividad__year=anio_sel, fecha_actividad__month=mes_sel, estado="RESUELTO").count(),
        "en_proceso": TicketBitacora.objects.filter(fecha_actividad__year=anio_sel, fecha_actividad__month=mes_sel, estado="EN_PROCESO").count(),
        "pendientes": TicketBitacora.objects.filter(fecha_actividad__year=anio_sel, fecha_actividad__month=mes_sel, estado="PENDIENTE").count(),
    }

    fallas = FallaEquipo.objects.select_related("equipo", "funcionario").order_by("-fecha_reporte")
    fallas_reportadas = fallas.filter(estado="REPORTADA")
    fallas_en_revision = fallas.filter(estado="EN_REVISION")
    fallas_reparadas = fallas.filter(estado="REPARADA")

    fechas_tickets = TicketBitacora.objects.dates("fecha_actividad", "month", order="DESC")
    meses_set = {mes_filtro, hoy.strftime("%Y-%m")}
    for f in fechas_tickets:
        meses_set.add(f.strftime("%Y-%m"))

    meses_disponibles = []
    for m_str in sorted(list(meses_set), reverse=True):
        y, m = map(int, m_str.split("-"))
        meses_disponibles.append({
            "codigo": m_str,
            "nombre": f"{MESES_ES.get(m, '')} {y}",
            "seleccionado": m_str == mes_filtro
        })

    funcionarios = CustomUser.objects.filter(is_active=True).order_by("first_name", "last_name")
    lugares = LugarEquipo.objects.filter(activo=True).order_by("nombre")
    equipos = Equipo.objects.all().order_by("tipo", "marca", "modelo")

    context = {
        "tickets": tickets,
        "stats_tickets": stats_tickets,
        "mes_seleccionado": mes_filtro,
        "nombre_mes_seleccionado": f"{MESES_ES.get(mes_sel, '')} {anio_sel}",
        "meses_disponibles": meses_disponibles,
        "categoria_filtro": categoria_filtro,
        "estado_filtro": estado_filtro,
        "search": search,
        "CATEGORIA_CHOICES": TicketBitacora.CATEGORIA_CHOICES,
        "ESTADO_CHOICES": TicketBitacora.ESTADO_CHOICES,
        "PRIORIDAD_CHOICES": TicketBitacora.PRIORIDAD_CHOICES,
        "funcionarios": funcionarios,
        "lugares": lugares,
        "equipos": equipos,
        "hoy": hoy,
        "fallas": fallas,
        "fallas_reportadas": fallas_reportadas,
        "fallas_en_revision": fallas_en_revision,
        "fallas_reparadas": fallas_reparadas,
        "total_fallas": fallas.count(),
        "total_reportadas": fallas_reportadas.count(),
        "total_en_revision": fallas_en_revision.count(),
        "total_reparadas": fallas_reparadas.count(),
        "ESTADO_FALLA_CHOICES": FallaEquipo.ESTADO_FALLA_CHOICES,
    }
    return render(request, "equipos/gestion_fallas.html", context)


@login_required
def crear_ticket_bitacora(request):
    """Registra una nueva actividad o ticket en la bitácora TI"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect("dashboard")

    if request.method == "POST":
        try:
            titulo = request.POST.get("titulo", "").strip()
            descripcion = request.POST.get("descripcion", "").strip()
            categoria = request.POST.get("categoria", "SOPORTE_FUNCIONARIO")
            lugar_id = request.POST.get("lugar_id") or None
            lugar_personalizado = request.POST.get("lugar_personalizado", "").strip()
            funcionario_id = request.POST.get("funcionario_id") or None
            equipo_id = request.POST.get("equipo_id") or None
            falla_id = request.POST.get("falla_id") or None
            estado = request.POST.get("estado", "RESUELTO")
            prioridad = request.POST.get("prioridad", "NORMAL")
            fecha_actividad = request.POST.get("fecha_actividad") or timezone.now().date()
            hora_actividad = request.POST.get("hora_actividad") or None
            resolucion = request.POST.get("resolucion", "").strip()

            if not titulo or not descripcion:
                messages.error(request, "El título y la descripción de la actividad son obligatorios.")
                return redirect("gestion_tickets_bitacora")

            ticket = TicketBitacora.objects.create(
                titulo=titulo,
                descripcion=descripcion,
                categoria=categoria,
                lugar_id=lugar_id,
                lugar_personalizado=lugar_personalizado,
                funcionario_id=funcionario_id,
                equipo_id=equipo_id,
                falla_asociada_id=falla_id,
                estado=estado,
                prioridad=prioridad,
                fecha_actividad=fecha_actividad,
                hora_actividad=hora_actividad,
                resolucion=resolucion,
                creado_por=request.user
            )

            if falla_id:
                falla = FallaEquipo.objects.filter(id=falla_id).first()
                if falla:
                    if estado == "RESUELTO":
                        falla.estado = "REPARADA"
                    elif estado == "EN_PROCESO":
                        falla.estado = "EN_REVISION"
                    falla.comentarios_tecnicos = resolucion or descripcion
                    falla.resuelto_por = request.user
                    falla.save()

            registrar_log(
                usuario=request.user,
                tipo="CREATE",
                accion="Registro de Ticket / Bitácora TI",
                descripcion=f"Se registró ticket TI: {ticket.titulo} ({ticket.get_categoria_display()})",
                ip_address=get_client_ip(request)
            )
            messages.success(request, f"Actividad '{ticket.titulo}' registrada en la bitácora correctamente.")
        except Exception as e:
            messages.error(request, f"Error al registrar ticket: {str(e)}")

    return redirect("gestion_tickets_bitacora")


@login_required
def editar_ticket_bitacora(request, ticket_id):
    """Actualiza un ticket o actividad de la bitácora TI"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect("dashboard")

    ticket = get_object_or_404(TicketBitacora, id=ticket_id)

    if request.method == "POST":
        ticket.titulo = request.POST.get("titulo", "").strip() or ticket.titulo
        ticket.descripcion = request.POST.get("descripcion", "").strip() or ticket.descripcion
        ticket.categoria = request.POST.get("categoria", ticket.categoria)
        ticket.lugar_id = request.POST.get("lugar_id") or None
        ticket.lugar_personalizado = request.POST.get("lugar_personalizado", "").strip()
        ticket.funcionario_id = request.POST.get("funcionario_id") or None
        ticket.equipo_id = request.POST.get("equipo_id") or None
        ticket.estado = request.POST.get("estado", ticket.estado)
        ticket.prioridad = request.POST.get("prioridad", ticket.prioridad)
        if request.POST.get("fecha_actividad"):
            ticket.fecha_actividad = request.POST.get("fecha_actividad")
        if request.POST.get("hora_actividad"):
            ticket.hora_actividad = request.POST.get("hora_actividad")
        ticket.resolucion = request.POST.get("resolucion", "").strip()
        ticket.save()

        registrar_log(
            usuario=request.user,
            tipo="UPDATE",
            accion="Actualización de Ticket / Bitácora TI",
            descripcion=f"Se actualizó ticket TI: {ticket.titulo}",
            ip_address=get_client_ip(request)
        )
        messages.success(request, f"Ticket '{ticket.titulo}' actualizado correctamente.")

    return redirect("gestion_tickets_bitacora")


@login_required
def eliminar_ticket_bitacora(request, ticket_id):
    """Elimina un ticket de la bitácora TI (solo administradores)"""
    if request.user.role != 'ADMIN':
        messages.error(request, "No tienes permisos para eliminar tickets.")
        return redirect("gestion_tickets_bitacora")

    ticket = get_object_or_404(TicketBitacora, id=ticket_id)
    if request.method == "POST":
        titulo = ticket.titulo
        ticket.delete()
        registrar_log(
            usuario=request.user,
            tipo="DELETE",
            accion="Eliminación de Ticket TI",
            descripcion=f"Se eliminó ticket TI: {titulo}",
            ip_address=get_client_ip(request)
        )
        messages.success(request, f"Ticket '{titulo}' eliminado correctamente.")

    return redirect("gestion_tickets_bitacora")


@login_required
def reporte_bitacora_pdf(request):
    """Genera informe oficial PDF de la Bitácora General de Actividades TI por mes"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a este reporte.")
        return redirect("dashboard")

    hoy = timezone.now().date()
    mes_filtro = request.GET.get("mes") or hoy.strftime("%Y-%m")

    try:
        anio_sel, mes_sel = map(int, mes_filtro.split("-"))
    except ValueError:
        anio_sel, mes_sel = hoy.year, hoy.month

    tickets = TicketBitacora.objects.filter(
        fecha_actividad__year=anio_sel,
        fecha_actividad__month=mes_sel
    ).select_related("funcionario", "lugar", "equipo", "creado_por").order_by("fecha_actividad", "hora_actividad")

    nombre_mes = f"{MESES_ES.get(mes_sel, '')} {anio_sel}"
    titulo = f"Bitácora Mensual de Actividades y Soporte TI - {nombre_mes}"

    html_string = render(request, "equipos/reporte_bitacora_pdf.html", {
        "tickets": tickets,
        "nombre_mes": nombre_mes,
        "titulo": titulo,
        "total": tickets.count(),
        "resueltos": tickets.filter(estado="RESUELTO").count(),
        "en_proceso": tickets.filter(estado="EN_PROCESO").count(),
        "pendientes": tickets.filter(estado="PENDIENTE").count(),
        "fecha_emision": timezone.now(),
        "generado_por": request.user,
    }).content.decode("utf-8")

    pdf = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    filename = f"bitacora_ti_{anio_sel}_{mes_sel:02d}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ==========================================
# 5. CRUD DE EQUIPOS Y LUGARES
# ==========================================

@login_required
def crear_equipo(request):
    """Crear nuevo equipo (solo administradores)"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    if request.method == "POST":
        try:
            funcionario_id = request.POST.get("funcionario")
            estado_inicial = "ASIGNADO" if funcionario_id else "DISPONIBLE"

            equipo = Equipo.objects.create(
                tipo=request.POST.get("tipo"),
                marca=request.POST.get("marca"),
                modelo=request.POST.get("modelo"),
                numero_serie=request.POST.get("numero_serie", "").upper(),
                numero_inventario=request.POST.get("numero_inventario", "").upper(),
                observaciones=request.POST.get("observaciones", ""),
                estado=estado_inicial,
                fecha_adquisicion=request.POST.get("fecha_adquisicion") or None,
                fecha_origen=request.POST.get("fecha_origen") or None,
                lugar_id=request.POST.get("lugar") or None,
                creado_por=request.user
            )

            if funcionario_id:
                funcionario = CustomUser.objects.get(id=funcionario_id)
                PrestamoEquipo.objects.create(
                    equipo=equipo,
                    funcionario=funcionario,
                    asignado_por=request.user,
                    activo=True
                )

            registrar_log(
                usuario=request.user,
                tipo="CREATE",
                accion="Creación de Equipo",
                descripcion=f"Se creó equipo {equipo.get_tipo_display()} {equipo.marca} {equipo.modelo} (Inv: {equipo.numero_inventario})",
                ip_address=get_client_ip(request)
            )
            messages.success(request, f"Equipo {equipo} creado exitosamente.")
            return redirect("inventario_equipos")
        except Exception as e:
            messages.error(request, f"Error al crear equipo: {str(e)}")

    funcionarios = CustomUser.objects.filter(is_active=True).order_by("first_name", "last_name")
    lugares = LugarEquipo.objects.filter(activo=True).order_by("nombre")

    context = {
        "tipos": Equipo.TIPO_CHOICES,
        "estados": Equipo.ESTADO_CHOICES,
        "lugares": lugares,
        "funcionarios": funcionarios,
    }
    return render(request, "equipos/crear_equipo.html", context)


@login_required
def crear_lugar_equipo(request):
    """Crear lugar de equipo (solo administradores)"""
    if request.user.role != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("inventario_equipos")
    
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip().upper()
        descripcion = request.POST.get("descripcion", "").strip()
        
        if not nombre:
            messages.error(request, "El nombre del lugar es obligatorio.")
            return redirect("inventario_equipos")
        
        lugar, created = LugarEquipo.objects.get_or_create(
            nombre=nombre,
            defaults={"descripcion": descripcion, "creado_por": request.user}
        )
        
        if created:
            messages.success(request, f"Lugar '{lugar.nombre}' creado exitosamente.")
            registrar_log(
                usuario=request.user,
                tipo="CREATE",
                accion="Creación de Lugar de Equipo",
                descripcion=f"Se creó lugar de equipo: {lugar.nombre}",
                ip_address=get_client_ip(request)
            )
        else:
            messages.info(request, f"El lugar '{lugar.nombre}' ya existe.")
        
    return redirect("inventario_equipos")


@login_required
def editar_lugar_equipo(request, lugar_id):
    """Editar lugar de equipo (solo administradores)"""
    if request.user.role != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("inventario_equipos")
    
    lugar = get_object_or_404(LugarEquipo, id=lugar_id)
    
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip().upper()
        descripcion = request.POST.get("descripcion", "").strip()
        activo = request.POST.get("activo") == "on"
        
        if not nombre:
            messages.error(request, "El nombre del lugar es obligatorio.")
            return redirect("inventario_equipos")
        
        lugar.nombre = nombre
        lugar.descripcion = descripcion
        lugar.activo = activo
        lugar.save()
        
        messages.success(request, f"Lugar '{lugar.nombre}' actualizado exitosamente.")
        registrar_log(
            usuario=request.user,
            tipo="UPDATE",
            accion="Actualización de Lugar de Equipo",
            descripcion=f"Se actualizó lugar de equipo: {lugar.nombre}",
            ip_address=get_client_ip(request)
        )
        
    return redirect("inventario_equipos")


@login_required
def eliminar_lugar_equipo(request, lugar_id):
    """Eliminar/Desactivar lugar de equipo (solo administradores)"""
    if request.user.role != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("inventario_equipos")
    
    lugar = get_object_or_404(LugarEquipo, id=lugar_id)
    
    if request.method == "POST":
        lugar.activo = False
        lugar.save()
        
        messages.success(request, f"Lugar '{lugar.nombre}' desactivado exitosamente.")
        registrar_log(
            usuario=request.user,
            tipo="UPDATE",
            accion="Desactivación de Lugar de Equipo",
            descripcion=f"Se desactivó lugar de equipo: {lugar.nombre}",
            ip_address=get_client_ip(request)
        )
    
    return redirect("inventario_equipos")


@login_required
def editar_equipo(request, equipo_id):
    """Editar equipo (solo administradores)"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    try:
        equipo_id = int(equipo_id)
    except (ValueError, TypeError):
        messages.error(request, "ID de equipo inválido.")
        return redirect("inventario_equipos")

    equipo = get_object_or_404(Equipo, id=equipo_id)

    if request.method == "POST":
        equipo.tipo = request.POST.get("tipo")
        equipo.marca = request.POST.get("marca")
        equipo.modelo = request.POST.get("modelo")
        equipo.numero_serie = request.POST.get("numero_serie", "").upper()
        equipo.numero_inventario = request.POST.get("numero_inventario", "").upper()
        equipo.observaciones = request.POST.get("observaciones", "")
        equipo.estado = request.POST.get("estado")
        equipo.fecha_adquisicion = request.POST.get("fecha_adquisicion") or None
        equipo.fecha_origen = request.POST.get("fecha_origen") or None
        equipo.lugar_id = request.POST.get("lugar") or None
        equipo.save()

        funcionario_id = request.POST.get("funcionario")
        prestamo_actual = equipo.prestamos.filter(activo=True).first()

        if funcionario_id:
            if not prestamo_actual or prestamo_actual.funcionario.id != int(funcionario_id):
                if prestamo_actual:
                    prestamo_actual.activo = False
                    prestamo_actual.fecha_devolucion = timezone.now().date()
                    prestamo_actual.save()
                
                PrestamoEquipo.objects.create(
                    equipo=equipo,
                    funcionario_id=funcionario_id,
                    asignado_por=request.user,
                    activo=True
                )
                equipo.estado = "ASIGNADO"
                equipo.save()
        else:
            if prestamo_actual:
                prestamo_actual.activo = False
                prestamo_actual.fecha_devolucion = timezone.now().date()
                prestamo_actual.save()
                
                if equipo.estado == "ASIGNADO":
                    equipo.estado = "DISPONIBLE"
                    equipo.save()

        registrar_log(
            usuario=request.user,
            tipo="UPDATE",
            accion="Actualización de Equipo",
            descripcion=f"Se actualizó equipo {equipo.get_tipo_display()} {equipo.marca} {equipo.modelo} (Inv: {equipo.numero_inventario})",
            ip_address=get_client_ip(request)
        )
        messages.success(request, "Equipo actualizado exitosamente.")
        return redirect("inventario_equipos")

    prestamo_actual = equipo.prestamos.filter(activo=True).first()
    funcionarios = CustomUser.objects.filter(is_active=True).order_by("first_name", "last_name")
    lugares = LugarEquipo.objects.filter(activo=True).order_by("nombre")

    context = {
        "equipo": equipo,
        "tipos": Equipo.TIPO_CHOICES,
        "estados": Equipo.ESTADO_CHOICES,
        "lugares": lugares,
        "funcionarios": funcionarios,
        "prestamo_actual": prestamo_actual,
    }
    return render(request, "equipos/editar_equipo.html", context)


@login_required
def eliminar_equipo(request, equipo_id):
    """Eliminar equipo (solo administradores)"""
    if request.user.role != 'ADMIN':
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    equipo = get_object_or_404(Equipo, id=equipo_id)

    if request.method == "POST":
        registrar_log(
            usuario=request.user,
            tipo="DELETE",
            accion="Eliminación de Equipo",
            descripcion=f"Se eliminó equipo {equipo.get_tipo_display()} {equipo.marca} {equipo.modelo} (Inv: {equipo.numero_inventario})",
            ip_address=get_client_ip(request)
        )
        equipo.delete()
        messages.success(request, "Equipo eliminado exitosamente.")
        return redirect("inventario_equipos")

    context = {"equipo": equipo}
    return render(request, "equipos/eliminar_equipo.html", context)


@login_required
def detalle_equipo(request, equipo_id):
    """Ver el detalle de un equipo, su historial de préstamos, fallas y mantenimiento"""
    equipo = get_object_or_404(Equipo, id=equipo_id)
    
    es_admin = request.user.role in ('ADMIN', 'SECRETARIA')
    tiene_prestamo = PrestamoEquipo.objects.filter(equipo=equipo, funcionario=request.user, activo=True).exists()
    
    if not (es_admin or tiene_prestamo):
        messages.error(request, "No tienes permisos para ver el historial de este equipo.")
        return redirect("dashboard")
    
    hitos = HitoMantenimiento.objects.filter(equipo=equipo).select_related("creado_por")
    fallas = equipo.fallas.all().select_related("funcionario", "resuelto_por")
    prestamos = equipo.prestamos.all().select_related("funcionario", "asignado_por").order_by("-fecha_asignacion")
    
    from itertools import chain
    from operator import attrgetter
    
    for h in hitos:
        h.tipo_evento = "HITO"
        h.fecha_orden = h.fecha
    
    for f in fallas:
        f.tipo_evento = "FALLA"
        f.fecha_orden = f.fecha_reporte.date()
        
    for p in prestamos:
        p.tipo_evento = "PRESTAMO"
        p.fecha_orden = p.fecha_asignacion.date()

    linea_tiempo = sorted(
        chain(hitos, fallas, prestamos),
        key=attrgetter("fecha_orden"),
        reverse=True
    )
    
    context = {
        "equipo": equipo,
        "linea_tiempo": linea_tiempo,
        "prestamo_actual": prestamos.filter(activo=True).first(),
        "es_admin": es_admin,
    }
    return render(request, "equipos/detalle_equipo.html", context)


@login_required
def agregar_hito(request, equipo_id):
    """Agregar un hito o mantenimiento a un equipo (solo administradores)"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")
        
    equipo = get_object_or_404(Equipo, id=equipo_id)
    
    if request.method == "POST":
        tipo = request.POST.get("tipo")
        fecha = request.POST.get("fecha")
        descripcion = request.POST.get("descripcion")
        costo_str = request.POST.get("costo")
        
        costo = None
        if costo_str and costo_str.strip():
            try:
                costo = float(costo_str.replace(",", "."))
            except ValueError:
                messages.error(request, "El costo ingresado no es válido.")
                return redirect("agregar_hito", equipo_id=equipo.id)
                
        HitoMantenimiento.objects.create(
            equipo=equipo,
            tipo=tipo,
            fecha=fecha,
            descripcion=descripcion,
            costo=costo,
            creado_por=request.user
        )
        
        messages.success(request, "Hito de mantenimiento registrado exitosamente.")
        return redirect("detalle_equipo", equipo_id=equipo.id)
        
    context = {
        "equipo": equipo,
        "tipos_hito": HitoMantenimiento.TIPO_HITO_CHOICES,
        "fecha_hoy": timezone.now().date().strftime("%Y-%m-%d")
    }
    return render(request, "equipos/agregar_hito.html", context)


@login_required
def asignar_equipo(request, equipo_id):
    """Asignar equipo permanente a un funcionario (solo administradores)"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    equipo = get_object_or_404(Equipo, id=equipo_id)

    if equipo.estado == "ASIGNADO":
        messages.warning(request, "Este equipo ya está asignado.")
        return redirect("lista_equipos")

    if request.method == "POST":
        funcionario_id = request.POST.get("funcionario")
        try:
            funcionario = CustomUser.objects.get(id=funcionario_id)

            PrestamoEquipo.objects.create(
                equipo=equipo,
                funcionario=funcionario,
                observaciones=request.POST.get("observaciones", ""),
                asignado_por=request.user,
                es_prestamo_diario=False,
                activo=True
            )

            registrar_log(
                usuario=request.user,
                tipo="UPDATE",
                accion="Asignación de Equipo",
                descripcion=f"Se asignó {equipo} a {funcionario.get_full_name()}",
                ip_address=get_client_ip(request)
            )
            messages.success(request, f"Equipo asignado a {funcionario.get_full_name()}")
            return redirect("lista_equipos")
        except CustomUser.DoesNotExist:
            messages.error(request, "Funcionario no encontrado.")

    funcionarios = CustomUser.objects.filter(is_active=True).order_by("first_name", "last_name")

    context = {
        "equipo": equipo,
        "funcionarios": funcionarios
    }
    return render(request, "equipos/asignar_equipo.html", context)


@login_required
def devolver_equipo(request, prestamo_id):
    """Devolver equipo asignado (solo administradores)"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")

    prestamo = get_object_or_404(PrestamoEquipo, id=prestamo_id)

    if request.method == "POST":
        prestamo.activo = False
        prestamo.fecha_devolucion = timezone.now().date()
        prestamo.fecha_devolucion_real = timezone.now()
        prestamo.save()
        registrar_log(
            usuario=request.user,
            tipo="UPDATE",
            accion="Devolución de Equipo",
            descripcion=f"Se devolvió equipo {prestamo.equipo} por {prestamo.funcionario.get_full_name()}",
            ip_address=get_client_ip(request)
        )
        messages.success(request, f"Equipo {prestamo.equipo} devuelto exitosamente.")
        return redirect("lista_equipos")

    context = {"prestamo": prestamo}
    return render(request, "equipos/devolver_equipo.html", context)


@login_required
def mis_equipos(request):
    """Ver equipos asignados al usuario actual (funcionario)"""
    prestamos = PrestamoEquipo.objects.filter(
        funcionario=request.user,
        activo=True
    ).select_related("equipo", "equipo__lugar")

    laptops_count = prestamos.filter(equipo__tipo="LAPTOP").count()

    context = {
        "prestamos": prestamos,
        "laptops_count": laptops_count
    }
    return render(request, "equipos/mis_equipos.html", context)


@login_required
def reportar_falla(request, equipo_id):
    """Permite a un funcionario reportar una falla en un equipo que tiene asignado"""
    equipo = get_object_or_404(Equipo, id=equipo_id)
    
    if not PrestamoEquipo.objects.filter(equipo=equipo, funcionario=request.user, activo=True).exists():
        messages.error(request, "No puedes reportar fallas de un equipo que no tienes asignado.")
        return redirect("mis_equipos")
    
    if request.method == "POST":
        descripcion = request.POST.get("descripcion")
        if not descripcion:
            messages.error(request, "Debes proporcionar una descripción de la falla.")
        else:
            FallaEquipo.objects.create(
                equipo=equipo,
                funcionario=request.user,
                descripcion=descripcion
            )
            messages.success(request, "Falla reportada correctamente. El administrador será notificado.")
        return redirect("mis_equipos")
    
    return redirect("mis_equipos")


@login_required
def actualizar_estado_falla(request, falla_id):
    """Actualiza el estado de un reporte de falla"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para esta acción.")
        return redirect("dashboard")
    
    falla = get_object_or_404(FallaEquipo, id=falla_id)
    
    if request.method == "POST":
        nuevo_estado = request.POST.get("estado")
        comentarios = request.POST.get("comentarios_tecnicos")
        
        if nuevo_estado in dict(FallaEquipo.ESTADO_FALLA_CHOICES):
            falla.estado = nuevo_estado
            falla.comentarios_tecnicos = comentarios
            falla.resuelto_por = request.user
            falla.save()
            messages.success(request, f"Estado de la falla actualizado a {falla.get_estado_display()}.")
        else:
            messages.error(request, "Estado no válido.")
            
    registrar_log(
        usuario=request.user,
        tipo="UPDATE",
        accion="Actualización Estado Falla",
        descripcion=f"Se actualizó falla de equipo {falla.equipo} a estado {falla.get_estado_display()}",
        ip_address=get_client_ip(request)
    )
    return redirect("gestion_fallas")


# ==========================================
# 6. EXPORTACIONES DE INVENTARIO Y REPORTES
# ==========================================

@login_required
def reporte_prestamos_pdf(request, usuario_id=None):
    """Generar reporte PDF de préstamos individuales o generales"""
    uid = usuario_id or request.GET.get("usuario_id")

    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        if not uid or str(request.user.id) != str(uid):
            messages.error(request, "No tienes permisos para acceder a esta sección.")
            return redirect("dashboard")

    if uid:
        funcionario = get_object_or_404(CustomUser, id=uid)
        prestamos = PrestamoEquipo.objects.filter(
            funcionario=funcionario,
            activo=True
        ).select_related("equipo", "funcionario", "equipo__lugar").order_by("-fecha_asignacion")
        
        for prestamo in prestamos:
            equipo = prestamo.equipo
            equipo.hitos_recientes = HitoMantenimiento.objects.filter(equipo=equipo).order_by("-fecha")[:5]
            equipo.fallas_recientes = FallaEquipo.objects.filter(equipo=equipo).order_by("-fecha_reporte")[:5]

        titulo = f"Certificado de Préstamo - {funcionario.get_full_name()}"
        filename = f"prestamo_equipos_{funcionario.last_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        html_string = render(request, "equipos/comprobante_prestamo_pdf.html", {
            "funcionario": funcionario,
            "prestamos": prestamos,
            "titulo": titulo,
            "fecha": timezone.now()
        }).content.decode("utf-8")
    else:
        prestamos = PrestamoEquipo.objects.select_related("equipo", "funcionario", "equipo__lugar").all()
        titulo = "Reporte General de Préstamos de Equipos"
        filename = f"reporte_prestamos_{datetime.now().strftime('%Y%m%d')}.pdf"

        prestamos_por_usuario = {}
        for prestamo in prestamos:
            key = prestamo.funcionario.id
            if key not in prestamos_por_usuario:
                prestamos_por_usuario[key] = {
                    "funcionario": prestamo.funcionario,
                    "prestamos": []
                }
            prestamos_por_usuario[key]["prestamos"].append(prestamo)

        html_string = render(request, "equipos/reporte_prestamos.html", {
            "prestamos_por_usuario": prestamos_por_usuario,
            "titulo": titulo,
            "fecha": timezone.now()
        }).content.decode("utf-8")

    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_inventario_excel(request):
    """Exportar inventario de equipos a Excel con asignaciones"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")
    
    equipos = Equipo.objects.all().prefetch_related(
        Prefetch(
            "prestamos",
            queryset=PrestamoEquipo.objects.filter(activo=True).select_related("funcionario"),
            to_attr="prestamo_activo_list"
        )
    ).select_related("lugar").order_by("tipo", "marca", "modelo")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Equipos"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    headers = [
        "Tipo", "Marca", "Modelo", "N° Serie", "N° Inventario",
        "Estado", "Funcionario Asignado", "RUT Funcionario",
        "Fecha Asignación", "Fecha Adquisición", "Fecha Origen", "Lugar", "Observaciones"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, equipo in enumerate(equipos, 2):
        prestamo_activo = equipo.prestamo_activo_list[0] if equipo.prestamo_activo_list else None
        
        ws.cell(row=row, column=1, value=equipo.get_tipo_display()).border = thin_border
        ws.cell(row=row, column=2, value=equipo.marca).border = thin_border
        ws.cell(row=row, column=3, value=equipo.modelo).border = thin_border
        ws.cell(row=row, column=4, value=equipo.numero_serie).border = thin_border
        ws.cell(row=row, column=5, value=equipo.numero_inventario).border = thin_border
        ws.cell(row=row, column=6, value=equipo.get_estado_display()).border = thin_border
        
        if prestamo_activo:
            ws.cell(row=row, column=7, value=prestamo_activo.funcionario.get_full_name()).border = thin_border
            ws.cell(row=row, column=8, value=prestamo_activo.funcionario.run).border = thin_border
            ws.cell(row=row, column=9, value=prestamo_activo.fecha_asignacion.strftime("%d/%m/%Y") if prestamo_activo.fecha_asignacion else "").border = thin_border
        else:
            ws.cell(row=row, column=7, value="-").border = thin_border
            ws.cell(row=row, column=8, value="-").border = thin_border
            ws.cell(row=row, column=9, value="-").border = thin_border
        
        ws.cell(row=row, column=10, value=equipo.fecha_adquisicion.strftime("%d/%m/%Y") if equipo.fecha_adquisicion else "").border = thin_border
        ws.cell(row=row, column=11, value=equipo.fecha_origen.strftime("%d/%m/%Y") if equipo.fecha_origen else "").border = thin_border
        ws.cell(row=row, column=12, value=equipo.lugar.nombre if equipo.lugar else "").border = thin_border
        ws.cell(row=row, column=13, value=equipo.observaciones or "").border = thin_border
    
    column_widths = [15, 15, 20, 20, 18, 15, 25, 15, 15, 15, 15, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=inventario_equipos_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(response)
    
    return response


@login_required
def export_inventario_pdf(request):
    """Exportar inventario de equipos a PDF con asignaciones"""
    if request.user.role not in ('ADMIN', 'SECRETARIA'):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect("dashboard")
    
    equipos = Equipo.objects.all().prefetch_related(
        Prefetch(
            "prestamos",
            queryset=PrestamoEquipo.objects.filter(activo=True).select_related("funcionario"),
            to_attr="prestamo_activo_list"
        )
    ).select_related("lugar").order_by("tipo", "marca", "modelo")
    
    equipos_data = []
    for equipo in equipos:
        prestamo_activo = equipo.prestamo_activo_list[0] if equipo.prestamo_activo_list else None
        equipos_data.append({
            "equipo": equipo,
            "prestamo": prestamo_activo
        })
    
    html_string = render(request, "equipos/inventario_pdf.html", {
        "equipos_data": equipos_data,
        "fecha": timezone.now(),
        "total": equipos.count()
    }).content.decode("utf-8")
    
    pdf = HTML(string=html_string).write_pdf()
    
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f"attachment; filename=inventario_equipos_{datetime.now().strftime('%Y%m%d')}.pdf"
    return response
