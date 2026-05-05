from django.shortcuts import render, redirect
from django.urls import reverse, reverse_lazy
from django.views.generic import TemplateView, FormView, ListView, View, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Avg, Sum
from django.http import HttpResponse
import openpyxl
from datetime import datetime, time, timedelta
from django.utils import timezone
from licencias.models import LicenciaMedica
from permisos.models import SolicitudPermiso
from .forms import EditarRegistroAsistenciaForm
import zipfile
import io
import re
import logging

# PDF generation
from weasyprint import HTML, CSS
from django.template.loader import render_to_string

# Import xlrd conditionally
try:
    import xlrd
    from xlrd import XLRDError
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    xlrd = None
    XLRDError = Exception

# Import pypdf for PDF processing (like payroll system)
try:
    import pypdf
    PYPDF_AVAILABLE = True
except ImportError:
    PYPDF_AVAILABLE = False
    pypdf = None
from .models import HorarioFuncionario, RegistroAsistencia, DiaFestivo, AlegacionAsistencia, AnoEscolar, DiaHorario, HorarioExcepcional
from .forms import CargaHorariosForm, HorarioFuncionarioForm, CargaRegistrosAsistenciaForm, DiaFestivoForm, HorarioExcepcionalForm
from django.shortcuts import get_object_or_404, redirect
from users.models import CustomUser
from core.utils import normalize_rut
from admin_dashboard.utils import registrar_log, get_client_ip

logger = logging.getLogger(__name__)


def find_user_by_rut(rut_encontrado: str):
    """
    Busca un usuario por RUT con matching inteligente que maneja múltiples formatos.
    """
    try:
        # Limpiar el RUT encontrado (remover espacios, mayúsculas)
        rut_limpio = rut_encontrado.upper().replace(' ', '').strip()

        # Crear diferentes variaciones para buscar
        variaciones_rut = set()

        # 1. RUT original limpio
        variaciones_rut.add(rut_limpio)

        # 2. RUT normalizado (con puntos)
        rut_normalizado = normalize_rut(rut_encontrado)
        variaciones_rut.add(rut_normalizado)

        # 3. RUT sin puntos
        rut_sin_puntos = rut_normalizado.replace('.', '')
        variaciones_rut.add(rut_sin_puntos)

        # 4. Si tiene puntos, intentar sin ellos
        if '.' in rut_limpio:
            variaciones_rut.add(rut_limpio.replace('.', ''))

        # 5. Si no tiene puntos pero tiene guión, intentar con puntos
        if '.' not in rut_limpio and '-' in rut_limpio:
            # Intentar agregar puntos automáticamente
            parts = rut_limpio.split('-')
            if len(parts) == 2:
                cuerpo, dv = parts
                cuerpo = cuerpo.replace('.', '')  # Remover puntos existentes si los hay

                if len(cuerpo) == 8:  # RUT de 8 dígitos: 12345678 -> 12.345.678
                    cuerpo_con_puntos = f"{cuerpo[:2]}.{cuerpo[2:5]}.{cuerpo[5:]}"
                    variaciones_rut.add(f"{cuerpo_con_puntos}-{dv}")
                elif len(cuerpo) == 7:  # RUT de 7 dígitos: 1234567 -> 1.234.567
                    cuerpo_con_puntos = f"{cuerpo[:1]}.{cuerpo[1:4]}.{cuerpo[4:]}"
                    variaciones_rut.add(f"{cuerpo_con_puntos}-{dv}")

        # Intentar cada variación
        for rut_variacion in variaciones_rut:
            try:
                user = CustomUser.objects.get(run=rut_variacion)
                logger.info(f"✅ RUT encontrado: '{rut_encontrado}' → '{rut_variacion}' ({user.get_full_name()})")
                return user
            except CustomUser.DoesNotExist:
                continue

        # Si ninguna variación funcionó, mostrar debug info
        logger.warning(f"❌ RUT '{rut_encontrado}' no encontrado. Variaciones probadas: {sorted(variaciones_rut)}")

        # Mostrar algunos RUTs de la base de datos para comparación
        sample_users = CustomUser.objects.all()[:10]
        logger.info(f"Muestra de RUTs en BD ({len(sample_users)} usuarios):")
        for i, user in enumerate(sample_users):
            logger.info(f"  {i+1}. RUT: '{user.run}' - Nombre: {user.get_full_name()}")

        # Mostrar todos los RUTs únicos en la BD para debugging
        all_runs = list(CustomUser.objects.values_list('run', flat=True).distinct())
        logger.info(f"Todos los RUTs en BD ({len(all_runs)}): {sorted(all_runs)}")

        # Buscar RUTs similares (primeros 8 dígitos)
        base_rut = rut_sin_puntos.replace('-', '')[:8]
        similar_users = CustomUser.objects.filter(run__icontains=base_rut)[:5]
        if similar_users:
            logger.info("RUTs similares encontrados:")
            for user in similar_users:
                logger.info(f"  '{user.run}' - {user.get_full_name()}")

        return None

    except Exception as e:
        logger.error(f"Error finding user by RUT {rut_encontrado}: {e}")
        return None


def load_data_file(archivo, mes=None, anio=None):
    """Carga datos de archivos Excel (.xlsx/.xls) o PDF y retorna filas de datos"""
    # Asegurar que el puntero del archivo esté al inicio
    if hasattr(archivo, 'seek'):
        archivo.seek(0)

    filename = archivo.name.lower()

    try:
        if filename.endswith(('.xlsx', '.xls')):
            # Procesar archivos Excel
            if filename.endswith('.xlsx'):
                # Usar openpyxl para .xlsx
                wb = openpyxl.load_workbook(archivo, data_only=True)
                ws = wb.active
                # Convertir a lista de filas
                rows = list(ws.iter_rows(min_row=2, values_only=True))
            elif filename.endswith('.xls'):
                # Intentar usar xlrd para .xls
                if not XLRD_AVAILABLE:
                    raise Exception("Los archivos .xls no son soportados actualmente. Por favor, convierta su archivo .xls a .xlsx usando Excel o Google Sheets y vuelva a intentarlo.")
                # Usar xlrd para .xls
                wb = xlrd.open_workbook(file_contents=archivo.read())
                ws = wb.sheet_by_index(0)  # Primera hoja
                # Convertir a lista de filas
                rows = [tuple(ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols))
                       for row_idx in range(1, ws.nrows)]  # Skip header row

            return rows

        elif filename.endswith('.pdf'):
            # Procesar archivos PDF (similar al sistema de liquidaciones)
            if not PYPDF_AVAILABLE:
                raise Exception("Los archivos PDF no son soportados actualmente. Instale pypdf para habilitar esta funcionalidad.")

            # Extraer datos del PDF
            rows = []
            pdf_reader = pypdf.PdfReader(archivo)

            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    text = page.extract_text()
                    if not text.strip():
                        continue

                    # Intentar extraer datos tabulares del texto
                    # Buscar patrones de asistencia: RUT, Nombre, Horario
                    lines = text.split('\n')

                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue

                        # Intentar parsear línea como datos de asistencia
                        # Formato esperado: "RUT, Nombre Horario"
                        # Ejemplo: "12345678-9, Juan Pérez 08:30-17:30"

                        # Buscar patrón: RUT seguido de coma, luego nombre, luego horario
                        match = re.match(r'^(\d{7,8}-[\dKk]|\d{1,2}\.\d{3}\.\d{3}-[\dKk]|\d{8,9})\s*,\s*(.+?)\s+(\d{1,2}:\d{2}(?::\d{2})?(?:\s*-\s*\d{1,2}:\d{2}(?::\d{2})?)?)$', line)
                        if match:
                            rut = match.group(1).strip()
                            nombre = match.group(2).strip()
                            horario = match.group(3).strip()

                            # Parsear el horario (ej: "08:30-17:30" o solo "08:30")
                            horario_parts = horario.split('-')
                            hora_entrada_str = horario_parts[0].strip()
                            hora_salida_str = horario_parts[1].strip() if len(horario_parts) > 1 else None

                            # Para PDFs con formato "RUT, Nombre Horario", usamos la fecha del formulario
                            if mes and anio:
                                # Crear fecha del primer día del mes especificado
                                fecha_str = f"01/{mes:02d}/{anio}"
                            else:
                                fecha_str = datetime.now().strftime("%d/%m/%Y")

                            rows.append((rut, fecha_str, hora_entrada_str, hora_salida_str or ''))
                        else:
                            # Intentar otros formatos posibles
                            # Formato alternativo: RUT Nombre Horario (sin coma)
                            alt_match = re.match(r'^(\d{7,8}-[\dKk]|\d{1,2}\.\d{3}\.\d{3}-[\dKk]|\d{8,9})\s+(.+?)\s+(\d{1,2}:\d{2}(?::\d{2})?(?:\s*-\s*\d{1,2}:\d{2}(?::\d{2})?)?)$', line)
                            if alt_match:
                                rut = alt_match.group(1).strip()
                                nombre = alt_match.group(2).strip()
                                horario = alt_match.group(3).strip()

                                horario_parts = horario.split('-')
                                hora_entrada_str = horario_parts[0].strip()
                                hora_salida_str = horario_parts[1].strip() if len(horario_parts) > 1 else None

                                # Usar fecha del formulario para este formato también
                                if mes and anio:
                                    fecha_str = f"01/{mes:02d}/{anio}"
                                else:
                                    fecha_str = datetime.now().strftime("%d/%m/%Y")

                                rows.append((rut, fecha_str, hora_entrada_str, hora_salida_str or ''))

                except Exception as e:
                    # Continuar con la siguiente página si hay error
                    continue

            if not rows:
                raise Exception("No se encontraron datos de asistencia en el archivo PDF. Asegúrese de que el PDF contenga información de asistencia en formato tabular o de texto estructurado.")

            return rows

        else:
            raise Exception("Formato de archivo no soportado. Use .xlsx, .xls o .pdf")

    except Exception as e:
        # Mejorar el manejo de errores para identificar el tipo de archivo
        error_msg = str(e)

        # Si es un error de Excel pero el archivo podría ser PDF
        if "File is not a zip file" in error_msg or "BadZipFile" in error_msg:
            # Intentar detectar si es un PDF mal etiquetado
            archivo.seek(0)
            first_bytes = archivo.read(8)
            if first_bytes.startswith(b'%PDF-'):
                raise Exception("El archivo parece ser un PDF pero tiene extensión .xls. Cambie la extensión a .pdf o use un archivo Excel válido.")
            else:
                raise Exception("El archivo no es un archivo Excel válido. Verifique que no esté corrupto.")

        # Si es un error de xlrd
        if "XLRDError" in str(type(e)) or "xlrd" in error_msg.lower():
            raise Exception(f"Error al leer el archivo Excel: {error_msg}")

        # Si es un error de PDF
        if "pdf" in error_msg.lower() or "PDF" in error_msg:
            raise Exception(f"Error al procesar el archivo PDF: {error_msg}")

        # Error genérico
        raise Exception(f"Error al procesar el archivo: {error_msg}")

class AsistenciaAdminView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista principal de administración de asistencia"""
    template_name = 'asistencia/admin_dashboard.html'

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas generales
        total_funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']).count()
        funcionarios_con_horario = HorarioFuncionario.objects.filter(activo=True).count()
        registros_hoy = RegistroAsistencia.objects.filter(fecha=datetime.now().date()).count()

        context.update({
            'total_funcionarios': total_funcionarios,
            'funcionarios_con_horario': funcionarios_con_horario,
            'registros_hoy': registros_hoy,
            'horarios_pendientes': total_funcionarios - funcionarios_con_horario,
        })

        return context


class GestionHorariosView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para gestionar horarios de funcionarios"""
    template_name = 'asistencia/gestion_horarios.html'

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Búsqueda
        search = self.request.GET.get('search', '').strip()

        # Obtener todos los usuarios del sistema
        funcionarios = CustomUser.objects.filter(
            role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']
        ).order_by('first_name')

        # Obtener datos del mes
        registros_mes = RegistroAsistencia.objects.filter(
            fecha__year=anio,
            fecha__month=mes
        ).select_related('funcionario', 'horario_asignado')

        # Crear mapa de registros por funcionario
        registros_por_funcionario = {}
        for registro in registros_mes:
            func_id = registro.funcionario.id
            if func_id not in registros_por_funcionario:
                registros_por_funcionario[func_id] = []
            registros_por_funcionario[func_id].append(registro)

        # Procesar cada funcionario y solo incluir aquellos con atrasos o inasistencias sin justificar
        funcionarios_lista = []
        for funcionario in todos_funcionarios:
            func_id = funcionario.id
            registros_funcionario = registros_por_funcionario.get(func_id, [])

            # Inicializar datos del funcionario
            func_data = {
                'funcionario': funcionario,
                'atrasos': [],
                'inasistencias': [],
                'tiene_registros': len(registros_funcionario) > 0,
                'total_atrasos': 0,
                'total_minutos_retraso': 0,
                'total_inasistencias_sin_justificar': 0,
            }

            # Procesar registros del funcionario
            for registro in registros_funcionario:
                if registro.estado == 'RETRASO':
                    atraso_info = {
                        'fecha': registro.fecha,
                        'hora_entrada': registro.hora_entrada_real,
                        'minutos_retraso': registro.minutos_retraso,
                    }
                    func_data['atrasos'].append(atraso_info)
                    func_data['total_atrasos'] += 1
                    func_data['total_minutos_retraso'] += registro.minutos_retraso or 0
                elif registro.estado == 'AUSENTE':
                    # Ignorar si es antes de su ingreso
                    if registro.fecha < funcionario.date_joined.date():
                        continue
                        
                    inasistencia_info = {
                        'fecha': registro.fecha,
                        'hora_esperada': registro.horario_asignado.hora_entrada if registro.horario_asignado else None,
                        'justificada': False,
                    }
                    func_data['inasistencias'].append(inasistencia_info)
                    func_data['total_inasistencias_sin_justificar'] += 1
                # Ignorar registros justificados - solo mostrar injustificados

            # Detectar días sin registro que son inasistencias
            fechas_con_registro = {r.fecha for r in registros_funcionario}
            today = datetime.now().date()
            num_dias = cal.monthrange(anio, mes)[0]
            for dia in range(1, num_dias + 1):
                fecha = datetime(anio, mes, dia).date()
                if fecha >= today:
                    continue
                if fecha in fechas_con_registro:
                    continue
                if DiaFestivo.objects.filter(fecha=fecha).exists():
                    continue
                if fecha.weekday() >= 5 and not (funcionario.funcion == 'SERENO' or funcionario.tipo_funcionario == 'SERENO'):
                    continue
                if fecha < funcionario.date_joined.date():
                    continue
                inasistencia_info = {
                    'fecha': fecha,
                    'hora_esperada': None,
                    'justificada': False,
                }
                func_data['inasistencias'].append(inasistencia_info)
                func_data['total_inasistencias_sin_justificar'] += 1

            # Solo incluir funcionarios que tienen atrasos o inasistencias sin justificar
            if func_data['atrasos'] or func_data['inasistencias']:
                # Ordenar listas por fecha
                func_data['atrasos'].sort(key=lambda x: x['fecha'])
                func_data['inasistencias'].sort(key=lambda x: x['fecha'])
                funcionarios_lista.append(func_data)

        # Nombre del mes
        meses = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        nombre_mes = meses[mes - 1]

        # Renderizar template HTML para PDF
        html_content = render_to_string('asistencia/reporte_mensual_pdf.html', {
            'funcionarios': funcionarios_lista,
            'anio': anio,
            'mes': mes,
            'nombre_mes': nombre_mes,
            'fecha_actual': datetime.now(),
        })

        # Generar PDF
        pdf_file = HTML(string=html_content).write_pdf()

        # Crear respuesta HTTP
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'reporte_asistencia_{anio}_{mes:02d}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class RecalcularEstadoAsistenciaView(LoginRequiredMixin, View):
    """Vista para recalcular el estado de todos los registros de asistencia del usuario actual"""

    def post(self, request):
        # Obtener todos los registros del usuario
        registros = RegistroAsistencia.objects.filter(funcionario=request.user)

        if not registros.exists():
            messages.warning(request, 'No tiene registros de asistencia para recalcular.')
            return redirect('asistencia:mi_asistencia')

        registros_actualizados = 0

        # Recalcular estado para cada registro
        for registro in registros:
            # Actualizar horario_asignado al horario actual del usuario (si existe)
            try:
                horario_actual = HorarioFuncionario.objects.filter(
                    funcionario=registro.funcionario, activo=True
                ).first()
                if horario_actual:
                    registro.horario_asignado = horario_actual
            except:
                pass

            # Forzar recálculo del estado llamando al método save
            registro.save()

            registros_actualizados += 1

        messages.success(
            request,
            f'Se recalcularon {registros_actualizados} registros de asistencia. Los estados ahora reflejan su horario actual y permisos/licencias vigentes.'
        )

        # Redirigir de vuelta a la vista de asistencia con los filtros actuales
        mes = request.GET.get('mes')
        anio = request.GET.get('anio')

        if mes and anio:
            return redirect(f'/asistencia/mi-asistencia/?mes={mes}&anio={anio}')
        else:
            return redirect('asistencia:mi_asistencia')


class RecalcularTodaAsistenciaView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para recalcular TODOS los registros de asistencia de TODOS los funcionarios"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def post(self, request):
        registros = RegistroAsistencia.objects.select_related('funcionario').all()

        if not registros.exists():
            messages.warning(request, 'No hay registros de asistencia para recalcular.')
            return redirect('asistencia:gestion_asistencia')

        registros_actualizados = 0

        for registro in registros:
            try:
                horario_actual = HorarioFuncionario.objects.filter(
                    funcionario=registro.funcionario, activo=True
                ).first()
                if horario_actual:
                    registro.horario_asignado = horario_actual
            except Exception:
                pass

            registro.save()
            registros_actualizados += 1

        messages.success(
            request,
            f'Se recalcularon {registros_actualizados} registros de asistencia de todos los funcionarios.'
        )

        registrar_log(
            usuario=request.user,
            tipo='UPDATE',
           accion='Recálculo Masivo de Asistencia',
            descripcion=f'Se recalcularon {registros_actualizados} registros de asistencia',
            ip_address=get_client_ip(request)
        )

        return redirect('asistencia:gestion_asistencia')


class RecalcularAsistenciaUsuarioView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para recalcular la asistencia de un usuario en particular"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def post(self, request, user_id):
        usuario = get_object_or_404(CustomUser, id=user_id)
        mes = request.POST.get('mes')
        anio = request.POST.get('anio')

        registros = RegistroAsistencia.objects.filter(funcionario=usuario)

        if mes and anio:
            try:
                mes = int(mes)
                anio = int(anio)
                registros = registros.filter(fecha__month=mes, fecha__year=anio)
            except ValueError:
                pass

        if not registros.exists():
            messages.warning(request, f'No hay registros de asistencia para {usuario.get_full_name()} en ese periodo.')
            return redirect(f'/asistencia/usuario/{user_id}/')

        registros_actualizados = 0

        for registro in registros:
            try:
                horario_actual = HorarioFuncionario.objects.filter(
                    funcionario=registro.funcionario, activo=True
                ).first()
                if horario_actual:
                    registro.horario_asignado = horario_actual
            except Exception:
                pass

            registro.save()
            registros_actualizados += 1

        # Formatear mes y anio para el redirect y el log si existen
        texto_periodo = ""
        url_redirect = f'/asistencia/usuario/{user_id}/'
        
        if mes and anio:
            texto_periodo = f" (Mes {mes}/{anio})"
            url_redirect += f"?anio={anio}"

        messages.success(
            request,
            f'Se recalcularon {registros_actualizados} registros de asistencia para {usuario.get_full_name()}{texto_periodo}.'
        )

        registrar_log(
            usuario=request.user,
            tipo='UPDATE',
            accion='Recálculo de Asistencia Individual',
            descripcion=f'Se recalcularon {registros_actualizados} registros para el usuario {usuario.run}{texto_periodo}',
            ip_address=get_client_ip(request)
        )

        return redirect(url_redirect)


class ReporteAsistenciaIndividualView(LoginRequiredMixin, View):
    """Vista para generar reporte individual de asistencia en PDF"""

    def get(self, request, anio, mes):
        import calendar as cal

        # Obtener datos del usuario actual para el mes
        registros_mes = list(RegistroAsistencia.objects.filter(
            funcionario=request.user,
            fecha__year=anio,
            fecha__month=mes
        ).select_related('horario_asignado').order_by('fecha'))

        # Obtener horario del funcionario
        try:
            horario = HorarioFuncionario.objects.get(funcionario=request.user, activo=True)
        except HorarioFuncionario.DoesNotExist:
            horario = None

        # Obtener festivos del mes
        festivos = set(
            DiaFestivo.objects.filter(
                fecha__year=anio, fecha__month=mes
            ).values_list('fecha', flat=True)
        )

        # Determinar si es sereno
        es_sereno = request.user.role == 'FUNCIONARIO' and request.user.funcion == 'SERENO'
        if request.user.tipo_funcionario == 'SERENO':
            es_sereno = True

        # Recopilar detalles de atrasos, inasistencias y justificaciones
        atrasos_detalle = []
        inasistencias_detalle = []
        justificaciones_detalle = []

        # Obtener fechas con registro de forma confiable
        fechas_con_registro = set()
        for r in registros_mes:
            fechas_con_registro.add(r.fecha)

        # Primero: procesar registros existentes por estado
        for registro in registros_mes:
            if registro.estado == 'RETRASO':
                atrasos_detalle.append({
                    'fecha': registro.fecha,
                    'hora_entrada': registro.hora_entrada_real,
                    'minutos_retraso': registro.minutos_retraso,
                })
            elif registro.estado == 'AUSENTE':
                # Ignorar si es antes de su ingreso
                if registro.fecha < request.user.date_joined.date():
                    continue
                    
                inasistencias_detalle.append({
                    'fecha': registro.fecha,
                    'hora_esperada': registro.horario_asignado.hora_entrada if registro.horario_asignado else None,
                })
            elif registro.estado in ['JUSTIFICADO', 'DIA_ADMINISTRATIVO', 'LICENCIA_MEDICA']:
                if registro.estado == 'DIA_ADMINISTRATIVO':
                    tipo = 'permiso'
                elif registro.estado == 'LICENCIA_MEDICA':
                    tipo = 'licencia'
                else:
                    tipo = 'permiso' if registro.tiene_permiso_aprobado() else 'licencia' if registro.tiene_licencia_medica() else 'otro'
                justificaciones_detalle.append({
                    'fecha': registro.fecha,
                    'tipo': tipo,
                })

        # Segundo: detectar días sin registro que son inasistencias
        # Días pasados sin registro, que no sean festivos ni fines de semana (para no serenos)
        today = datetime.now().date()
        num_dias = cal.monthrange(anio, mes)[0]
        ano_escolar_activo = AnoEscolar.get_activo()
        for dia in range(1, num_dias + 1):
            fecha = datetime(anio, mes, dia).date()
            es_pasado = fecha < today
            if not es_pasado:
                continue
            tiene_registro = fecha in fechas_con_registro
            if tiene_registro:
                continue
            es_festivo = fecha in festivos
            if es_festivo:
                continue
            # Verificar año escolar SOLO si hay uno activo configurado
            if ano_escolar_activo and not AnoEscolar.es_dia_escolar(fecha):
                continue
            # Mismo filtro de fin de semana que la página
            dia_semana = fecha.weekday()
            es_fin_de_semana = dia_semana >= 5
            if es_fin_de_semana and not es_sereno:
                continue
            
            # No contar inasistencia si es antes de su ingreso
            if fecha < request.user.date_joined.date():
                continue

            # Verificar si tiene permiso administrativo aprobado
            if SolicitudPermiso.objects.filter(
                usuario=request.user,
                estado='APROBADO',
                fecha_inicio__lte=fecha,
                fecha_termino__gte=fecha
            ).exists():
                justificaciones_detalle.append({'fecha': fecha, 'tipo': 'permiso'})
                continue

            # Verificar si tiene licencia médica
            licencia_cubre = False
            for lic in LicenciaMedica.objects.filter(usuario=request.user, fecha_inicio__lte=fecha):
                fecha_fin_lic = lic.fecha_inicio + timedelta(days=lic.dias - 1)
                if fecha <= fecha_fin_lic:
                    licencia_cubre = True
                    break
            if licencia_cubre:
                justificaciones_detalle.append({'fecha': fecha, 'tipo': 'licencia'})
                continue

            # Es una inasistencia sin registro
            inasistencias_detalle.append({
                'fecha': fecha,
                'hora_esperada': horario.hora_entrada if horario else None,
            })

        # Ordenar por fecha
        atrasos_detalle.sort(key=lambda x: x['fecha'])
        inasistencias_detalle.sort(key=lambda x: x['fecha'])
        justificaciones_detalle.sort(key=lambda x: x['fecha'])

        # Totales
        total_atrasos = len(atrasos_detalle)
        total_inasistencias = len(inasistencias_detalle)
        total_justificados = len(justificaciones_detalle)
        total_minutos_retraso = sum(a['minutos_retraso'] for a in atrasos_detalle)

        # Nombre del mes
        meses = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        nombre_mes = meses[mes - 1]

        # Renderizar template HTML para PDF
        html_content = render_to_string('asistencia/reporte_individual_pdf.html', {
            'funcionario': request.user,
            'anio': anio,
            'mes': mes,
            'nombre_mes': nombre_mes,
            'atrasos_detalle': atrasos_detalle,
            'inasistencias_detalle': inasistencias_detalle,
            'justificaciones_detalle': justificaciones_detalle,
            'total_atrasos': total_atrasos,
            'total_inasistencias': total_inasistencias,
            'total_justificados': total_justificados,
            'total_minutos_retraso': total_minutos_retraso,
            'fecha_actual': datetime.now(),
            'ano_escolar': ano_escolar_activo,
        })

        # Generar PDF
        pdf_file = HTML(string=html_content).write_pdf()

        # Crear respuesta HTTP
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'mi_asistencia_{anio}_{mes:02d}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class ExportarRetrasosExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exporta atrasos a Excel - individual (con user_id) o masivo (sin user_id)"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get(self, request, user_id=None):
        anio = int(request.GET.get('anio', datetime.now().year))
        mes = int(request.GET.get('mes', datetime.now().month))

        wb = openpyxl.Workbook()

        if user_id:
            usuario = get_object_or_404(CustomUser, id=user_id)
            filename = f'atrasos_{usuario.last_name}_{anio}_{mes:02d}.xlsx'

            ws = wb.active
            ws.title = 'Atrasos'
            headers = ['RUT', 'Nombre', 'Fecha', 'Horario Est.', 'Entrada Real', 'Min. Retraso', 'Observación']
            header_fill = openpyxl.styles.PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_font = openpyxl.styles.Font(color='FFFFFF', bold=True, size=10)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            registros = RegistroAsistencia.objects.filter(
                funcionario=usuario,
                fecha__year=anio,
                fecha__month=mes,
                estado='RETRASO'
            ).order_by('fecha')

            row = 2
            for reg in registros:
                ws.cell(row=row, column=1, value=usuario.run)
                ws.cell(row=row, column=2, value=usuario.get_full_name())
                ws.cell(row=row, column=3, value=reg.fecha.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=4, value=reg.horario_asignado.hora_entrada.strftime('%H:%M') if reg.horario_asignado else '-')
                ws.cell(row=row, column=5, value=reg.hora_entrada_real.strftime('%H:%M') if reg.hora_entrada_real else '-')
                ws.cell(row=row, column=6, value=reg.minutos_retraso)
                ws.cell(row=row, column=7, value=reg.justificacion_manual or '')
                row += 1

            # Resumen
            row += 1
            ws.cell(row=row, column=1, value='TOTAL ATRASOS:').font = openpyxl.styles.Font(bold=True)
            ws.cell(row=row, column=2, value=registros.count()).font = openpyxl.styles.Font(bold=True, color='FF0000')
            total_min = sum(r.minutos_retraso for r in registros)
            ws.cell(row=row, column=4, value='TOTAL MIN.').font = openpyxl.styles.Font(bold=True)
            ws.cell(row=row, column=5, value=total_min).font = openpyxl.styles.Font(bold=True, color='FF0000')

            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 30)

        else:
            # Masivo: resumen por usuario
            filename = f'atrasos_todos_{anio}_{mes:02d}.xlsx'
            ws = wb.active
            ws.title = 'Resumen Atrasos'

            usuarios = CustomUser.objects.filter(
                registros_asistencia__fecha__year=anio,
                registros_asistencia__fecha__month=mes,
                registros_asistencia__estado='RETRASO'
            ).distinct().order_by('first_name', 'last_name')

            headers = ['N°', 'RUT', 'Nombre', 'Cargo', 'Días con Atraso', 'Total Min. Retraso']
            header_fill = openpyxl.styles.PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_font = openpyxl.styles.Font(color='FFFFFF', bold=True, size=10)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            row = 2
            num = 0
            total_general_min = 0
            for usuario in usuarios:
                registros = RegistroAsistencia.objects.filter(
                    funcionario=usuario,
                    fecha__year=anio,
                    fecha__month=mes,
                    estado='RETRASO'
                )
                dias_atraso = registros.count()
                total_min = sum(r.minutos_retraso for r in registros)
                total_general_min += total_min
                num += 1

                ws.cell(row=row, column=1, value=num)
                ws.cell(row=row, column=2, value=usuario.run)
                ws.cell(row=row, column=3, value=usuario.get_full_name())
                ws.cell(row=row, column=4, value=usuario.get_funcion_display() or usuario.get_role_display())
                ws.cell(row=row, column=5, value=dias_atraso)
                ws.cell(row=row, column=6, value=total_min)

                if total_min >= 60:
                    ws.cell(row=row, column=6).font = openpyxl.styles.Font(color='FF0000', bold=True)

                row += 1

            # Resumen
            row += 1
            ws.cell(row=row, column=4, value='TOTAL GENERAL:').font = openpyxl.styles.Font(bold=True)
            ws.cell(row=row, column=6, value=total_general_min).font = openpyxl.styles.Font(bold=True, color='FF0000', size=12)

            # Ajustar anchos
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 18

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportarRetrasosPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exporta atrasos a PDF - individual (con user_id) o masivo (sin user_id)"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get(self, request, user_id=None):
        anio = int(request.GET.get('anio', datetime.now().year))
        mes = int(request.GET.get('mes', datetime.now().month))

        meses = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        nombre_mes = meses[mes - 1]

        if user_id:
            usuario = get_object_or_404(CustomUser, id=user_id)
            registros = RegistroAsistencia.objects.filter(
                funcionario=usuario,
                fecha__year=anio,
                fecha__month=mes,
                estado='RETRASO'
            ).order_by('fecha')
            usuarios_data = [{'usuario': usuario, 'registros': registros, 'total': registros.count()}]
            filename = f'atrasos_{usuario.last_name}_{anio}_{mes:02d}.pdf'
            titulo = f'Reporte de Atrasos - {usuario.get_full_name()}'
            template = 'asistencia/reporte_retrasos_pdf.html'
        else:
            usuarios = CustomUser.objects.filter(
                registros_asistencia__fecha__year=anio,
                registros_asistencia__fecha__month=mes,
                registros_asistencia__estado='RETRASO'
            ).distinct().order_by('first_name', 'last_name')

            usuarios_data = []
            total_general = 0
            for usuario in usuarios:
                regs = RegistroAsistencia.objects.filter(
                    funcionario=usuario,
                    fecha__year=anio,
                    fecha__month=mes,
                    estado='RETRASO'
                )
                total_min = sum(r.minutos_retraso for r in regs)
                total_general += total_min
                usuarios_data.append({
                    'usuario': usuario,
                    'dias_atraso': regs.count(),
                    'total_minutos': total_min,
                })

            filename = f'atrasos_todos_{anio}_{mes:02d}.pdf'
            titulo = 'Reporte Masivo de Atrasos'
            template = 'asistencia/reporte_retrasos_masivo_pdf.html'

        html_content = render_to_string(template, {
            'usuarios_data': usuarios_data,
            'total_general': total_general if not user_id else None,
            'anio': anio,
            'mes': mes,
            'nombre_mes': nombre_mes,
            'titulo': titulo,
            'fecha_actual': datetime.now(),
        })

        pdf_file = HTML(string=html_content).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class GestionAnoEscolarView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para gestionar la configuración del año escolar"""
    template_name = 'asistencia/gestion_ano_escolar.html'

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ano_activo'] = AnoEscolar.get_activo()
        context['todos_anos'] = AnoEscolar.objects.all()
        return context

    def post(self, request):
        from django.core.exceptions import ValidationError

        ano = request.POST.get('ano')
        sem1_inicio = request.POST.get('sem1_inicio')
        sem1_fin = request.POST.get('sem1_fin')
        sem2_inicio = request.POST.get('sem2_inicio')
        sem2_fin = request.POST.get('sem2_fin')
        accion = request.POST.get('accion')

        if accion == 'eliminar':
            pk = request.POST.get('pk')
            try:
                ano_obj = AnoEscolar.objects.get(pk=pk)
                ano_obj.delete()
                messages.success(request, f'Año escolar {pk} eliminado correctamente.')
            except AnoEscolar.DoesNotExist:
                messages.error(request, 'Año escolar no encontrado.')
            return redirect('asistencia:gestion_ano_escolar')

        if accion == 'activar':
            pk = request.POST.get('pk')
            try:
                AnoEscolar.objects.update(activo=False)
                ano_obj = AnoEscolar.objects.get(pk=pk)
                ano_obj.activo = True
                ano_obj.save()
                messages.success(request, f'Año escolar {ano_obj.ano} activado correctamente.')
            except AnoEscolar.DoesNotExist:
                messages.error(request, 'Año escolar no encontrado.')
            return redirect('asistencia:gestion_ano_escolar')

        if accion == 'desactivar':
            pk = request.POST.get('pk')
            try:
                ano_obj = AnoEscolar.objects.get(pk=pk)
                ano_obj.activo = False
                ano_obj.save()
                messages.success(request, f'Año escolar {ano_obj.ano} desactivado.')
            except AnoEscolar.DoesNotExist:
                messages.error(request, 'Año escolar no encontrado.')
            return redirect('asistencia:gestion_ano_escolar')

        # Crear o actualizar año escolar
        if not all([ano, sem1_inicio, sem1_fin, sem2_inicio, sem2_fin]):
            messages.error(request, 'Todos los campos son obligatorios.')
            return redirect('asistencia:gestion_ano_escolar')

        try:
            ano = int(ano)
            sem1_inicio = datetime.strptime(sem1_inicio, '%Y-%m-%d').date()
            sem1_fin = datetime.strptime(sem1_fin, '%Y-%m-%d').date()
            sem2_inicio = datetime.strptime(sem2_inicio, '%Y-%m-%d').date()
            sem2_fin = datetime.strptime(sem2_fin, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Formato de fecha inválido.')
            return redirect('asistencia:gestion_ano_escolar')

        # Verificar si ya existe
        existente = AnoEscolar.objects.filter(ano=ano).first()
        if existente:
            existente.sem1_inicio = sem1_inicio
            existente.sem1_fin = sem1_fin
            existente.sem2_inicio = sem2_inicio
            existente.sem2_fin = sem2_fin
            existente.save()
            messages.success(request, f'Año escolar {ano} actualizado correctamente.')
        else:
            AnoEscolar.objects.create(
                ano=ano,
                sem1_inicio=sem1_inicio,
                sem1_fin=sem1_fin,
                sem2_inicio=sem2_inicio,
                sem2_fin=sem2_fin,
                activo=False,
                creado_por=request.user,
            )
            messages.success(request, f'Año escolar {ano} creado correctamente.')

        return redirect('asistencia:gestion_ano_escolar')

class GuardarHorarioSemanalView(LoginRequiredMixin, UserPassesTestMixin, View):
    """API para guardar la configuración del horario semanal de un usuario"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def post(self, request, user_id):
        import json
        from django.http import JsonResponse
        from django.db import transaction

        try:
            usuario = get_object_or_404(CustomUser, pk=user_id)
            data = json.loads(request.body)
            
            with transaction.atomic():
                # Obtener o crear el HorarioFuncionario base
                horario_base, created = HorarioFuncionario.objects.get_or_create(
                    funcionario=usuario,
                    defaults={
                        'hora_entrada': time(7, 55),
                        'activo': True
                    }
                )

                dias_data = data.get('dias', [])
                for dia_data in dias_data:
                    dia_semana = int(dia_data.get('dia_semana'))
                    activo = bool(dia_data.get('activo', False))
                    hora_entrada_str = dia_data.get('hora_entrada')
                    hora_salida_str = dia_data.get('hora_salida')

                    hora_entrada = None
                    hora_salida = None

                    if activo:
                        if hora_entrada_str:
                            try:
                                h, m = map(int, hora_entrada_str.split(':'))
                                hora_entrada = time(h, m)
                            except ValueError:
                                pass
                        
                        if hora_salida_str:
                            try:
                                h, m = map(int, hora_salida_str.split(':'))
                                hora_salida = time(h, m)
                            except ValueError:
                                pass
                        
                    # Validar tope de 44 horas semanales antes de guardar
                    if data.get('total_minutos', 0) > 44 * 60:
                         return JsonResponse({
                             'status': 'error', 
                             'message': 'No se puede exceder el límite de 44 horas semanales.'
                         }, status=400)

                    DiaHorario.objects.update_or_create(
                        horario=horario_base,
                        dia_semana=dia_semana,
                        defaults={
                            'activo': activo,
                            'hora_entrada': hora_entrada,
                            'hora_salida': hora_salida
                        }
                    )
                
                # Recalcular todos los registros de asistencia para este usuario
                # Esto permite que la vista 'mi_asistencia' refleje instantáneamente el nuevo horario
                for registro in RegistroAsistencia.objects.filter(funcionario=usuario):
                    registro.save()

            return JsonResponse({'status': 'success', 'message': 'Horario semanal guardado correctamente.'})
            
        except Exception as e:
            logger.error(f"Error guardando horario semanal para {user_id}: {str(e)}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


class GestionHorariosExcepcionalesView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para gestionar horarios excepcionales globales del establecimiento"""
    template_name = 'asistencia/gestion_excepcionales.html'

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['excepcionales'] = HorarioExcepcional.objects.select_related('creado_por').all()
        context['form'] = HorarioExcepcionalForm()
        return context


class CrearHorarioExcepcionalView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para crear un horario excepcional y recalcular los registros del día afectado"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def post(self, request):
        form = HorarioExcepcionalForm(request.POST)
        if form.is_valid():
            excepcional = form.save(commit=False)
            excepcional.creado_por = request.user
            excepcional.save()

            # Recalcular todos los registros de ese día específico
            registros_del_dia = RegistroAsistencia.objects.filter(fecha=excepcional.fecha)
            count = 0
            for registro in registros_del_dia:
                registro.save()
                count += 1

            registrar_log(
                usuario=request.user,
                tipo='CREATE',
                accion='Creación de Horario Excepcional',
                descripcion=f'Se creó horario excepcional para {excepcional.fecha}: {excepcional.motivo}. '
                            f'Se recalcularon {count} registros.',
                ip_address=get_client_ip(request)
            )

            messages.success(
                request,
                f'Horario excepcional creado para el {excepcional.fecha.strftime("%d/%m/%Y")}. '
                f'Se recalcularon {count} registros de asistencia.'
            )
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')

        return redirect('asistencia:gestion_excepcionales')


class EliminarHorarioExcepcionalView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para eliminar un horario excepcional y recalcular los registros del día afectado"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def post(self, request, pk):
        excepcional = get_object_or_404(HorarioExcepcional, pk=pk)
        fecha = excepcional.fecha
        motivo = excepcional.motivo
        excepcional.delete()

        # Recalcular registros del día ahora que ya no hay excepción
        registros_del_dia = RegistroAsistencia.objects.filter(fecha=fecha)
        count = 0
        for registro in registros_del_dia:
            registro.save()
            count += 1

        registrar_log(
            usuario=request.user,
            tipo='DELETE',
            accion='Eliminación de Horario Excepcional',
            descripcion=f'Se eliminó horario excepcional para {fecha}: {motivo}. '
                        f'Se recalcularon {count} registros.',
            ip_address=get_client_ip(request)
        )

        messages.success(
            request,
            f'Horario excepcional del {fecha.strftime("%d/%m/%Y")} eliminado. '
            f'Se recalcularon {count} registros de asistencia.'
        )
        return redirect('asistencia:gestion_excepcionales')


class ReporteAsistenciaAdministrativaView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para generar reporte administrativo de asistencia (días administrativos)"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get(self, request, anio=None, mes=None):
        import calendar as cal
        # Si no se pasan como parámetros de URL, obtener de GET
        if not anio or not mes or anio == '0':
            anio_str = request.GET.get('anio')
            mes_str = request.GET.get('mes')

            if anio_str and mes_str:
                try:
                    anio = int(anio_str)
                    mes = int(mes_str)
                except ValueError:
                    from django.contrib import messages
                    messages.error(request, 'Los valores de mes y año deben ser números válidos.')
                    return redirect(reverse('asistencia:gestion_asistencia'))
            else:
                from django.contrib import messages
                messages.error(request, 'Debe seleccionar mes y año para generar el reporte.')
                return redirect(reverse('asistencia:gestion_asistencia'))

        # Obtener permisos administrativos del mes
        from permisos.models import SolicitudPermiso
        permisos = SolicitudPermiso.objects.filter(
            estado='APROBADO',
            fecha_inicio__year=anio,
            fecha_inicio__month=mes
        ).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')

        # Preparar datos
        empleados_data = []
        for permiso in permisos:
            empleados_data.append({
                'funcionario': permiso.usuario.get_full_name() or permiso.usuario.username,
                'run': permiso.usuario.run,
                'establecimiento': 'Colegio Los Alerces',
                'dias_solicitados': permiso.dias_solicitados,
                'dias_disponibles': permiso.usuario.dias_disponibles if permiso.usuario.dias_disponibles else 0,
                'fecha_desde': permiso.fecha_inicio,
                'fecha_hasta': permiso.fecha_termino,
            })

        # Renderizar template HTML para PDF
        html_content = render_to_string('asistencia/reporte_administrativo_pdf.html', {
            'empleados_data': empleados_data,
            'anio': anio,
            'mes': mes,
            'mes_nombre': {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}.get(mes, ''),
            'fecha_actual': datetime.now(),
        })

        # Generar PDF
        pdf_file = HTML(string=html_content).write_pdf()

        # Crear respuesta HTTP
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'reporte_administrativo_{anio}_{mes:02d}.pdf'
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        return response


class ReporteAsistenciaAdministrativaExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para generar reporte administrativo de asistencia en Excel"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get(self, request):
        year = request.GET.get('year', '')
        mes = request.GET.get('mes', '')

        # Obtener permisos administrativos
        from permisos.models import SolicitudPermiso
        permisos = SolicitudPermiso.objects.filter(
            estado='APROBADO'
        ).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')

        if year:
            permisos = permisos.filter(fecha_inicio__year=year)
        if mes:
            permisos = permisos.filter(fecha_inicio__month=mes)

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Administrativo {mes}-{year}" if mes and year else "Administrativo"

        # Headers
        headers = ['N°', 'Funcionario', 'RUN', 'Establecimiento', 'Días Solicitados', 'Días Disponibles', 'Fecha Desde', 'Fecha Hasta']
        ws.append(headers)

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12

        for i, permiso in enumerate(permisos, 1):
            ws.append([
                i,
                permiso.usuario.get_full_name() or permiso.usuario.username,
                permiso.usuario.run,
                'Colegio Los Alerces',
                permiso.dias_solicitados,
                permiso.usuario.dias_disponibles if permiso.usuario.dias_disponibles else 0,
                permiso.fecha_inicio.strftime("%d/%m/%Y") if permiso.fecha_inicio else "",
                permiso.fecha_termino.strftime("%d/%m/%Y") if permiso.fecha_termino else "",
            ])

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"reporte_administrativo"
        if mes and year:
            filename += f"_{mes}_{year}"
        elif year:
            filename += f"_{year}"
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

        wb.save(response)
        return response


class ReporteDAEM3View(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para generar reporte DAEM3 (asistencia laboral por tipo de funcionario)"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get(self, request, anio=None, mes=None):
        import calendar as cal
        # Si no se pasan como parámetros de URL, obtener de GET
        if not anio or not mes or anio == '0':
            anio_str = request.GET.get('year', '') or request.GET.get('anio', '')
            mes_str = request.GET.get('mes')
            tipo = request.GET.get('tipo', 'DOCENTE')  # DOCENTE o ASISTENTE

            if anio_str and mes_str:
                try:
                    anio = int(anio_str)
                    mes = int(mes_str)
                except ValueError:
                    from django.contrib import messages
                    messages.error(request, 'Los valores de mes y año deben ser números válidos.')
                    return redirect(reverse('asistencia:gestion_asistencia'))
            else:
                from django.contrib import messages
                messages.error(request, 'Debe seleccionar mes y año para generar el reporte.')
                return redirect(reverse('asistencia:gestion_asistencia'))
        else:
            tipo = request.GET.get('tipo', 'DOCENTE')

        # Obtener funcionarios del tipo especificado
        from users.models import CustomUser
        if tipo == 'DOCENTE':
            funcionarios = CustomUser.objects.filter(
                role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'],
                tipo_funcionario='DOCENTE'
            ).order_by('first_name')
            tipo_descripcion = "Docentes"
        else:  # ASISTENTE
            funcionarios = CustomUser.objects.filter(
                role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'],
                tipo_funcionario='ASISTENTE'
            ).order_by('first_name')
            tipo_descripcion = "Asistentes De La Educación"

        # Obtener datos del mes
        registros_mes = RegistroAsistencia.objects.filter(
            fecha__year=anio,
            fecha__month=mes,
            funcionario__in=funcionarios
        ).select_related('funcionario')

        # Crear mapa de registros por funcionario
        registros_por_funcionario = {}
        for registro in registros_mes:
            func_id = registro.funcionario.id
            if func_id not in registros_por_funcionario:
                registros_por_funcionario[func_id] = []
            registros_por_funcionario[func_id].append(registro)

        # Procesar cada funcionario
        empleados_data = []
        for funcionario in funcionarios:
            func_id = funcionario.id
            registros_funcionario = registros_por_funcionario.get(func_id, [])

            atrasos_total = 0
            inasistencias_injustificadas = 0

            # Procesar registros del funcionario
            for registro in registros_funcionario:
                if registro.estado == 'RETRASO':
                    atrasos_total += registro.minutos_retraso or 0
                elif registro.estado == 'AUSENTE':
                    # Solo contar inasistencias injustificadas (no permisos, licencias, etc.)
                    if registro.fecha < funcionario.date_joined.date():
                        continue
                    inasistencias_injustificadas += 1

            # Detectar días sin registro que son inasistencias injustificadas
            fechas_con_registro = {r.fecha for r in registros_funcionario}
            today = datetime.now().date()
            num_dias = cal.monthrange(anio, mes)[0]
            for dia in range(1, num_dias + 1):
                fecha = datetime(anio, mes, dia).date()
                if fecha >= today:
                    continue
                if fecha in fechas_con_registro:
                    continue
                if DiaFestivo.objects.filter(fecha=fecha).exists():
                    continue
                if fecha.weekday() >= 5 and not (funcionario.funcion == 'SERENO' or funcionario.tipo_funcionario == 'SERENO'):
                    continue
                if fecha < funcionario.date_joined.date():
                    continue
                inasistencias_injustificadas += 1

            # Solo incluir si tiene registros o ausencias
            if registros_funcionario or inasistencias_injustificadas > 0:
                empleados_data.append({
                    'nombre_completo': funcionario.get_full_name() or funcionario.username,
                    'run': funcionario.run,
                    'atrasos': atrasos_total,
                    'inasistencias': inasistencias_injustificadas,
                })

        # Nombre del mes
        meses = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]
        nombre_mes = meses[mes - 1]

        # Renderizar template HTML para PDF
        html_content = render_to_string('asistencia/reporte_daem3_pdf.html', {
            'empleados_data': empleados_data,
            'anio': anio,
            'mes': mes,
            'nombre_mes': nombre_mes,
            'tipo_descripcion': tipo_descripcion,
            'fecha_actual': datetime.now(),
        })

        # Generar PDF
        pdf_file = HTML(string=html_content).write_pdf()

        # Crear respuesta HTTP
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'informe_asistencia_laboral_{tipo_descripcion.lower().replace(" ", "_")}_{anio}_{mes:02d}.pdf'
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        return response


class ReporteDAEM3ExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para generar reporte DAEM3 en Excel"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get(self, request):
        year = request.GET.get('year', '') or request.GET.get('anio', '')
        mes = request.GET.get('mes', '')
        tipo = request.GET.get('tipo', 'DOCENTE')

        # Obtener funcionarios del tipo especificado
        from users.models import CustomUser
        if tipo == 'DOCENTE':
            funcionarios = CustomUser.objects.filter(
                role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'],
                tipo_funcionario='DOCENTE'
            ).order_by('first_name')
            tipo_descripcion = "Docentes"
        else:  # ASISTENTE
            funcionarios = CustomUser.objects.filter(
                role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'],
                tipo_funcionario='ASISTENTE'
            ).order_by('first_name')
            tipo_descripcion = "Asistentes de la Educación"

        # Obtener datos del mes
        registros_mes = RegistroAsistencia.objects.filter(
            fecha__year=year,
            fecha__month=mes,
            funcionario__in=funcionarios
        ).select_related('funcionario')

        # Crear mapa de registros por funcionario
        registros_por_funcionario = {}
        for registro in registros_mes:
            func_id = registro.funcionario.id
            if func_id not in registros_por_funcionario:
                registros_por_funcionario[func_id] = []
            registros_por_funcionario[func_id].append(registro)

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Informe Asistencia {tipo_descripcion}"

        # Headers
        headers = ['N°', 'Nombre y Apellidos', 'RUN', 'Atrasos (min)', 'Inasistencias']
        ws.append(headers)

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Procesar cada funcionario
        empleados_data = []
        for funcionario in funcionarios:
            func_id = funcionario.id
            registros_funcionario = registros_por_funcionario.get(func_id, [])

            atrasos_total = 0
            inasistencias_injustificadas = 0

            # Procesar registros del funcionario
            for registro in registros_funcionario:
                if registro.estado == 'RETRASO':
                    atrasos_total += registro.minutos_retraso or 0
                elif registro.estado == 'AUSENTE':
                    # Solo contar inasistencias injustificadas
                    if registro.fecha < funcionario.date_joined.date():
                        continue
                    inasistencias_injustificadas += 1

            # Detectar días sin registro que son inasistencias injustificadas
            import calendar as cal
            fechas_con_registro = {r.fecha for r in registros_funcionario}
            today = datetime.now().date()
            num_dias = cal.monthrange(int(year), int(mes))[0]
            for dia in range(1, num_dias + 1):
                fecha = datetime(int(year), int(mes), dia).date()
                if fecha >= today:
                    continue
                if fecha in fechas_con_registro:
                    continue
                if DiaFestivo.objects.filter(fecha=fecha).exists():
                    continue
                if fecha.weekday() >= 5 and not (funcionario.funcion == 'SERENO' or funcionario.tipo_funcionario == 'SERENO'):
                    continue
                if fecha < funcionario.date_joined.date():
                    continue
                inasistencias_injustificadas += 1

            # Solo incluir si tiene registros o ausencias
            if registros_funcionario or inasistencias_injustificadas > 0:
                empleados_data.append({
                    'nombre_completo': funcionario.get_full_name() or funcionario.username,
                    'run': funcionario.run,
                    'atrasos': atrasos_total,
                    'inasistencias': inasistencias_injustificadas,
                })

        # Agregar datos a Excel
        for i, empleado in enumerate(empleados_data, 1):
            ws.append([
                i,
                empleado['nombre_completo'],
                empleado['run'],
                empleado['atrasos'],
                empleado['inasistencias']
            ])

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"informe_asistencia_laboral_{tipo_descripcion.lower().replace(' ', '_')}_{year}_{mes}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response


class ReporteAsistenciaMensualExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para generar reporte mensual de asistencia en Excel"""

    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA', 'DIRECTOR', 'DIRECTIVO']

    def get(self, request, anio=None, mes=None):
        import calendar as cal
        # Si no se pasan como parámetros de URL, obtener de GET
        if not anio or not mes or anio == '0':
            anio_str = request.GET.get('anio')
            mes_str = request.GET.get('mes')

            if anio_str and mes_str:
                try:
                    anio = int(anio_str)
                    mes = int(mes_str)
                except ValueError:
                    from django.contrib import messages
                    messages.error(request, 'Los valores de mes y año deben ser números válidos.')
                    return redirect(reverse('asistencia:gestion_asistencia'))
            else:
                from django.contrib import messages
                messages.error(request, 'Debe seleccionar mes y año para generar el reporte.')
                return redirect(reverse('asistencia:gestion_asistencia'))

        # Obtener todos los funcionarios que deben tener asistencia
        from users.models import CustomUser
        todos_funcionarios = CustomUser.objects.filter(
            role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN']
        ).order_by('last_name', 'first_name')

        # Obtener datos del mes
        registros_mes = RegistroAsistencia.objects.filter(
            fecha__year=anio,
            fecha__month=mes
        ).select_related('funcionario', 'horario_asignado')

        # Crear mapa de registros por funcionario
        registros_por_funcionario = {}
        for registro in registros_mes:
            func_id = registro.funcionario.id
            if func_id not in registros_por_funcionario:
                registros_por_funcionario[func_id] = []
            registros_por_funcionario[func_id].append(registro)

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Asistencia {mes:02d}-{anio}"

        # Encabezados
        headers = ['Nombre Completo', 'RUN', 'Cargo', 'Fecha', 'Tipo', 'Detalle', 'Total Atrasos', 'Min. Retraso', 'Inasistencias']
        ws.append(headers)

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15

        # Procesar cada funcionario
        for funcionario in todos_funcionarios:
            func_id = funcionario.id
            registros_funcionario = registros_por_funcionario.get(func_id, [])

            # Inicializar datos del funcionario
            atrasos = []
            inasistencias = []
            total_atrasos = 0
            total_minutos_retraso = 0
            total_inasistencias = 0

            # Procesar registros del funcionario
            for registro in registros_funcionario:
                if registro.estado == 'RETRASO':
                    atraso_info = {
                        'fecha': registro.fecha,
                        'hora_entrada': registro.hora_entrada_real,
                        'minutos_retraso': registro.minutos_retraso,
                    }
                    atrasos.append(atraso_info)
                    total_atrasos += 1
                    total_minutos_retraso += registro.minutos_retraso or 0
                elif registro.estado == 'AUSENTE':
                    # Ignorar si es antes de su ingreso
                    if registro.fecha < funcionario.date_joined.date():
                        continue

                    inasistencia_info = {
                        'fecha': registro.fecha,
                        'hora_esperada': registro.horario_asignado.hora_entrada if registro.horario_asignado else None,
                    }
                    inasistencias.append(inasistencia_info)
                    total_inasistencias += 1

            # Detectar días sin registro que son inasistencias
            fechas_con_registro = {r.fecha for r in registros_funcionario}
            today = datetime.now().date()
            num_dias = cal.monthrange(anio, mes)[0]
            for dia in range(1, num_dias + 1):
                fecha = datetime(anio, mes, dia).date()
                if fecha >= today:
                    continue
                if fecha in fechas_con_registro:
                    continue
                if DiaFestivo.objects.filter(fecha=fecha).exists():
                    continue
                if fecha.weekday() >= 5 and not (funcionario.funcion == 'SERENO' or funcionario.tipo_funcionario == 'SERENO'):
                    continue
                if fecha < funcionario.date_joined.date():
                    continue
                inasistencia_info = {
                    'fecha': fecha,
                    'hora_esperada': None,
                }
                inasistencias.append(inasistencia_info)
                total_inasistencias += 1

            # Solo incluir si tiene atrasos o inasistencias
            if atrasos or inasistencias:
                # Agregar filas para atrasos
                for atraso in atrasos:
                    ws.append([
                        funcionario.get_full_name(),
                        funcionario.run,
                        funcionario.get_funcion_display() or "",
                        atraso['fecha'].strftime("%d/%m/%Y"),
                        "Atraso",
                        f"{atraso['hora_entrada'].strftime('%H:%M') if atraso['hora_entrada'] else '-'} - {atraso['minutos_retraso']} min retraso",
                        total_atrasos,
                        total_minutos_retraso,
                        total_inasistencias
                    ])

                # Agregar filas para inasistencias
                for inasistencia in inasistencias:
                    ws.append([
                        funcionario.get_full_name(),
                        funcionario.run,
                        funcionario.get_funcion_display() or "",
                        inasistencia['fecha'].strftime("%d/%m/%Y"),
                        "Inasistencia",
                        f"Sin justificar - Hora esperada: {inasistencia['hora_esperada'].strftime('%H:%M') if inasistencia['hora_esperada'] else '-'}",
                        total_atrasos,
                        total_minutos_retraso,
                        total_inasistencias
                    ])

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill

        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'reporte_atrasos_inasistencias_{anio}_{mes:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response
