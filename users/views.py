from django.views.generic import ListView, CreateView, UpdateView, DeleteView, FormView, TemplateView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.core.management import call_command
from django.core.management import call_command
from django import forms
from .models import CustomUser
from .forms import UserCreateForm, UserEditForm, BulkUserImportForm
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random
import string
from io import BytesIO

from core.security import audit_log
from admin_dashboard.utils import registrar_log, get_client_ip

from django.db.models import Q

class UserListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = CustomUser
    template_name = 'users/user_list.html'
    context_object_name = 'users'
    paginate_by = None  # Sin paginación - mostrar todos los usuarios
    
    def test_func(self):
        return self.request.user.role in ['SECRETARIA', 'ADMIN']
    
    def get_queryset(self):
        # Orden por defecto: Más reciente primero
        queryset = CustomUser.objects.all().select_related('horario').order_by('-date_joined')
        
        # Búsqueda
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) | 
                Q(last_name__icontains=search_query) | 
                Q(run__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        
        # Aplicar ordenamiento si se solicita
        sort_by = self.request.GET.get('sort', 'recent')
        
        if sort_by == 'name':
            queryset = queryset.order_by('first_name', 'last_name')
        elif sort_by == 'name_desc':
            queryset = queryset.order_by('-first_name', '-last_name')
        elif sort_by == 'role':
            queryset = queryset.order_by('role', 'first_name')
        elif sort_by == 'role_desc':
            queryset = queryset.order_by('-role', 'first_name')
        elif sort_by == 'tipo':
            queryset = queryset.order_by('tipo_funcionario', 'first_name')
        elif sort_by == 'tipo_desc':
            queryset = queryset.order_by('-tipo_funcionario', 'first_name')
        elif sort_by == 'dias':
            queryset = queryset.order_by('-dias_disponibles')
        elif sort_by == 'dias_asc':
            queryset = queryset.order_by('dias_disponibles')
        elif sort_by == 'recent':
            queryset = queryset.order_by('-date_joined')
        else:
            queryset = queryset.order_by('-date_joined')
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'recent')
        
        # Recuperar contraseñas de importación masiva y limpiar sesión
        if 'bulk_passwords' in self.request.session:
            context['bulk_passwords'] = self.request.session.pop('bulk_passwords')
            
        return context

class UserCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = CustomUser
    form_class = UserCreateForm
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return self.request.user.role in ['SECRETARIA', 'ADMIN']
    
    def form_valid(self, form):
        response = super().form_valid(form)
        user = self.object
        registrar_log(
            usuario=self.request.user,
            tipo='CREATE',
            accion='Creación de Usuario',
            descripcion=f'Se creó usuario: {user.run} - {user.email}',
            ip_address=get_client_ip(self.request)
        )
        return response

class UserUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    form_class = UserEditForm
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return self.request.user.role in ['SECRETARIA', 'ADMIN']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Pass the editing user to the form
        kwargs['editing_user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        registrar_log(
            usuario=self.request.user,
            tipo='UPDATE',
            accion='Actualización de Usuario',
            descripcion=f'Se actualizó usuario: {self.object.run} - {self.object.email}',
            ip_address=get_client_ip(self.request)
        )
        return super().form_valid(form)


class UserDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = CustomUser
    template_name = 'users/user_confirm_delete.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return self.request.user.role == 'ADMIN'
    
    def delete(self, request, *args, **kwargs):
        user_to_delete = self.get_object()
        registrar_log(
            usuario=self.request.user,
            tipo='DELETE',
            accion='Eliminación de Usuario',
            descripcion=f'Se eliminó usuario: {user_to_delete.run} - {user_to_delete.email}',
            ip_address=get_client_ip(self.request)
        )
        messages.success(self.request, 'Usuario eliminado exitosamente')
        return super().delete(request, *args, **kwargs)


class BulkUserImportView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    """Vista para importación masiva de usuarios desde Excel"""
    template_name = 'users/bulk_import.html'
    form_class = BulkUserImportForm
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        # Solo admins y secretarias
        return self.request.user.role in ['ADMIN', 'SECRETARIA']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['download_template'] = True
        return context
    
    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        
        try:
            # Leer archivo Excel
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_users = []
            errors = []
            passwords = {}  # Almacenar contraseñas generadas
            
            # Procesar cada fila (empezar desde fila 2, asumiendo headers en fila 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                try:
                    # Rellenar row con None si tiene menos de 7 elementos
                    row_list = list(row) + [None] * (7 - len(row))
                    run, first_name, last_name, email, role, tipo_funcionario, dias = row_list[:7]
                    
                    # Validaciones básicas
                    if not run or not first_name or not last_name:
                        errors.append(f"Fila {row_num}: Faltan datos obligatorios (RUN, nombre o apellido)")
                        continue
                    
                    # Validar email
                    if email and not email.endswith('@losalercespuertomontt.cl'):
                        errors.append(f"Fila {row_num}: Email debe ser @losalercespuertomontt.cl")
                        continue
                    
                    # Crear email si no existe
                    if not email:
                        email = f"{run}@losalercespuertomontt.cl"
                    
                    # Verificar si ya existe
                    if CustomUser.objects.filter(run=run).exists():
                        errors.append(f"Fila {row_num}: Usuario con RUN {run} ya existe")
                        continue
                    
                    # Generar contraseña
                    password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                    
                    # Crear usuario
                    user = CustomUser.objects.create_user(
                        run=run,
                        username=run,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        password=password,
                        role=role or 'FUNCIONARIO',
                        tipo_funcionario=tipo_funcionario or 'PLANTA',
                        dias_disponibles=dias or 6.0
                    )
                    
                    # Crear horario por defecto (07:45 AM con 15 min tolerancia)
                    from asistencia.models import HorarioFuncionario
                    from datetime import time
                    HorarioFuncionario.objects.create(
                        funcionario=user,
                        hora_entrada=time(7, 45),
                        tolerancia_minutos=15,
                        activo=True
                    )
                    
                    created_users.append(user)
                    passwords[run] = password
                    
                except Exception as e:
                    errors.append(f"Fila {row_num}: Error - {str(e)}")
            
            # Resultado
            if created_users:
                messages.success(
                    self.request,
                    f"Se crearon {len(created_users)} usuarios exitosamente"
                )
                
                # Guardar contraseñas para mostrar
                self.request.session['bulk_passwords'] = passwords
                
                registrar_log(
                    usuario=self.request.user,
                    tipo='IMPORT',
                    accion='Importación Masiva de Usuarios',
                    descripcion=f'Se importaron exitosamente {len(created_users)} usuarios desde Excel',
                    ip_address=get_client_ip(self.request)
                )
                
            if errors:
                for error in errors:
                    messages.warning(self.request, error)
            
            return redirect(self.success_url)
            
        except Exception as e:
            messages.error(self.request, f"Error al procesar archivo: {str(e)}")
            return self.form_invalid(form)


def download_template(request):
    """Generar y descargar plantilla Excel para importación"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Headers
    headers = ['RUN', 'Nombres', 'Apellidos', 'Email', 'Rol', 'Tipo Funcionario', 'Días Disponibles']
    ws.append(headers)
    
    # Ejemplo
    ws.append([
        '12345678-9',
        'Juan',
        'Pérez',
        'juan.perez@losalercespuertomontt.cl',
        'FUNCIONARIO',
        'PLANTA',
        6.0
    ])
    
    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_usuarios.xlsx'
    
    return response


class ResetUserPasswordView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """
    Vista para que administradores reseteen la contraseña de un usuario.
    Genera una nueva contraseña temporal y la muestra al administrador.
    
    Permisos:
    - ADMIN: puede resetear cualquier contraseña
    - SECRETARIA: puede resetear contraseñas de FUNCIONARIOS, DIRECTIVOS y SECRETARIAS
                  (NO puede resetear DIRECTOR ni ADMIN)
    """
    template_name = 'users/reset_password.html'
    
    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA']
    
    def can_reset_password(self, target_user):
        """Verifica si el usuario actual puede resetear la contraseña del usuario objetivo"""
        current_user = self.request.user
        
        # ADMIN puede resetear cualquier contraseña
        if current_user.role == 'ADMIN':
            return True
        
        # SECRETARIA puede resetear solo de FUNCIONARIO, DIRECTIVO y SECRETARIA
        if current_user.role == 'SECRETARIA':
            forbidden_roles = ['DIRECTOR', 'ADMIN']
            return target_user.role not in forbidden_roles
        
        return False
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_id = self.kwargs.get('user_id')
        target_user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar permisos
        if not self.can_reset_password(target_user):
            if self.request.user.role == 'SECRETARIA':
                messages.error(
                    self.request,
                    'No tienes permisos para resetear la contraseña de directores o administradores.'
                )
            else:
                messages.error(self.request, 'No tienes permisos para realizar esta acción.')
            context['permission_denied'] = True
            return context
        
        context['user'] = target_user
        return context
    
    def post(self, request, *args, **kwargs):
        user_id = self.kwargs.get('user_id')
        target_user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar permisos
        if not self.can_reset_password(target_user):
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('user_list')
        
        # Generar nueva contraseña temporal basada en RUN (sin puntos ni guión)
        new_password = target_user.run.replace('.', '').replace('-', '')
        target_user.set_password(new_password)
        target_user.save()
        
        # Mantener la sesión activa si se está reseteando la propia contraseña
        if target_user == request.user:
            update_session_auth_hash(request, target_user)
        
        registrar_log(
            usuario=self.request.user,
            tipo='UPDATE',
            accion='Reset Password Admin',
            descripcion=f'Admin reseteó contraseña para: {target_user.run}',
            ip_address=get_client_ip(self.request)
        )
        
        messages.success(
            request,
            f'Contraseña de {target_user.get_full_name()} ha sido restablecida a su RUN (sin puntos ni guión). '
            f'Nueva contraseña temporal: {new_password}'
        )
        return redirect('user_list')


class ChangeOwnPasswordView(LoginRequiredMixin, TemplateView):
    """
    Vista para que los usuarios cambien su propia contraseña.
    Mantiene la sesión activa después del cambio.
    """
    template_name = 'users/change_password.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = PasswordChangeForm(self.request.user)
        return context
    
    def post(self, request, *args, **kwargs):
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Mantener la sesión activa después de cambiar la contraseña
            update_session_auth_hash(request, user)
            
            registrar_log(
                usuario=request.user,
                tipo='UPDATE',
                accion='Cambio Propia Contraseña',
                descripcion='Usuario cambió su propia contraseña',
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            return redirect('dashboard')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
        
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)


class SetPasswordForm(forms.Form):
    """Formulario para establecer una contraseña específica"""
    password1 = forms.CharField(
        label='Nueva contraseña',
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
        min_length=8
    )
    password2 = forms.CharField(
        label='Confirmar contraseña',
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
        min_length=8
    )
    
    def clean(self):
        cleaned_data = super().clean()
        password1 = cleaned_data.get('password1')
        password2 = cleaned_data.get('password2')
        
        if password1 and password2:
            if password1 != password2:
                raise forms.ValidationError("Las contraseñas no coinciden.")
            if len(password1) < 8:
                raise forms.ValidationError("La contraseña debe tener al menos 8 caracteres.")
        
        return cleaned_data


class AdminChangePasswordView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """
    Vista para que ADMIN y SECRETARIA cambien la contraseña de un usuario.
    
    Permisos:
    - ADMIN: puede cambiar cualquier contraseña
    - SECRETARIA: puede cambiar contraseñas de FUNCIONARIOS, DIRECTIVOS y SECRETARIAS
                  (NO puede cambiar DIRECTOR ni ADMIN)
    """
    template_name = 'users/admin_change_password.html'
    
    def test_func(self):
        return self.request.user.role in ['ADMIN', 'SECRETARIA']
    
    def can_change_password(self, target_user):
        """Verifica si el usuario actual puede cambiar la contraseña del usuario objetivo"""
        current_user = self.request.user
        
        # ADMIN puede cambiar cualquier contraseña
        if current_user.role == 'ADMIN':
            return True
        
        # SECRETARIA puede cambiar solo de FUNCIONARIO, DIRECTIVO y SECRETARIA
        if current_user.role == 'SECRETARIA':
            forbidden_roles = ['DIRECTOR', 'ADMIN']
            return target_user.role not in forbidden_roles
        
        return False
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_id = self.kwargs.get('user_id')
        target_user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar permisos
        if not self.can_change_password(target_user):
            if self.request.user.role == 'SECRETARIA':
                messages.error(
                    self.request,
                    'No tienes permisos para cambiar la contraseña de directores o administradores.'
                )
            else:
                messages.error(self.request, 'No tienes permisos para realizar esta acción.')
            context['permission_denied'] = True
            return context
        
        context['target_user'] = target_user
        context['form'] = SetPasswordForm()
        return context
    
    def post(self, request, *args, **kwargs):
        user_id = self.kwargs.get('user_id')
        target_user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar permisos
        if not self.can_change_password(target_user):
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('user_list')
        
        form = SetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['password1']
            target_user.set_password(new_password)
            target_user.save()
            
            messages.success(
                request,
                f'La contraseña de {target_user.get_full_name()} ha sido cambiada exitosamente.'
            )
            return redirect('user_list')
        
        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


class EmailDirectoryView(LoginRequiredMixin, ListView):
    """Vista para que todos los funcionarios vean el directorio de correos institucionales"""
    model = CustomUser
    template_name = 'users/email_directory.html'
    context_object_name = 'users'

    def get_template_names(self):
        return ['users/email_directory.html']

    def get_queryset(self):
        queryset = CustomUser.objects.filter(is_active=True).order_by('first_name', 'last_name')

        # Búsqueda
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['active_tab'] = self.request.GET.get('tab', 'usuarios')

        # Obtener grupos de correo
        from users.models import GrupoCorreo
        grupos = GrupoCorreo.objects.filter(activo=True).order_by('nombre')

        grupo_search = self.request.GET.get('grupo_search', '')
        if grupo_search:
            grupos = grupos.filter(
                Q(nombre__icontains=grupo_search) |
                Q(correo__icontains=grupo_search) |
                Q(descripcion__icontains=grupo_search)
            )

        context['grupos'] = grupos
        context['grupo_search'] = grupo_search

        # Teléfonos: usuarios con teléfono registrado
        phone_search = self.request.GET.get('phone_search', '')
        telefonos = CustomUser.objects.filter(
            is_active=True,
            telefono__gt=''
        ).order_by('first_name', 'last_name')
        
        if phone_search:
            telefonos = telefonos.filter(
                Q(first_name__icontains=phone_search) |
                Q(last_name__icontains=phone_search) |
                Q(telefono__icontains=phone_search)
            )
        
        context['telefonos'] = telefonos
        context['phone_search'] = phone_search

        # Directorio Telefónico (lugar y anexo)
        from users.models import DirectorioTelefonico
        directorio = DirectorioTelefonico.objects.filter(activo=True).order_by('lugar')
        
        directorio_search = self.request.GET.get('directorio_search', '')
        if directorio_search:
            directorio = directorio.filter(
                Q(lugar__icontains=directorio_search) |
                Q(anexo__icontains=directorio_search)
            )
        
        context['directorio'] = directorio
        context['directorio_search'] = directorio_search

        return context



class CrearGrupoCorreoView(LoginRequiredMixin, View):
    """Vista para que ADMIN cree grupos de correo desde el directorio"""
    
    def post(self, request):
        
        # Only ADMIN can create groups
        if request.user.role != 'ADMIN':
            messages.error(request, 'No tienes permisos para crear grupos de correo.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        nombre = request.POST.get('nombre', '').strip()
        correo = request.POST.get('correo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        if not nombre or not correo:
            messages.error(request, 'El nombre y correo son requeridos.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        from users.models import GrupoCorreo
        
        # Check if already exists
        if GrupoCorreo.objects.filter(correo__iexact=correo).exists():
            messages.error(request, 'Ya existe un grupo con ese correo.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        if GrupoCorreo.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, 'Ya existe un grupo con ese nombre.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        try:
            grupo = GrupoCorreo.objects.create(
                nombre=nombre,
                correo=correo,
                descripcion=descripcion,
                creado_por=request.user,
                activo=True
            )
            messages.success(request, f'Grupo "{grupo.nombre}" creado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al crear el grupo: {str(e)}')
        
        return redirect(reverse('email_directory') + '?tab=grupos')


class CrearDirectorioTelefonicoView(LoginRequiredMixin, View):
    """Vista para que ADMIN cree entradas del directorio telefónico desde el directorio"""
    
    def post(self, request):
        
        if request.user.role != 'ADMIN':
            messages.error(request, 'No tienes permisos para agregar números telefónicos.')
            return redirect('email_directory')
        
        lugar = request.POST.get('lugar', '').strip()
        anexo = request.POST.get('anexo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        if not lugar or not anexo:
            messages.error(request, 'El lugar y anexo son requeridos.')
            return redirect(reverse('email_directory') + '?tab=telefonos')
        
        from users.models import DirectorioTelefonico
        
        if DirectorioTelefonico.objects.filter(anexo__iexact=anexo).exists():
            messages.error(request, 'Ya existe un anexo con ese número.')
            return redirect(reverse('email_directory') + '?tab=telefonos')
        
        try:
            telefono = DirectorioTelefonico.objects.create(
                lugar=lugar,
                anexo=anexo,
                descripcion=descripcion,
                creado_por=request.user,
                activo=True
            )
            messages.success(request, f'Teléfono "{telefono.lugar}" ({telefono.anexo}) creado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al crear el teléfono: {str(e)}')
        
        return redirect(reverse('email_directory') + '?tab=telefonos')


class EditarGrupoCorreoView(LoginRequiredMixin, View):
    """Vista para que ADMIN edite grupos de correo"""
    
    def post(self, request, grupo_id):
        if request.user.role != 'ADMIN':
            messages.error(request, 'No tienes permisos para editar grupos de correo.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        from users.models import GrupoCorreo
        
        try:
            grupo = GrupoCorreo.objects.get(pk=grupo_id)
        except GrupoCorreo.DoesNotExist:
            messages.error(request, 'Grupo no encontrado.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        nombre = request.POST.get('nombre', '').strip()
        correo = request.POST.get('correo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        if not nombre or not correo:
            messages.error(request, 'El nombre y correo son requeridos.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        # Check if email exists in another group
        if GrupoCorreo.objects.exclude(pk=grupo_id).filter(correo__iexact=correo).exists():
            messages.error(request, 'Ya existe otro grupo con ese correo.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        if GrupoCorreo.objects.exclude(pk=grupo_id).filter(nombre__iexact=nombre).exists():
            messages.error(request, 'Ya existe otro grupo con ese nombre.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        try:
            grupo.nombre = nombre
            grupo.correo = correo
            grupo.descripcion = descripcion
            grupo.save()
            messages.success(request, f'Grupo "{grupo.nombre}" actualizado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar el grupo: {str(e)}')
        
        return redirect(reverse('email_directory') + '?tab=grupos')


class EliminarGrupoCorreoView(LoginRequiredMixin, View):
    """Vista para que ADMIN elimine (desactive) grupos de correo"""
    
    def post(self, request, grupo_id):
        if request.user.role != 'ADMIN':
            messages.error(request, 'No tienes permisos para eliminar grupos de correo.')
            return redirect(reverse('email_directory') + '?tab=grupos')
        
        from users.models import GrupoCorreo
        
        try:
            grupo = GrupoCorreo.objects.get(pk=grupo_id)
            grupo.activo = False
            grupo.save()
            messages.success(request, f'Grupo "{grupo.nombre}" eliminado exitosamente.')
        except GrupoCorreo.DoesNotExist:
            messages.error(request, 'Grupo no encontrado.')
        except Exception as e:
            messages.error(request, f'Error al eliminar el grupo: {str(e)}')
        
        return redirect(reverse('email_directory') + '?tab=grupos')


class EditarDirectorioTelefonicoView(LoginRequiredMixin, View):
    """Vista para que ADMIN edite entradas del directorio telefónico"""
    
    def post(self, request, telefono_id):
        if request.user.role != 'ADMIN':
            messages.error(request, 'No tienes permisos para editar números telefónicos.')
            return redirect(reverse('email_directory') + '?tab=telefonos')
        
        from users.models import DirectorioTelefonico
        
        try:
            telefono = DirectorioTelefonico.objects.get(pk=telefono_id)
        except DirectorioTelefonico.DoesNotExist:
            messages.error(request, 'Teléfono no encontrado.')
            return redirect(reverse('email_directory') + '?tab=telefonos')
        
        lugar = request.POST.get('lugar', '').strip()
        anexo = request.POST.get('anexo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        if not lugar or not anexo:
            messages.error(request, 'El lugar y anexo son requeridos.')
            return redirect(reverse('email_directory') + '?tab=telefonos')
        
        # Check if anexo exists in another entry
        if DirectorioTelefonico.objects.exclude(pk=telefono_id).filter(anexo__iexact=anexo).exists():
            messages.error(request, 'Ya existe otro teléfono con ese anexo.')
            return redirect(reverse('email_directory') + '?tab=telefonos')
        
        try:
            telefono.lugar = lugar
            telefono.anexo = anexo
            telefono.descripcion = descripcion
            telefono.save()
            messages.success(request, f'Teléfono "{telefono.lugar}" actualizado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar el teléfono: {str(e)}')
        
        return redirect(reverse('email_directory') + '?tab=telefonos')


class EliminarDirectorioTelefonicoView(LoginRequiredMixin, View):
    """Vista para que ADMIN elimine (desactive) entradas del directorio telefónico"""
    
    def post(self, request, telefono_id):
        if request.user.role != 'ADMIN':
            messages.error(request, 'No tienes permisos para eliminar números telefónicos.')
            return redirect(reverse('email_directory') + '?tab=telefonos')
        
        from users.models import DirectorioTelefonico
        
        try:
            telefono = DirectorioTelefonico.objects.get(pk=telefono_id)
            telefono.activo = False
            telefono.save()
            messages.success(request, f'Teléfono "{telefono.lugar}" eliminado exitosamente.')
        except DirectorioTelefonico.DoesNotExist:
            messages.error(request, 'Teléfono no encontrado.')
        except Exception as e:
            messages.error(request, f'Error al eliminar el teléfono: {str(e)}')
        
        return redirect(reverse('email_directory') + '?tab=telefonos')


class BackupExportUsersView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para exportar un respaldo exacto de todos los usuarios (incluyendo contraseñas)"""

    def test_func(self):
        return self.request.user.role == 'ADMIN'

    def get(self, request):
        audit_log(request, 'BACKUP_EXPORT', f'Admin {request.user.run} exported user backup')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Respaldo_Usuarios"
        
        headers = [
            'run', 'username', 'first_name', 'last_name', 'email', 
            'password_hash', 'role', 'tipo_funcionario', 
            'dias_disponibles', 'telefono', 'is_active', 'is_blocked'
        ]
        ws.append(headers)
        
        for user in CustomUser.objects.all().order_by('run'):
            ws.append([
                user.run,
                user.username,
                user.first_name,
                user.last_name,
                user.email,
                user.password,
                user.role,
                user.tipo_funcionario,
                user.dias_disponibles,
                user.telefono,
                str(user.is_active),
                str(user.is_blocked)
            ])
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from datetime import datetime
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=respaldo_usuarios_{fecha_str}.xlsx'
        return response


class BackupRestoreUsersView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista para importar un respaldo y restaurar la base de usuarios"""

    def test_func(self):
        return self.request.user.role == 'ADMIN'

    def get(self, request):
        return render(request, 'users/restore_backup.html')

    def post(self, request):
        if 'backup_file' not in request.FILES:
            messages.error(request, "No se seleccionó ningún archivo de respaldo.")
            return redirect('backup_restore_users')
            
        excel_file = request.FILES['backup_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue
                
                row_list = list(row) + [None] * (12 - len(row))
                run = row_list[0]
                if not run: continue
                
                is_active = str(row_list[10]).lower() == 'true' if row_list[10] is not None else True
                is_blocked = str(row_list[11]).lower() == 'true' if row_list[11] is not None else False
                
                defaults = {
                    'username': row_list[1] or run,
                    'first_name': row_list[2] or '',
                    'last_name': row_list[3] or '',
                    'email': row_list[4] or f"{run}@losalercespuertomontt.cl",
                    'password': row_list[5] or '',
                    'role': row_list[6] or 'FUNCIONARIO',
                    'tipo_funcionario': row_list[7] or 'PLANTA',
                    'dias_disponibles': float(row_list[8]) if row_list[8] is not None else 6.0,
                    'telefono': row_list[9] or '',
                    'is_active': is_active,
                    'is_blocked': is_blocked
                }
                
                user, created = CustomUser.objects.update_or_create(
                    run=run,
                    defaults=defaults
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            audit_log(request, 'BACKUP_RESTORED', f'Admin {request.user.run} restored user backup')
            messages.success(request, f"Respaldo restaurado: {created_count} creados, {updated_count} actualizados.")
            return redirect('user_list')
            
        except Exception as e:
            messages.error(request, f"Error al procesar archivo de respaldo: {str(e)}")
            return redirect('backup_restore_users')
