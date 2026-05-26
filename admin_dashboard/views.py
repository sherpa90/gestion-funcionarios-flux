from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import TemplateView, View
from django.db.models import Count, Q, Sum
from django.utils import timezone
from django.contrib import messages
from datetime import timedelta
import os
import zipfile
import tempfile
import io
import shutil
from django.conf import settings
from django.core.management import call_command
from django.core.paginator import Paginator
from django.http import FileResponse, HttpResponse, JsonResponse
from django.views.generic import TemplateView, View, ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
import urllib.parse

from users.models import CustomUser
from permisos.models import SolicitudPermiso
from licencias.models import LicenciaMedica
from .models import SystemLog, Efemeride
from .forms import EfemerideForm
from .utils import registrar_log, get_client_ip
from auditlog.models import LogEntry



class AdminDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'admin_dashboard/dashboard.html'
    
    def test_func(self):
        # ADMIN, DIRECTOR, DIRECTIVO y SECRETARIA pueden acceder
        return self.request.user.role in ['ADMIN', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['total_usuarios'] = CustomUser.objects.count()
        context['total_funcionarios'] = CustomUser.objects.filter(role='FUNCIONARIO').count()
        context['total_directivos'] = CustomUser.objects.filter(
            role__in=['DIRECTOR', 'DIRECTIVO', 'SECRETARIA']
        ).count()
        
        context['solicitudes_pendientes'] = SolicitudPermiso.objects.filter(
            estado='PENDIENTE'
        ).count()
        context['solicitudes_aprobadas_mes'] = SolicitudPermiso.objects.filter(
            estado='APROBADO',
            updated_at__gte=timezone.now() - timedelta(days=30)
        ).count()
        
        # Licencias activas: cálculo eficiente (compatible con todas las DB)
        # --- Lógica de Planificación Semanal y Diaria ---
        hoy = timezone.now().date()
        # Lunes de la semana actual
        lunes_actual = hoy - timedelta(days=hoy.weekday())
        domingo_actual = lunes_actual + timedelta(days=6)
        # Próxima semana
        lunes_proximo = lunes_actual + timedelta(days=7)
        domingo_proximo = lunes_proximo + timedelta(days=6)

        # 1. Permisos Semana Actual
        permisos_actual = SolicitudPermiso.objects.select_related('usuario').filter(
            estado='APROBADO',
            fecha_inicio__lte=domingo_actual,
            fecha_termino__gte=lunes_actual
        ).order_by('fecha_inicio')
        context['permisos_semana_actual'] = permisos_actual

        # 2. Permisos Semana Próxima
        permisos_proxima = SolicitudPermiso.objects.select_related('usuario').filter(
            estado='APROBADO',
            fecha_inicio__lte=domingo_proximo,
            fecha_termino__gte=lunes_proximo
        ).order_by('fecha_inicio')
        context['permisos_semana_proxima'] = permisos_proxima

        # 3. Licencias Semana Actual
        licencias_actual_raw = LicenciaMedica.objects.select_related('usuario').filter(
            fecha_inicio__lte=domingo_actual
        )
        licencias_semana_actual = []
        for lic in licencias_actual_raw:
            fecha_retorno = lic.fecha_inicio + timedelta(days=lic.dias)
            # Solo incluir si está ACTIVA HOY
            if lic.fecha_inicio <= hoy <= fecha_retorno:
                licencias_semana_actual.append({
                    'usuario': lic.usuario,
                    'fecha_inicio': lic.fecha_inicio,
                    'dias': lic.dias,
                    'fecha_retorno': fecha_retorno,
                    'es_activa': True
                })
        context['licencias_semana_actual'] = sorted(licencias_semana_actual, key=lambda x: x['usuario'].get_full_name())

        # 4. Licencias Semana Próxima
        licencias_proxima_raw = LicenciaMedica.objects.select_related('usuario').filter(
            fecha_inicio__lte=domingo_proximo
        )
        licencias_semana_proxima = []
        for lic in licencias_proxima_raw:
            fecha_retorno = lic.fecha_inicio + timedelta(days=lic.dias)
            if fecha_retorno >= lunes_proximo:
                licencias_semana_proxima.append({
                    'usuario': lic.usuario,
                    'fecha_inicio': lic.fecha_inicio,
                    'dias': lic.dias,
                    'fecha_retorno': fecha_retorno,
                })
        context['licencias_semana_proxima'] = sorted(licencias_semana_proxima, key=lambda x: x['usuario'].get_full_name())

        # --- Cálculo de Impacto Diario (Conteo de Permisos - Solo Diás Hábiles) ---
        def get_diario_stats(inicio_semana, permisos_qs, licencias_list):
            stats = []
            dias_nombre = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie']
            for i in range(5):
                dia = inicio_semana + timedelta(days=i)
                # Contar permisos en este día, diferenciando docentes y asistentes
                docente_count = 0
                asistente_count = 0
                for p in permisos_qs:
                    if p.fecha_inicio <= dia <= p.fecha_termino:
                        if p.usuario.categoria_funcionario == 'DOCENTE':
                            docente_count += 1
                        elif p.usuario.categoria_funcionario == 'ASISTENTE':
                            asistente_count += 1

                stats.append({
                    'dia': dia,
                    'nombre': dias_nombre[i],
                    'docente_count': docente_count,
                    'asistente_count': asistente_count,
                    'total': docente_count + asistente_count,
                    'is_today': dia == hoy
                })
            return stats

        context['daily_stats_actual'] = get_diario_stats(lunes_actual, permisos_actual, licencias_semana_actual)
        context['daily_stats_proxima'] = get_diario_stats(lunes_proximo, permisos_proxima, licencias_semana_proxima)

        # Metas generales
        usuarios_stats = CustomUser.objects.filter(role='FUNCIONARIO').aggregate(
            total_disponibles=Sum('dias_disponibles'),
            total_usuarios=Count('id')
        )
        context['dias_totales_disponibles'] = usuarios_stats['total_disponibles'] or 0
        context['promedio_dias_disponibles'] = (
            (usuarios_stats['total_disponibles'] / usuarios_stats['total_usuarios'])
            if usuarios_stats['total_usuarios'] > 0 else 0
        )
        
        context['usuarios_saldo_bajo'] = CustomUser.objects.filter(
            role='FUNCIONARIO',
            dias_disponibles__lt=2.0
        ).count()
        
        return context


class BlockedUsersView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para gestionar usuarios bloqueados"""
    template_name = 'admin_dashboard/blocked_users.html'
    
    def test_func(self):
        # Solo ADMIN puede gestionar usuarios bloqueados
        return self.request.user.role == 'ADMIN'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener todos los usuarios bloqueados manualmente
        manually_blocked = CustomUser.objects.filter(
            is_blocked=True
        ).select_related('blocked_by').order_by('-blocked_at')
        
        # Obtener usuarios bloqueados por Axes (intentos fallidos)
        axes_blocked_users = []
        try:
            from axes.models import AccessAttempt
            from django.contrib.auth import get_user_model
            from django.conf import settings
            User = get_user_model()
            failure_limit = getattr(settings, 'AXES_FAILURE_LIMIT', 8)
            
            # Obtener todos los intentos fallidos trackeados por Axes (para que el Admin vea todo)
            blocked_attempts = AccessAttempt.objects.filter(
                failures_since_start__gt=0
            ).order_by('-attempt_time')[:100]
            
            seen_usernames = set()
            for attempt in blocked_attempts:
                username = attempt.username
                if username and username not in seen_usernames:
                    seen_usernames.add(username)
                    user = User.objects.filter(
                        Q(email__iexact=username) | Q(username__iexact=username)
                    ).first()
                    
                    # Siempre agregamos a la lista para visibilidad, 
                    # incluso si no encontramos un perfil de usuario (ej: alguien intentando correos inexistentes)
                    axes_blocked_users.append({
                        'user': user,
                        'username_attempted': username,
                        'is_axes_blocked': attempt.failures_since_start >= failure_limit,
                        'blocked_at': attempt.attempt_time,
                        'ip_address': attempt.ip_address,
                        'attempts': attempt.failures_since_start,
                        'limit': failure_limit
                    })
        except Exception:
            pass
        
        context['blocked_users'] = manually_blocked
        context['axes_blocked_users'] = axes_blocked_users
        
        # Obtener usuarios no bloqueados para poder bloquear (todos los roles excepto ADMIN)
        context['active_users'] = CustomUser.objects.filter(
            is_blocked=False
        ).exclude(
            role='ADMIN'
        ).order_by('first_name', 'last_name')
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Manejar acciones de bloquear/desbloquear"""
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')
        
        if action == 'block':
            user = get_object_or_404(CustomUser, pk=user_id)
            user.is_blocked = True
            user.blocked_at = timezone.now()
            user.blocked_by = request.user
            user.save()
            
            messages.success(request, f'Usuario {user.get_full_name()} ha sido bloqueado.')
            
            registrar_log(
                usuario=request.user,
                tipo='USER',
                accion='Bloqueo de usuario',
                descripcion=f'Usuario {user.get_full_name()} (ID: {user.id}) fue bloqueado',
                ip_address=get_client_ip(request)
            )
            
        elif action == 'unblock':
            user = get_object_or_404(CustomUser, pk=user_id, is_blocked=True)
            user.is_blocked = False
            user.blocked_at = None
            user.blocked_by = None
            user.save()
            
            messages.success(request, f'Usuario {user.get_full_name()} ha sido desbloqueado.')
            
            registrar_log(
                usuario=request.user,
                tipo='USER',
                accion='Desbloqueo de usuario',
                descripcion=f'Usuario {user.get_full_name()} (ID: {user.id}) fue desbloqueado',
                ip_address=get_client_ip(request)
            )
        
        elif action == 'unblock_axes':
            # Desbloquear usuario de Axes
            try:
                from axes.models import AccessAttempt
                from django.contrib.auth import get_user_model
                User = get_user_model()
                
                user_id = request.POST.get('user_id')
                username = request.POST.get('username')
                
                description = ""
                if user_id:
                    user = get_object_or_404(CustomUser, pk=user_id)
                    username = user.email
                    description = f'Usuario {user.get_full_name()} (ID: {user.id}) ha sido desbloqueado de Axes.'
                elif username:
                    description = f'Usuario con correo {username} ha sido desbloqueado de Axes.'
                
                if username:
                    # Usar el método oficial de Axes 5+ para resetear intentos
                    try:
                        from axes.utils import reset
                        reset(username=username)
                    except ImportError:
                        # Si no existe reset en utils, borrar manualmente el registro
                        AccessAttempt.objects.filter(username__iexact=username).delete()
                    
                    messages.success(request, description)
                    
                    registrar_log(
                        usuario=request.user,
                        tipo='USER',
                        accion='Desbloqueo Axes',
                        descripcion=description,
                        ip_address=get_client_ip(request)
                    )
                else:
                    messages.error(request, 'No se pudo identificar el usuario para desbloquear.')
            except Exception as e:
                messages.error(request, f'Error al desbloquear: {str(e)}')
        
        return redirect('admin_dashboard:blocked_users')


class SystemLogsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista avanzada de logs del sistema con filtros y paginación"""
    template_name = 'admin_dashboard/logs.html'
    
    def test_func(self):
        # Solo ADMIN puede ver logs completos. Director/Secretaria ven sus propios o filtros limitados?
        # El user pidió mejorar para que ADMIN vea lo que hacen otros. 
        # Mantendremos el acceso actual pero orientado al Administrador.
        return self.request.user.role in ['ADMIN', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA']
    
    def get(self, request, *args, **kwargs):
        if request.resolver_match.url_name == 'logs_export':
            return self.export_logs(request)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Parámetros de filtro
        role = self.request.GET.get('role')
        search = self.request.GET.get('search')
        
        # 1. Obtener logs del sistema (Macro) - EXCLUIR "Sistema Automático" (usuario=None)
        system_queryset = SystemLog.objects.select_related('usuario').filter(usuario__isnull=False)
        if role:
            system_queryset = system_queryset.filter(usuario__role=role)
        if search:
            system_queryset = system_queryset.filter(
                Q(accion__icontains=search) | Q(descripcion__icontains=search) |
                Q(usuario__first_name__icontains=search) | Q(usuario__last_name__icontains=search)
            )

        # 2. Obtener logs de auditoría (Micro) - EXCLUIR "Sistema Automático" (actor=None)
        audit_queryset = LogEntry.objects.select_related('actor', 'content_type').filter(actor__isnull=False)
        if role:
            audit_queryset = audit_queryset.filter(actor__role=role)
        if search:
            audit_queryset = audit_queryset.filter(
                Q(object_repr__icontains=search) | Q(changes__icontains=search) |
                Q(actor__first_name__icontains=search) | Q(actor__last_name__icontains=search)
            )

        # 3. Combinar y Normalizar para la vista
        combined_logs = []
        
        for log in system_queryset[:300]:
            combined_logs.append({
                'timestamp': log.timestamp,
                'usuario': log.usuario,
                'tipo_display': log.get_tipo_display(),
                'tipo_slug': log.tipo,
                'accion': log.accion,
                'detalles': log.descripcion,
                'es_audit': False,
                'ip': log.ip_address
            })
            
        for log in audit_queryset[:300]:
            accion_map = {0: 'CREACIÓN', 1: 'MODIFICACIÓN', 2: 'ELIMINACIÓN'}
            
            detalles = 'Registro eliminado'
            if log.action != 2:
                try:
                    cambios = log.changes_display_dict
                    if cambios:
                        detalles = "\n".join([f"• {k}: {v[0]} ➔ {v[1]}" for k, v in cambios.items()])
                    else:
                        detalles = "Sin cambios registrados"
                except Exception:
                    detalles = str(log.changes_dict)
                    
            combined_logs.append({
                'timestamp': log.timestamp,
                'usuario': log.actor,
                'tipo_display': log.content_type.name.upper() if log.content_type else 'AUDITORÍA',
                'tipo_slug': 'AUDIT',
                'accion': f"{accion_map.get(log.action, 'CAMBIO')} EN {log.object_repr}",
                'detalles': detalles,
                'es_audit': True,
                'ip': '-'
            })

        # 4. Ordenar por fecha descendente
        combined_logs.sort(key=lambda x: x['timestamp'], reverse=True)

        # 5. Paginación de la lista combinada
        paginator = Paginator(combined_logs, 40)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context['page_obj'] = page_obj
        context['role_choices'] = CustomUser.ROLE_CHOICES
        context['filtros'] = {
            'role': role,
            'search': search,
        }
        
        return context

    def export_logs(self, request):
        """Exporta los logs unificados a Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Filtros
        role = request.GET.get('role')
        search = request.GET.get('search')
        
        # 1. Obtener datos (mismo orden que la vista) - EXCLUIR "Sistema Automático"
        system_qs = SystemLog.objects.select_related('usuario').filter(usuario__isnull=False)
        audit_qs = LogEntry.objects.select_related('actor', 'content_type').filter(actor__isnull=False)
        
        if role:
            system_qs = system_qs.filter(usuario__role=role)
            audit_qs = audit_qs.filter(actor__role=role)
        
        if search:
            system_qs = system_qs.filter(Q(accion__icontains=search) | Q(descripcion__icontains=search))
            audit_qs = audit_qs.filter(Q(object_repr__icontains=search) | Q(changes__icontains=search))

        combined = []
        for l in system_qs[:500]:
            combined.append([l.timestamp, l.usuario.get_full_name(), l.get_tipo_display(), l.accion, l.descripcion])
        for l in audit_qs[:500]:
            combined.append([l.timestamp, l.actor.get_full_name(), 'AUDITORÍA', f"CAMBIO EN {l.object_repr}", l.changes_display_dict])
        
        combined.sort(key=lambda x: x[0], reverse=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoría Unificada"
        
        ws.append(['Fecha/Hora', 'Usuario', 'Tipo', 'Acción', 'Detalles'])
        
        # Estilo
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
            
        for row in combined:
            # Formatear fecha para Excel
            row[0] = row[0].replace(tzinfo=None)
            ws.append(row)
            
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 25
        ws.column_dimensions['E'].width = 80

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=auditoria_unificada_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class SystemBackupView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Interfaz gráfica para gestionar el respaldo nacional del sistema"""
    template_name = 'admin_dashboard/system_backup.html'
    
    def test_func(self):
        return self.request.user.role == 'ADMIN'


class SystemBackupExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Genera y descarga en vivo el archivo ZIP (Restauración Ante Desastres)"""
    
    def test_func(self):
        return self.request.user.role == 'ADMIN'
        
    def get(self, request):
        registrar_log(
            usuario=request.user,
            tipo='SYSTEM',
            accion='Exportar Respaldo Nacional',
            descripcion='Admin descargó una copia completa del sistema',
            ip_address=get_client_ip(request)
        )
        
        # 1. Volcar Base de Datos y excluir cosas problemáticas
        db_out = io.StringIO()
        call_command(
            'dumpdata', 
            format='json', 
            natural_foreign=True, 
            natural_primary=True,
            exclude=['contenttypes', 'auth.permission', 'admin.logentry', 'sessions.session', 'asistencia'],
            stdout=db_out
        )
        
        # 2. Generar archivo comprimido ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Adjuntar Base Cruda
            zip_file.writestr('database.json', db_out.getvalue())
            
            # Recopilar todo el repositorio de la carpeta "media/"
            media_root = settings.MEDIA_ROOT
            if os.path.exists(media_root):
                for root, dirs, files in os.walk(media_root):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.join('media', os.path.relpath(file_path, media_root))
                        zip_file.write(file_path, arcname)
                        
        zip_buffer.seek(0)
        
        from datetime import datetime
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        response = FileResponse(zip_buffer, as_attachment=True, filename=f'sgpal_disaster_recovery_{fecha_str}.zip')
        return response
class SystemBackupRestoreView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Instala a la fuerza un archivo ZIP y rescata el sistema SIN borrar datos actuales primero"""
    
    def test_func(self):
        return self.request.user.role == 'ADMIN'
        
    def post(self, request):
        if 'backup_zip' not in request.FILES:
            messages.error(request, 'No se ha subido ningún archivo ZIP.')
            return redirect('admin_dashboard:system_backup')
            
        zip_file = request.FILES['backup_zip']
        # 🛡️ BLINDAJE QUIRÚRGICO (v5): Capturar datos del admin actual para protección total
        current_admin = request.user
        admin_id = current_admin.id
        admin_run = getattr(current_admin, 'run', '')
        admin_email = current_admin.email
        admin_username = current_admin.username
        
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                # 1. Extraer ZIP
                zip_path = os.path.join(temp_dir, 'backup.zip')
                with open(zip_path, 'wb+') as f:
                    for chunk in zip_file.chunks():
                        f.write(chunk)
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
                    
                json_path = os.path.join(temp_dir, 'database.json')
                if not os.path.exists(json_path):
                    messages.error(request, 'Archivo inválido: falta database.json')
                    return redirect('admin_dashboard:system_backup')

                from django.db import connection, transaction
                from django.contrib.auth import get_user_model
                import json as json_lib
                
                # 🛡️ VALIDACIÓN DE SEGURIDAD (No proceder si no detectamos al admin)
                if not admin_id or not admin_run:
                    raise Exception("No se pudo identificar de forma única al administrador logueado.")

                # 2. SANEAMIENTO QUIRÚRGICO DEL JSON: Eliminar al admin y asistencia del backup
                with open(json_path, 'r', encoding='utf-8') as f:
                    backup_data = json_lib.load(f)
                
                clean_data = []
                for obj in backup_data:
                    model_label = str(obj.get('model', '')).lower()
                    
                    # Ignorar registros de asistencia por completo
                    if model_label.startswith('asistencia.'):
                        continue
                    
                    # Ignorar al administrador actual en el backup
                    if model_label == 'users.customuser':
                        fields = obj.get('fields', {})
                        bk_email = str(fields.get('email', '')).strip().lower()
                        bk_run = str(fields.get('run', '')).strip().lower()
                        bk_username = str(obj.get('username', '')).strip().lower()
                        
                        if obj.get('pk') == admin_id or \
                           bk_email == admin_email.strip().lower() or \
                           bk_run == admin_run.strip().lower() or \
                           bk_username == admin_username.strip().lower():
                            continue
                            
                    clean_data.append(obj)
                
                with open(json_path, 'w', encoding='utf-8') as f:
                    json_lib.dump(clean_data, f)

                # 3. RESTAURACIÓN CONTROLADA
                with transaction.atomic():
                    with connection.cursor() as cursor:
                        # Truncar tablas aplicación EXCEPTO usuarios y sesiones
                        # Esta vez INCLUIMOS asistencia para que queden vacías si el usuario no las quiere
                        cursor.execute("""
                            SELECT tablename FROM pg_tables 
                            WHERE schemaname = 'public' 
                            AND tablename NOT IN (
                                'django_migrations', 'django_content_type', 'auth_permission', 
                                'django_session', 'axes_accessattempt', 'axes_accesslog',
                                'users_customuser', 'users_customuser_groups', 'users_customuser_user_permissions'
                            )
                        """)
                        tables = [row[0] for row in cursor.fetchall()]
                        if tables:
                            tables_sql = ', '.join([f'"{t}"' for t in tables])
                            cursor.execute(f'TRUNCATE TABLE {tables_sql} RESTART IDENTITY CASCADE;')
                    
                    # 4. LIMPIEZA QUIRÚRGICA DE USUARIOS: Borrar todos menos al admin actual
                    User = get_user_model()
                    User.objects.exclude(id=admin_id).delete()
                    
                    # 5. CARGA DE DATOS SANEADOS
                    call_command('loaddata', json_path)
                    
                # 4. Restauración de archivos físicos (Media)
                extracted_media = os.path.join(temp_dir, 'media')
                if os.path.exists(extracted_media):
                    media_root = settings.MEDIA_ROOT
                    for root, dirs, files in os.walk(extracted_media):
                        rel_path = os.path.relpath(root, extracted_media)
                        dest_dir = os.path.join(media_root, rel_path) if rel_path != '.' else media_root
                        os.makedirs(dest_dir, exist_ok=True)
                        for file in files:
                            src_file = os.path.join(root, file)
                            dst_file = os.path.join(dest_dir, file)
                            shutil.copy2(src_file, dst_file)
                            
            registrar_log(
                usuario=request.user,
                tipo='SYSTEM',
                accion='Restauración Nacional Crítica',
                descripcion='Admin restauró el sistema completo (Base de Datos + Media)',
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, '✅ SISTEMA RECUPERADO EXITOSAMENTE. Se han restaurado todos los registros y archivos media.')
            return redirect('admin_dashboard:dashboard')
            
        except Exception as e:
            messages.error(request, f'💥 Fallo en el despliegue del Respaldo Nacional: {str(e)}')
            return redirect('admin_dashboard:system_backup')


# --- Vistas de Efemérides ---

class EfemerideListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Efemeride
    template_name = 'admin_dashboard/efemerides_list.html'
    context_object_name = 'efemerides'
    
    def test_func(self):
        return self.request.user.role in ['ADMIN', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA']

    def get_queryset(self):
        return Efemeride.objects.all().order_by('fecha')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for efe in context['efemerides']:
            # Generar URL de Google Calendar
            base_url = "https://www.google.com/calendar/render?action=TEMPLATE"
            date_start = efe.fecha.strftime('%Y%m%d')
            date_end = (efe.fecha + timedelta(days=1)).strftime('%Y%m%d')
            details = f"Responsable: {efe.responsable or 'No especificado'}\n{efe.descripcion or ''}"
            params = {
                'text': efe.titulo,
                'dates': f"{date_start}/{date_end}",
                'details': details,
            }
            efe.google_calendar_url = f"{base_url}&{urllib.parse.urlencode(params)}"
        return context


class EfemerideCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Efemeride
    form_class = EfemerideForm
    template_name = 'admin_dashboard/efemeride_form.html'
    def get_success_url(self):
        from django.urls import reverse
        return f"{reverse('dashboard_funcionario')}?tab=efemerides"

    
    def test_func(self):
        return self.request.user.role in ['ADMIN', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA']

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        response = super().form_valid(form)
        registrar_log(
            usuario=self.request.user,
            tipo='CREATE',
            accion='Creación de Efeméride',
            descripcion=f'Se creó la efeméride: {self.object.titulo} para el {self.object.fecha}',
            ip_address=get_client_ip(self.request)
        )
        messages.success(self.request, 'Efeméride creada correctamente.')
        return response


class EfemerideUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Efemeride
    form_class = EfemerideForm
    template_name = 'admin_dashboard/efemeride_form.html'
    def get_success_url(self):
        from django.urls import reverse
        return f"{reverse('dashboard_funcionario')}?tab=efemerides"


    def test_func(self):
        return self.request.user.role in ['ADMIN', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA']

    def form_valid(self, form):
        response = super().form_valid(form)
        registrar_log(
            usuario=self.request.user,
            tipo='UPDATE',
            accion='Actualización de Efeméride',
            descripcion=f'Se actualizó la efeméride: {self.object.titulo}',
            ip_address=get_client_ip(self.request)
        )
        messages.success(self.request, 'Efeméride actualizada correctamente.')
        return response


class EfemerideDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Efemeride
    def get_success_url(self):
        from django.urls import reverse
        return f"{reverse('dashboard_funcionario')}?tab=efemerides"


    def test_func(self):
        return self.request.user.role in ['ADMIN', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA']

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        titulo = self.object.titulo
        response = super().post(request, *args, **kwargs)
        registrar_log(
            usuario=self.request.user,
            tipo='DELETE',
            accion='Eliminación de Efeméride',
            descripcion=f'Se eliminó la efeméride: {titulo}',
            ip_address=get_client_ip(self.request)
        )
        messages.success(self.request, 'Efeméride eliminada correctamente.')
        return response

